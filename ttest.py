import datetime
import random
from datetime import datetime as dt
from tkinter import *
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill, Protection
from openpyxl.utils import get_column_letter
import os
import openpyxl
from tkinter import messagebox
from openpyxl.drawing.image import Image
import pandas as pd
from matplotlib import pyplot as plt
from matplotlib import ticker
import asyncio

dict = {"开工日期": "2022-12-12"}
numtest1 = 1

# 打开工作簿
workbook1 = openpyxl.load_workbook('./测量值/电梯竣工资料信息表.xlsx')
# 选择工作表
worksheet = workbook1['基本信息']
# 获取某个单元格的值
dict["工程名称"] = worksheet['B2'].value if worksheet['B2'].value is not None else ""
dict["工程地址"] = worksheet['B3'].value if worksheet['B3'].value is not None else ""
dict["施工许可证号"] = worksheet['B4'].value if worksheet['B4'].value is not None else ""
dict["建设单位"] = worksheet['B5'].value if worksheet['B5'].value is not None else ""
dict["建设单位项目负责人"] = worksheet['D5'].value if worksheet['D5'].value is not None else ""
dict["设计单位"] = worksheet['B6'].value if worksheet['B6'].value is not None else ""
dict["设计单位项目负责人"] = worksheet['D6'].value if worksheet['D6'].value is not None else ""
dict["监理单位"] = worksheet['B7'].value if worksheet['B7'].value is not None else ""
dict["总监理工程师"] = worksheet['D7'].value if worksheet['D7'].value is not None else ""
dict["总包单位"] = worksheet['B8'].value if worksheet['B8'].value is not None else ""
dict["总包单位项目负责人"] = worksheet['D8'].value if worksheet['D8'].value is not None else ""
dict["总包单位技术负责人"] = worksheet['F8'].value if worksheet['F8'].value is not None else ""
dict["分包单位"] = worksheet['B9'].value if worksheet['B9'].value is not None else ""
dict["分包单位项目负责人"] = worksheet['D9'].value if worksheet['D9'].value is not None else ""
dict["分包单位技术负责人"] = worksheet['F9'].value if worksheet['F9'].value is not None else ""
dict["供货商"] = worksheet['B10'].value if worksheet['B10'].value is not None else ""
dict["供应商专业负责人"] = worksheet['D10'].value if worksheet['D10'].value is not None else ""
dict["子分部工程名称"] = worksheet['B11'].value if worksheet['B11'].value is not None else ""
dict["分项工程名称"] = worksheet['B12'].value if worksheet['B12'].value is not None else ""
dict["分包内容"] = worksheet['B13'].value if worksheet['B13'].value is not None else ""
dict["建筑名称楼号"] = dict['工程名称']
dict["电梯数量"] = worksheet['B15'].value if worksheet['B15'].value is not None else ""
dict["分项工程数量"] = worksheet['B16'].value if worksheet['B16'].value is not None else ""
dict["子分部工程数量"] = worksheet['B17'].value if worksheet['B17'].value is not None else ""
dict["编制人"] = worksheet['B20'].value if worksheet['B20'].value is not None else ""
dict["审核人"] = worksheet['B21'].value if worksheet['B21'].value is not None else ""
dict["册数"] = worksheet['B22'].value if worksheet['B22'].value is not None else ""
date_value = worksheet['B24'].value
# 检查日期字段是否为None，并进行格式化处理
if date_value is not None:
    # 将日期格式化为字符串，例如"2023-01-13"
    date_str = date_value.strftime("%Y-%m-%d")
    # 将格式化后的日期字符串存储到字典中
    dict["日期"] = date_str
else:
    dict["日期"] = ""
dict["技术交底分项工程名称"] = worksheet['B28'].value if worksheet['B28'].value is not None else ""

date_value = worksheet['B29'].value
# 检查日期字段是否为None，并进行格式化处理
if date_value is not None:
    # 将日期格式化为字符串，例如"2023-01-13"
    date_str = date_value.strftime("%Y-%m-%d")
    # 将格式化后的日期字符串存储到字典中
    dict["交底日期"] = date_str
else:
    dict["交底日期"] = ""

date_value = worksheet['B30'].value
# 检查日期字段是否为None，并进行格式化处理
if date_value is not None:
    # 将日期格式化为字符串，例如"2023-01-13"
    date_str = date_value.strftime("%Y-%m-%d")
    # 将格式化后的日期字符串存储到字典中
    dict["开工日期"] = date_str
else:
    dict["开工日期"] = ""

date_value = worksheet['B31'].value
# 检查日期字段是否为None，并进行格式化处理
if date_value is not None:
    # 将日期格式化为字符串，例如"2023-01-13"
    date_str = date_value.strftime("%Y-%m-%d")
    # 将格式化后的日期字符串存储到字典中
    dict["竣工日期"] = date_str
else:
    dict["竣工日期"] = ""

# 打开工作簿
workbook1 = openpyxl.load_workbook('./测量值/电梯竣工资料信息表.xlsx')
# 选择工作表
worksheet = workbook1['电梯配置信息']
date_value = worksheet['B44'].value
# 检查日期字段是否为None，并进行格式化处理
if date_value is not None:
    # 将日期格式化为字符串，例如"2023-01-13"
    date_str = date_value.strftime("%Y-%m-%d")
    # 将格式化后的日期字符串存储到字典中
    dict["会审日期"] = date_str
else:
    dict["会审日期"] = ""
date_value = worksheet['B43'].value
# 检查日期字段是否为None，并进行格式化处理
if date_value is not None:
    # 将日期格式化为字符串，例如"2023-01-13"
    date_str = date_value.strftime("%Y-%m-%d")
    # 将格式化后的日期字符串存储到字典中
    dict["调查日期"] = date_str
else:
    dict["调查日期"] = ""


class InformationCompletion(Frame):
    def __init__(self, master=None):
        super().__init__(master)
        self.master.title("第一页")
        self.master = master
        self.pack()
        self.current_page = 0
        self.create_widgets()

    def create_widgets(self):

        # Labels and input boxes
        self.labels = ["开工日期", "工程名称", "建设单位", "监理单位", "总包单位",
                       "分包单位", "分包单位技术负责人", "编制人", "竣工日期"]
        self.entries = []
        for i in range(len(self.labels)):
            label = Label(self, text=self.labels[i], font=('Arial', 14))
            label.grid(row=i, column=0, padx=5, pady=5, sticky='w')
            entry = Entry(self, font=('Arial', 14), width=30)
            if self.labels[i] in dict:
                entry.insert(0, dict[self.labels[i]])
            entry.grid(row=i, column=1, padx=5, pady=5)
            self.entries.append(entry)

        # Next button
        self.next_button = Button(self, text="下一页", font=('Arial', 14), command=self.create_widgets3)
        self.next_button.grid(row=len(self.labels) + 1, column=0, columnspan=2, pady=5)

    def create_widgets3(self):
        for i in range(len(self.labels)):
            dict[self.labels[i]] = self.entries[i].get()

        self.pack_forget()  # 移除所有组件并释放内存
        self.destroy()
        # self.master.title("第二页")  # 更改窗口标题
        InformationCompletion2(self.master)  # 创建下一页

    def getdata(self, labels):
        for i in labels:
            if i in dict.keys():
                print()


class InformationCompletion2(Frame):
    def __init__(self, master=None):
        super().__init__(master)
        self.master.title("第二页")
        self.master = master
        self.pack()
        self.current_page = 0
        self.create_widgets()

    def create_widgets(self):
        # Labels and input boxes
        self.labels = ["工程名称", "分包单位", "建设单位", "监理单位", "开工日期"]
        self.entries = []
        for i in range(len(self.labels)):
            label = Label(self, text=self.labels[i], font=('Arial', 14))
            label.grid(row=i, column=0, padx=5, pady=5, sticky='w')
            entry = Entry(self, font=('Arial', 14), width=30)
            if self.labels[i] in dict:
                entry.insert(0, dict[self.labels[i]])
            entry.grid(row=i, column=1, padx=5, pady=5)
            self.entries.append(entry)

        # Next button
        self.next_button = Button(self, text="下一页", font=('Arial', 14), command=self.create_widgets3)
        self.next_button.grid(row=len(self.labels) + 1, column=0, columnspan=2, pady=5)

        self.agrens_button = Button(self, text="上一页", font=('Arial', 14), command=self.agrens_create)
        self.agrens_button.grid(row=len(self.labels) + 2, column=0, columnspan=2, pady=5)

    def agrens_create(self):
        self.destroy()
        InformationCompletion(self.master)

    def create_widgets3(self):
        for i in range(len(self.labels)):
            dict[self.labels[i]] = self.entries[i].get()

        self.destroy()
        InformationCompletion3(self.master)


class InformationCompletion3(Frame):
    def __init__(self, master=None):
        super().__init__(master)
        self.master.title("第三页")
        self.master = master
        self.pack()
        self.current_page = 0
        self.create_widgets()

    def create_widgets(self):

        # Labels and input boxes
        self.labels = ["开工日期", "工程名称", "施工许可证号", "建设单位",
                       "建设单位项目负责人", "设计单位", "设计单位项目负责人",
                       "监理单位", "总监理工程师", "分包单位", "分包单位项目负责人", "分包单位技术负责人"]
        self.entries = []
        for i in range(len(self.labels)):
            label = Label(self, text=self.labels[i], font=('Arial', 14))
            label.grid(row=i, column=0, padx=5, pady=5, sticky='w')
            entry = Entry(self, font=('Arial', 14), width=30)
            if self.labels[i] in dict:
                entry.insert(0, dict[self.labels[i]])
            entry.grid(row=i, column=1, padx=5, pady=5)
            self.entries.append(entry)

        # Next button
        self.next_button = Button(self, text="下一页", font=('Arial', 14), command=self.create_widgets3)
        self.next_button.grid(row=len(self.labels) + 1, column=0, columnspan=2, pady=5)

        self.agrens_button = Button(self, text="上一页", font=('Arial', 14), command=self.agrens_create)
        self.agrens_button.grid(row=len(self.labels) + 2, column=0, columnspan=2, pady=5)

    def agrens_create(self):
        self.destroy()
        InformationCompletion2(self.master)

    def create_widgets3(self):
        for i in range(len(self.labels)):
            dict[self.labels[i]] = self.entries[i].get()

        self.pack_forget()  # 隐藏第一页的部件
        self.destroy()
        InformationCompletion4(self.master)

    def getdata(self, labels):
        for i in labels:
            if i in dict.keys():
                print()


class InformationCompletion4(Frame):
    def __init__(self, master=None):
        super().__init__(master)
        self.master.title("第四页")
        self.master = master
        self.pack()
        self.current_page = 0
        self.create_widgets()

    def create_widgets(self):

        # Labels and input boxes
        self.labels = ["工程名称", ]
        self.entries = []
        for i in range(len(self.labels)):
            label = Label(self, text=self.labels[i], font=('Arial', 14))
            label.grid(row=i, column=0, padx=5, pady=5, sticky='w')
            entry = Entry(self, font=('Arial', 14), width=30)
            if self.labels[i] in dict:
                entry.insert(0, dict[self.labels[i]])
            entry.grid(row=i, column=1, padx=5, pady=5)
            self.entries.append(entry)

        # Next button
        self.next_button = Button(self, text="下一页", font=('Arial', 14), command=self.create_widgets3)
        self.next_button.grid(row=len(self.labels) + 1, column=0, columnspan=2, pady=5)

        self.agrens_button = Button(self, text="上一页", font=('Arial', 14), command=self.agrens_create)
        self.agrens_button.grid(row=len(self.labels) + 2, column=0, columnspan=2, pady=5)

    def agrens_create(self):
        self.destroy()
        InformationCompletion3(self.master)

    def create_widgets3(self):
        for i in range(len(self.labels)):
            dict[self.labels[i]] = self.entries[i].get()

        self.pack_forget()  # 隐藏第一页的部件
        self.destroy()
        InformationCompletion5(self.master)

    def getdata(self, labels):
        for i in labels:
            if i in dict.keys():
                print()


class InformationCompletion5(Frame):
    def __init__(self, master=None):
        super().__init__(master)
        self.master.title("第五页")
        self.master = master
        self.pack()
        self.current_page = 0
        self.create_widgets()

    def create_widgets(self):

        # Labels and input boxes
        self.labels = ["工程名称", "分包单位"]
        self.entries = []
        for i in range(len(self.labels)):
            label = Label(self, text=self.labels[i], font=('Arial', 14))
            label.grid(row=i, column=0, padx=5, pady=5, sticky='w')
            entry = Entry(self, font=('Arial', 14), width=30)
            if self.labels[i] in dict:
                entry.insert(0, dict[self.labels[i]])
            entry.grid(row=i, column=1, padx=5, pady=5)
            self.entries.append(entry)

        # Next button
        self.next_button = Button(self, text="下一页", font=('Arial', 14), command=self.create_widgets3)
        self.next_button.grid(row=len(self.labels) + 1, column=0, columnspan=2, pady=5)

        self.agrens_button = Button(self, text="上一页", font=('Arial', 14), command=self.agrens_create)
        self.agrens_button.grid(row=len(self.labels) + 2, column=0, columnspan=2, pady=5)

    def agrens_create(self):
        self.destroy()
        InformationCompletion4(self.master)

    def create_widgets3(self):
        for i in range(len(self.labels)):
            dict[self.labels[i]] = self.entries[i].get()

        self.pack_forget()  # 隐藏第一页的部件
        self.destroy()
        InformationCompletion6(self.master)

    def getdata(self, labels):
        for i in labels:
            if i in dict.keys():
                print()


class InformationCompletion6(Frame):
    def __init__(self, master=None):
        super().__init__(master)
        self.master.title("第六页")
        self.master = master
        self.pack()
        self.current_page = 0
        self.create_widgets()

    def create_widgets(self):

        # Labels and input boxes
        self.labels = ["开工日期", "工程名称", "分包单位", ]
        self.entries = []
        for i in range(len(self.labels)):
            label = Label(self, text=self.labels[i], font=('Arial', 14))
            label.grid(row=i, column=0, padx=5, pady=5, sticky='w')
            entry = Entry(self, font=('Arial', 14), width=30)
            if self.labels[i] in dict:
                entry.insert(0, dict[self.labels[i]])
            entry.grid(row=i, column=1, padx=5, pady=5)
            self.entries.append(entry)

        # Next button
        self.next_button = Button(self, text="下一页", font=('Arial', 14), command=self.create_widgets3)
        self.next_button.grid(row=len(self.labels) + 1, column=0, columnspan=2, pady=5)

        self.agrens_button = Button(self, text="上一页", font=('Arial', 14), command=self.agrens_create)
        self.agrens_button.grid(row=len(self.labels) + 2, column=0, columnspan=2, pady=5)

    def agrens_create(self):
        self.destroy()
        InformationCompletion5(self.master)

    def create_widgets3(self):
        for i in range(len(self.labels)):
            dict[self.labels[i]] = self.entries[i].get()

        self.pack_forget()  # 隐藏第一页的部件
        self.destroy()
        InformationCompletion7(self.master)

    def getdata(self, labels):
        for i in labels:
            if i in dict.keys():
                print()


class InformationCompletion7(Frame):
    def __init__(self, master=None):
        super().__init__(master)
        self.master.title("第七页")
        self.master = master
        self.pack()
        self.current_page = 0
        self.create_widgets()

    def create_widgets(self):

        # Labels and input boxes
        self.labels = ["工程名称", "监理单位", "分包单位"]
        self.entries = []
        for i in range(len(self.labels)):
            label = Label(self, text=self.labels[i], font=('Arial', 14))
            label.grid(row=i, column=0, padx=5, pady=5, sticky='w')
            entry = Entry(self, font=('Arial', 14), width=30)
            if self.labels[i] in dict:
                entry.insert(0, dict[self.labels[i]])
            entry.grid(row=i, column=1, padx=5, pady=5)
            self.entries.append(entry)

        # Save button

        # Next button
        self.next_button = Button(self, text="下一页", font=('Arial', 14), command=self.create_widgets3)
        self.next_button.grid(row=len(self.labels) + 1, column=0, columnspan=2, pady=5)

        self.agrens_button = Button(self, text="上一页", font=('Arial', 14), command=self.agrens_create)
        self.agrens_button.grid(row=len(self.labels) + 2, column=0, columnspan=2, pady=5)

    def agrens_create(self):
        self.destroy()
        InformationCompletion6(self.master)

    def create_widgets3(self):
        for i in range(len(self.labels)):
            dict[self.labels[i]] = self.entries[i].get()

        self.pack_forget()  # 隐藏第一页的部件
        self.destroy()
        InformationCompletion8(self.master)

    def getdata(self, labels):
        for i in labels:
            if i in dict.keys():
                print()


class InformationCompletion8(Frame):
    def __init__(self, master=None):
        super().__init__(master)
        self.master.title("第八页")
        self.master = master
        self.pack()
        self.current_page = 0
        self.create_widgets()

    def create_widgets(self):

        # Labels and input boxes
        self.labels = ["开工日期", "工程名称", "总包单位", "分包单位", "监理单位"]
        self.entries = []
        for i in range(len(self.labels)):
            label = Label(self, text=self.labels[i], font=('Arial', 14))
            label.grid(row=i, column=0, padx=5, pady=5, sticky='w')
            entry = Entry(self, font=('Arial', 14), width=30)
            if self.labels[i] in dict:
                entry.insert(0, dict[self.labels[i]])
            entry.grid(row=i, column=1, padx=5, pady=5)
            self.entries.append(entry)

        # Save button

        # Next button
        self.next_button = Button(self, text="下一页", font=('Arial', 14), command=self.create_widgets3)
        self.next_button.grid(row=len(self.labels) + 1, column=0, columnspan=2, pady=5)

        self.agrens_button = Button(self, text="上一页", font=('Arial', 14), command=self.agrens_create)
        self.agrens_button.grid(row=len(self.labels) + 2, column=0, columnspan=2, pady=5)

    def agrens_create(self):
        self.destroy()
        InformationCompletion7(self.master)

    def create_widgets3(self):
        for i in range(len(self.labels)):
            dict[self.labels[i]] = self.entries[i].get()

        self.pack_forget()  # 隐藏第一页的部件
        self.destroy()
        InformationCompletion9(self.master)

    def getdata(self, labels):
        for i in labels:
            if i in dict.keys():
                print()


class InformationCompletion9(Frame):
    def __init__(self, master=None):
        super().__init__(master)
        self.master.title("第九页")
        self.master = master
        self.pack()
        self.current_page = 0
        self.create_widgets()

    def create_widgets(self):

        # Labels and input boxes
        self.labels = ["调查日期", "工程名称", "调(勘)查地点", "调(勘)查时间起始时间",
                       "调(勘)查时间终止时间"]
        self.entries = []
        for i in range(len(self.labels)):
            label = Label(self, text=self.labels[i], font=('Arial', 14))
            label.grid(row=i, column=0, padx=5, pady=5, sticky='w')
            entry = Entry(self, font=('Arial', 14), width=30)
            if self.labels[i] in dict:
                entry.insert(0, dict[self.labels[i]])
            entry.grid(row=i, column=1, padx=5, pady=5)
            self.entries.append(entry)

        # Save button

        # Next button
        self.next_button = Button(self, text="下一页", font=('Arial', 14), command=self.create_widgets3)
        self.next_button.grid(row=len(self.labels) + 1, column=0, columnspan=2, pady=5)

        self.agrens_button = Button(self, text="上一页", font=('Arial', 14), command=self.agrens_create)
        self.agrens_button.grid(row=len(self.labels) + 2, column=0, columnspan=2, pady=5)

    def agrens_create(self):
        self.destroy()
        InformationCompletion8(self.master)

    def create_widgets3(self):
        for i in range(len(self.labels)):
            dict[self.labels[i]] = self.entries[i].get()

        self.pack_forget()  # 隐藏第一页的部件
        self.destroy()
        InformationCompletion10(self.master)

    def getdata(self, labels):
        for i in labels:
            if i in dict.keys():
                print()


class InformationCompletion10(Frame):
    def __init__(self, master=None):
        super().__init__(master)
        self.master.title("第十页")
        self.master = master
        self.pack()
        self.current_page = 0
        self.create_widgets()

    def create_widgets(self):

        # Labels and input boxes
        self.labels = ["工程名称", "分包单位", "建设单位",
                       "设计单位", "工程地址"]
        self.entries = []
        for i in range(len(self.labels)):
            label = Label(self, text=self.labels[i], font=('Arial', 14))
            label.grid(row=i, column=0, padx=5, pady=5, sticky='w')
            entry = Entry(self, font=('Arial', 14), width=30)
            if self.labels[i] in dict:
                entry.insert(0, dict[self.labels[i]])
            entry.grid(row=i, column=1, padx=5, pady=5)
            self.entries.append(entry)

        # Save button

        # Next button
        self.next_button = Button(self, text="下一页", font=('Arial', 14), command=self.create_widgets3)
        self.next_button.grid(row=len(self.labels) + 1, column=0, columnspan=2, pady=5)

        self.agrens_button = Button(self, text="上一页", font=('Arial', 14), command=self.agrens_create)
        self.agrens_button.grid(row=len(self.labels) + 2, column=0, columnspan=2, pady=5)

    def agrens_create(self):
        self.destroy()
        InformationCompletion9(self.master)

    def create_widgets3(self):
        for i in range(len(self.labels)):
            dict[self.labels[i]] = self.entries[i].get()

        self.pack_forget()  # 隐藏第一页的部件
        self.destroy()
        InformationCompletion11(self.master)

    def getdata(self, labels):
        for i in labels:
            if i in dict.keys():
                print()


class InformationCompletion11(Frame):
    def __init__(self, master=None):
        super().__init__(master)
        self.master.title("第十一页")
        self.master = master
        self.pack()
        self.current_page = 0
        self.create_widgets()

    def create_widgets(self):

        # Labels and input boxes
        self.labels = ["开工日期", "竣工日期", "工程名称", "施工内容"]
        self.entries = []
        for i in range(len(self.labels)):
            label = Label(self, text=self.labels[i], font=('Arial', 14))
            label.grid(row=i, column=0, padx=5, pady=5, sticky='w')
            entry = Entry(self, font=('Arial', 14), width=30)
            if self.labels[i] in dict:
                entry.insert(0, dict[self.labels[i]])
            entry.grid(row=i, column=1, padx=5, pady=5)
            self.entries.append(entry)

        # Save button

        # Next button
        self.next_button = Button(self, text="下一页", font=('Arial', 14), command=self.create_widgets3)
        self.next_button.grid(row=len(self.labels) + 1, column=0, columnspan=2, pady=5)

        self.agrens_button = Button(self, text="上一页", font=('Arial', 14), command=self.agrens_create)
        self.agrens_button.grid(row=len(self.labels) + 2, column=0, columnspan=2, pady=5)

    def agrens_create(self):
        self.destroy()
        InformationCompletion10(self.master)

    def create_widgets3(self):
        for i in range(len(self.labels)):
            dict[self.labels[i]] = self.entries[i].get()

        self.pack_forget()  # 隐藏第一页的部件
        self.destroy()
        InformationCompletion12(self.master)

    def getdata(self, labels):
        for i in labels:
            if i in dict.keys():
                print()


class InformationCompletion12(Frame):
    def __init__(self, master=None):
        super().__init__(master)
        self.master.title("第十二页")
        self.master = master
        self.pack()
        self.current_page = 0
        self.create_widgets()

    def create_widgets(self):

        # Labels and input boxes
        self.labels = ["工程名称",
                       "监理单位", "分包单位", "建设单位"]
        self.entries = []
        for i in range(len(self.labels)):
            label = Label(self, text=self.labels[i], font=('Arial', 14))
            label.grid(row=i, column=0, padx=5, pady=5, sticky='w')
            entry = Entry(self, font=('Arial', 14), width=30)
            if self.labels[i] in dict:
                entry.insert(0, dict[self.labels[i]])
            entry.grid(row=i, column=1, padx=5, pady=5)
            self.entries.append(entry)

        # Save button

        # Next button
        self.next_button = Button(self, text="下一页", font=('Arial', 14), command=self.create_widgets3)
        self.next_button.grid(row=len(self.labels) + 1, column=0, columnspan=2, pady=5)

        self.agrens_button = Button(self, text="上一页", font=('Arial', 14), command=self.agrens_create)
        self.agrens_button.grid(row=len(self.labels) + 2, column=0, columnspan=2, pady=5)

    def agrens_create(self):
        self.destroy()
        InformationCompletion11(self.master)

    def create_widgets3(self):
        for i in range(len(self.labels)):
            dict[self.labels[i]] = self.entries[i].get()

        self.pack_forget()  # 隐藏第一页的部件
        self.destroy()
        InformationCompletion13(self.master)

    def getdata(self, labels):
        for i in labels:
            if i in dict.keys():
                print()


class InformationCompletion13(Frame):
    def __init__(self, master=None):
        super().__init__(master)
        self.master.title("第十三页")
        self.master = master
        self.pack()
        self.current_page = 0
        self.create_widgets()

    def create_widgets(self):

        # Labels and input boxes
        self.labels = ["工程名称"]
        self.entries = []
        for i in range(len(self.labels)):
            label = Label(self, text=self.labels[i], font=('Arial', 14))
            label.grid(row=i, column=0, padx=5, pady=5, sticky='w')
            entry = Entry(self, font=('Arial', 14), width=30)
            if self.labels[i] in dict:
                entry.insert(0, dict[self.labels[i]])
            entry.grid(row=i, column=1, padx=5, pady=5)
            self.entries.append(entry)

        # Save button

        # Next button
        self.next_button = Button(self, text="下一页", font=('Arial', 14), command=self.create_widgets3)
        self.next_button.grid(row=len(self.labels) + 1, column=0, columnspan=2, pady=5)

        self.agrens_button = Button(self, text="上一页", font=('Arial', 14), command=self.agrens_create)
        self.agrens_button.grid(row=len(self.labels) + 2, column=0, columnspan=2, pady=5)

    def agrens_create(self):
        self.destroy()
        InformationCompletion12(self.master)

    def create_widgets3(self):
        for i in range(len(self.labels)):
            dict[self.labels[i]] = self.entries[i].get()

        self.pack_forget()  # 隐藏第一页的部件
        self.destroy()
        InformationCompletion14(self.master)

    def getdata(self, labels):
        for i in labels:
            if i in dict.keys():
                print()


class InformationCompletion14(Frame):
    def __init__(self, master=None):
        super().__init__(master)
        self.master.title("第十四页")
        self.master = master
        self.pack()
        self.current_page = 0
        self.create_widgets()

    def create_widgets(self):

        # Labels and input boxes
        self.labels = ["工程名称", "分包单位", "分项工程名称"]
        self.entries = []
        for i in range(len(self.labels)):
            label = Label(self, text=self.labels[i], font=('Arial', 14))
            label.grid(row=i, column=0, padx=5, pady=5, sticky='w')
            entry = Entry(self, font=('Arial', 14), width=30)
            if self.labels[i] in dict:
                entry.insert(0, dict[self.labels[i]])
            entry.grid(row=i, column=1, padx=5, pady=5)
            self.entries.append(entry)

        # Save button

        # Next button
        self.next_button = Button(self, text="下一页", font=('Arial', 14), command=self.create_widgets3)
        self.next_button.grid(row=len(self.labels) + 1, column=0, columnspan=2, pady=5)

        self.agrens_button = Button(self, text="上一页", font=('Arial', 14), command=self.agrens_create)
        self.agrens_button.grid(row=len(self.labels) + 2, column=0, columnspan=2, pady=5)

    def agrens_create(self):
        self.destroy()
        InformationCompletion13(self.master)

    def create_widgets3(self):
        for i in range(len(self.labels)):
            dict[self.labels[i]] = self.entries[i].get()

        self.pack_forget()  # 隐藏第一页的部件
        self.destroy()
        InformationCompletion15(self.master)

    def getdata(self, labels):
        for i in labels:
            if i in dict.keys():
                print()


class InformationCompletion15(Frame):
    def __init__(self, master=None):
        super().__init__(master)
        self.master.title("第十五页")
        self.master = master
        self.pack()
        self.current_page = 0
        self.create_widgets()

    def create_widgets(self):

        # Labels and input boxes
        self.labels = ["工程名称", "会审日期"]
        self.entries = []
        for i in range(len(self.labels)):
            label = Label(self, text=self.labels[i], font=('Arial', 14))
            label.grid(row=i, column=0, padx=5, pady=5, sticky='w')
            entry = Entry(self, font=('Arial', 14), width=30)
            if self.labels[i] in dict:
                entry.insert(0, dict[self.labels[i]])
            entry.grid(row=i, column=1, padx=5, pady=5)
            self.entries.append(entry)

        # Save button

        # Next button
        self.next_button = Button(self, text="下一页", font=('Arial', 14), command=self.create_widgets3)
        self.next_button.grid(row=len(self.labels) + 1, column=0, columnspan=2, pady=5)

        self.agrens_button = Button(self, text="上一页", font=('Arial', 14), command=self.agrens_create)
        self.agrens_button.grid(row=len(self.labels) + 2, column=0, columnspan=2, pady=5)

    def agrens_create(self):
        self.destroy()
        InformationCompletion14(self.master)

    def create_widgets3(self):
        for i in range(len(self.labels)):
            dict[self.labels[i]] = self.entries[i].get()

        self.pack_forget()  # 隐藏第一页的部件
        self.destroy()
        InformationCompletion16(self.master)

    def getdata(self, labels):
        for i in labels:
            if i in dict.keys():
                print()


class InformationCompletion16(Frame):
    def __init__(self, master=None):
        super().__init__(master)
        self.master.title("第十六页")
        self.master = master
        self.pack()
        self.current_page = 0
        self.create_widgets()

    def create_widgets(self):

        # Labels and input boxes
        self.labels = ["会审日期", "工程名称", "子分部工程名称",
                       "建设单位", "设计单位", "监理单位", "分包单位"]
        self.entries = []
        for i in range(len(self.labels)):
            label = Label(self, text=self.labels[i], font=('Arial', 14))
            label.grid(row=i, column=0, padx=5, pady=5, sticky='w')
            entry = Entry(self, font=('Arial', 14), width=30)
            if self.labels[i] in dict:
                entry.insert(0, dict[self.labels[i]])
            entry.grid(row=i, column=1, padx=5, pady=5)
            self.entries.append(entry)

        # Save button

        # Next button
        self.next_button = Button(self, text="下一页", font=('Arial', 14), command=self.create_widgets3)
        self.next_button.grid(row=len(self.labels) + 1, column=0, columnspan=2, pady=5)

        self.agrens_button = Button(self, text="上一页", font=('Arial', 14), command=self.agrens_create)
        self.agrens_button.grid(row=len(self.labels) + 2, column=0, columnspan=2, pady=5)

    def agrens_create(self):
        self.destroy()
        InformationCompletion15(self.master)

    def create_widgets3(self):
        for i in range(len(self.labels)):
            dict[self.labels[i]] = self.entries[i].get()

        self.pack_forget()  # 隐藏第一页的部件
        self.destroy()
        InformationCompletion17(self.master)

    def getdata(self, labels):
        for i in labels:
            if i in dict.keys():
                print()


class InformationCompletion17(Frame):
    def __init__(self, master=None):
        super().__init__(master)
        self.master.title("第十七页")
        self.master = master
        self.pack()
        self.current_page = 0
        self.create_widgets()

    def create_widgets(self):

        # Labels and input boxes
        self.labels = ["工程名称", "子分部工程名称", "交底人", "交底地点",
                       "建设单位", "设计单位", "监理单位", "分包单位"]
        self.entries = []
        for i in range(len(self.labels)):
            label = Label(self, text=self.labels[i], font=('Arial', 14))
            label.grid(row=i, column=0, padx=5, pady=5, sticky='w')
            entry = Entry(self, font=('Arial', 14), width=30)
            if self.labels[i] in dict:
                entry.insert(0, dict[self.labels[i]])
            entry.grid(row=i, column=1, padx=5, pady=5)
            self.entries.append(entry)

        # Save button

        # Next button
        self.next_button = Button(self, text="下一页", font=('Arial', 14), command=self.create_widgets3)
        self.next_button.grid(row=len(self.labels) + 1, column=0, columnspan=2, pady=5)

        self.agrens_button = Button(self, text="上一页", font=('Arial', 14), command=self.agrens_create)
        self.agrens_button.grid(row=len(self.labels) + 2, column=0, columnspan=2, pady=5)

    def agrens_create(self):
        self.destroy()
        InformationCompletion16(self.master)

    def create_widgets3(self):
        for i in range(len(self.labels)):
            dict[self.labels[i]] = self.entries[i].get()

        self.pack_forget()  # 隐藏第一页的部件
        self.destroy()
        InformationCompletion18(self.master)

    def getdata(self, labels):
        for i in labels:
            if i in dict.keys():
                print()


class InformationCompletion18(Frame):
    def __init__(self, master=None):
        super().__init__(master)
        self.master.title("第十八页")
        self.master = master
        self.pack()
        self.current_page = 0
        self.create_widgets()

    def create_widgets(self):

        # Labels and input boxes
        self.labels = ["变更日期", "工程名称", "子分部工程名称", "设计单位",
                       "建设单位", "监理单位", "设计单位",
                       "分包单位"]
        self.entries = []
        for i in range(len(self.labels)):
            label = Label(self, text=self.labels[i], font=('Arial', 14))
            label.grid(row=i, column=0, padx=5, pady=5, sticky='w')
            entry = Entry(self, font=('Arial', 14), width=30)
            if self.labels[i] in dict:
                entry.insert(0, dict[self.labels[i]])
            entry.grid(row=i, column=1, padx=5, pady=5)
            self.entries.append(entry)

        # Save button

        # Next button
        self.next_button = Button(self, text="下一页", font=('Arial', 14), command=self.create_widgets3)
        self.next_button.grid(row=len(self.labels) + 1, column=0, columnspan=2, pady=5)

        self.agrens_button = Button(self, text="上一页", font=('Arial', 14), command=self.agrens_create)
        self.agrens_button.grid(row=len(self.labels) + 2, column=0, columnspan=2, pady=5)

    def agrens_create(self):
        self.destroy()
        InformationCompletion17(self.master)

    def create_widgets3(self):
        for i in range(len(self.labels)):
            dict[self.labels[i]] = self.entries[i].get()

        self.pack_forget()  # 隐藏第一页的部件
        self.destroy()
        InformationCompletion19(self.master)

    def getdata(self, labels):
        for i in labels:
            if i in dict.keys():
                print()


class InformationCompletion19(Frame):
    def __init__(self, master=None):
        super().__init__(master)
        self.master.title("第十九页")
        self.master = master
        self.pack()
        self.current_page = 0
        self.create_widgets()

    def create_widgets(self):

        # Labels and input boxes
        self.labels = ["工程名称", "子分部工程名称", "建设单位",
                       "分包单位", "设计单位", "监理单位",
                       "设计单位"]
        self.entries = []
        for i in range(len(self.labels)):
            label = Label(self, text=self.labels[i], font=('Arial', 14))
            label.grid(row=i, column=0, padx=5, pady=5, sticky='w')
            entry = Entry(self, font=('Arial', 14), width=30)
            if self.labels[i] in dict:
                entry.insert(0, dict[self.labels[i]])
            entry.grid(row=i, column=1, padx=5, pady=5)
            self.entries.append(entry)

        # Save button

        # Next button
        self.next_button = Button(self, text="下一页", font=('Arial', 14), command=self.create_widgets3)
        self.next_button.grid(row=len(self.labels) + 1, column=0, columnspan=2, pady=5)

        self.agrens_button = Button(self, text="上一页", font=('Arial', 14), command=self.agrens_create)
        self.agrens_button.grid(row=len(self.labels) + 2, column=0, columnspan=2, pady=5)

    def agrens_create(self):
        self.destroy()
        InformationCompletion18(self.master)

    def create_widgets3(self):
        for i in range(len(self.labels)):
            dict[self.labels[i]] = self.entries[i].get()

        self.pack_forget()  # 隐藏第一页的部件
        self.destroy()
        InformationCompletion20(self.master)

    def getdata(self, labels):
        for i in labels:
            if i in dict.keys():
                print()


class InformationCompletion20(Frame):
    def __init__(self, master=None):
        super().__init__(master)
        self.master.title("第20页")
        self.master = master
        self.pack()
        self.current_page = 0
        self.create_widgets()

    def create_widgets(self):

        # Labels and input boxes
        self.labels = ["工程名称", "监理单位",
                       "设备报审日期"]
        self.entries = []
        for i in range(len(self.labels)):
            label = Label(self, text=self.labels[i], font=('Arial', 14))
            label.grid(row=i, column=0, padx=5, pady=5, sticky='w')
            entry = Entry(self, font=('Arial', 14), width=30)
            if self.labels[i] in dict:
                entry.insert(0, dict[self.labels[i]])
            entry.grid(row=i, column=1, padx=5, pady=5)
            self.entries.append(entry)

        # Save button

        # Next button
        self.next_button = Button(self, text="下一页", font=('Arial', 14), command=self.create_widgets3)
        self.next_button.grid(row=len(self.labels) + 1, column=0, columnspan=2, pady=5)

        self.agrens_button = Button(self, text="上一页", font=('Arial', 14), command=self.agrens_create)
        self.agrens_button.grid(row=len(self.labels) + 2, column=0, columnspan=2, pady=5)

    def agrens_create(self):
        self.destroy()
        InformationCompletion19(self.master)

    def create_widgets3(self):
        for i in range(len(self.labels)):
            dict[self.labels[i]] = self.entries[i].get()

        self.pack_forget()  # 隐藏第一页的部件
        self.destroy()
        InformationCompletion21(self.master)

    def getdata(self, labels):
        for i in labels:
            if i in dict.keys():
                print()


class InformationCompletion21(Frame):
    def __init__(self, master=None):
        super().__init__(master)
        self.master.title("第21页")
        self.master = master
        self.pack()
        self.current_page = 0
        self.create_widgets()

    def create_widgets(self):

        # Labels and input boxes
        self.labels = ["检验日期", "工程名称"]
        self.entries = []
        for i in range(len(self.labels)):
            label = Label(self, text=self.labels[i], font=('Arial', 14))
            label.grid(row=i, column=0, padx=5, pady=5, sticky='w')
            entry = Entry(self, font=('Arial', 14), width=30)
            if self.labels[i] in dict:
                entry.insert(0, dict[self.labels[i]])
            entry.grid(row=i, column=1, padx=5, pady=5)
            self.entries.append(entry)

        # Save button

        # Next button
        self.next_button = Button(self, text="下一页", font=('Arial', 14), command=self.create_widgets3)
        self.next_button.grid(row=len(self.labels) + 1, column=0, columnspan=2, pady=5)

        self.agrens_button = Button(self, text="上一页", font=('Arial', 14), command=self.agrens_create)
        self.agrens_button.grid(row=len(self.labels) + 2, column=0, columnspan=2, pady=5)

    def agrens_create(self):
        self.destroy()
        InformationCompletion20(self.master)

    def create_widgets3(self):
        for i in range(len(self.labels)):
            dict[self.labels[i]] = self.entries[i].get()

        self.pack_forget()  # 隐藏第一页的部件
        self.destroy()
        InformationCompletion23(self.master)

    def getdata(self, labels):
        for i in labels:
            if i in dict.keys():
                print()


class InformationCompletion23(Frame):
    def __init__(self, master=None):
        super().__init__(master)
        self.master.title("第23页")
        self.master = master
        self.pack()
        self.current_page = 0
        self.create_widgets()

    def create_widgets(self):

        # Labels and input boxes
        self.labels = ["工程名称"]
        self.entries = []
        for i in range(len(self.labels)):
            label = Label(self, text=self.labels[i], font=('Arial', 14))
            label.grid(row=i, column=0, padx=5, pady=5, sticky='w')
            entry = Entry(self, font=('Arial', 14), width=30)
            if self.labels[i] in dict:
                entry.insert(0, dict[self.labels[i]])
            entry.grid(row=i, column=1, padx=5, pady=5)
            self.entries.append(entry)

        # Save button

        # Next button
        self.next_button = Button(self, text="下一页", font=('Arial', 14), command=self.create_widgets3)
        self.next_button.grid(row=len(self.labels) + 1, column=0, columnspan=2, pady=5)

        self.agrens_button = Button(self, text="上一页", font=('Arial', 14), command=self.agrens_create)
        self.agrens_button.grid(row=len(self.labels) + 2, column=0, columnspan=2, pady=5)

    def agrens_create(self):
        self.destroy()
        InformationCompletion21(self.master)

    def create_widgets3(self):
        for i in range(len(self.labels)):
            dict[self.labels[i]] = self.entries[i].get()

        self.pack_forget()  # 隐藏第一页的部件
        self.destroy()
        InformationCompletion24(self.master)

    def getdata(self, labels):
        for i in labels:
            if i in dict.keys():
                print()


class InformationCompletion24(Frame):
    def __init__(self, master=None):
        super().__init__(master)
        self.master.title("第24页")
        self.master = master
        self.pack()
        self.current_page = 0
        self.create_widgets()

    def create_widgets(self):

        # Labels and input boxes
        self.labels = ["材料名称", "合格证（原件）编号", "工程总需要量", "进货数量",
                       "供货单位", "到货日期", "合格证原件存放单位"]
        self.entries = []
        for i in range(len(self.labels)):
            label = Label(self, text=self.labels[i], font=('Arial', 14))
            label.grid(row=i, column=0, padx=5, pady=5, sticky='w')
            entry = Entry(self, font=('Arial', 14), width=30)
            if self.labels[i] in dict:
                entry.insert(0, dict[self.labels[i]])
            entry.grid(row=i, column=1, padx=5, pady=5)
            self.entries.append(entry)

        # Save button

        # Next button
        self.next_button = Button(self, text="下一页", font=('Arial', 14), command=self.create_widgets3)
        self.next_button.grid(row=len(self.labels) + 1, column=0, columnspan=2, pady=5)

        self.agrens_button = Button(self, text="上一页", font=('Arial', 14), command=self.agrens_create)
        self.agrens_button.grid(row=len(self.labels) + 2, column=0, columnspan=2, pady=5)

    def agrens_create(self):
        self.destroy()
        InformationCompletion23(self.master)

    def create_widgets3(self):
        for i in range(len(self.labels)):
            dict[self.labels[i]] = self.entries[i].get()

        self.pack_forget()  # 隐藏第一页的部件
        self.destroy()
        InformationCompletion25(self.master)

    def getdata(self, labels):
        for i in labels:
            if i in dict.keys():
                print()


class InformationCompletion25(Frame):
    def __init__(self, master=None):
        super().__init__(master)
        self.master.title("第25页")
        self.master = master
        self.pack()
        self.current_page = 0
        self.create_widgets()

    def create_widgets(self):

        # Labels and input boxes
        self.labels = ["工程名称", "监理单位", "分包单位"]
        self.entries = []
        for i in range(len(self.labels)):
            label = Label(self, text=self.labels[i], font=('Arial', 14))
            label.grid(row=i, column=0, padx=5, pady=5, sticky='w')
            entry = Entry(self, font=('Arial', 14), width=30)
            if self.labels[i] in dict:
                entry.insert(0, dict[self.labels[i]])
            entry.grid(row=i, column=1, padx=5, pady=5)
            self.entries.append(entry)

        # Save button

        # Next button
        self.next_button = Button(self, text="下一页", font=('Arial', 14), command=self.create_widgets3)
        self.next_button.grid(row=len(self.labels) + 1, column=0, columnspan=2, pady=5)

        self.agrens_button = Button(self, text="上一页", font=('Arial', 14), command=self.agrens_create)
        self.agrens_button.grid(row=len(self.labels) + 2, column=0, columnspan=2, pady=5)

    def agrens_create(self):
        self.destroy()
        InformationCompletion24(self.master)

    def create_widgets3(self):
        for i in range(len(self.labels)):
            dict[self.labels[i]] = self.entries[i].get()

        self.pack_forget()  # 隐藏第一页的部件
        self.destroy()
        InformationCompletion26(self.master)

    def getdata(self, labels):
        for i in labels:
            if i in dict.keys():
                print()


class InformationCompletion26(Frame):
    def __init__(self, master=None):
        super().__init__(master)
        self.master.title("第26页")
        self.master = master
        self.pack()
        self.current_page = 0
        self.create_widgets()

    def create_widgets(self):

        # Labels and input boxes
        self.labels = ["工程名称", "取样部位", "样品名称", "取样数量",
                       "取样地点", "取样日期", "执行标准、规范",
                       "试验项目"]
        self.entries = []
        for i in range(len(self.labels)):
            label = Label(self, text=self.labels[i], font=('Arial', 14))
            label.grid(row=i, column=0, padx=5, pady=5, sticky='w')
            entry = Entry(self, font=('Arial', 14), width=30)
            if self.labels[i] in dict:
                entry.insert(0, dict[self.labels[i]])
            entry.grid(row=i, column=1, padx=5, pady=5)
            self.entries.append(entry)

        # Save button

        # Next button
        self.next_button = Button(self, text="下一页", font=('Arial', 14), command=self.create_widgets3)
        self.next_button.grid(row=len(self.labels) + 1, column=0, columnspan=2, pady=5)

        self.agrens_button = Button(self, text="上一页", font=('Arial', 14), command=self.agrens_create)
        self.agrens_button.grid(row=len(self.labels) + 2, column=0, columnspan=2, pady=5)

    def agrens_create(self):
        self.destroy()
        InformationCompletion25(self.master)

    def create_widgets3(self):
        for i in range(len(self.labels)):
            dict[self.labels[i]] = self.entries[i].get()

        self.pack_forget()  # 隐藏第一页的部件
        self.destroy()
        InformationCompletion28(self.master)

    def getdata(self, labels):
        for i in labels:
            if i in dict.keys():
                print()


class InformationCompletion28(Frame):
    def __init__(self, master=None):
        super().__init__(master)
        self.master.title("第28页")
        self.master = master
        self.pack()
        self.current_page = 0
        self.create_widgets()

    def create_widgets(self):

        # Labels and input boxes
        self.labels = ["工程名称", "检查项目", "检查部位", "检查日期", "分包单位"]
        self.entries = []
        for i in range(len(self.labels)):
            label = Label(self, text=self.labels[i], font=('Arial', 14))
            label.grid(row=i, column=0, padx=5, pady=5, sticky='w')
            entry = Entry(self, font=('Arial', 14), width=30)
            if self.labels[i] in dict:
                entry.insert(0, dict[self.labels[i]])
            entry.grid(row=i, column=1, padx=5, pady=5)
            self.entries.append(entry)

        # Save button

        # Next button
        self.next_button = Button(self, text="下一页", font=('Arial', 14), command=self.create_widgets3)
        self.next_button.grid(row=len(self.labels) + 1, column=0, columnspan=2, pady=5)

        self.agrens_button = Button(self, text="上一页", font=('Arial', 14), command=self.agrens_create)
        self.agrens_button.grid(row=len(self.labels) + 2, column=0, columnspan=2, pady=5)

    def agrens_create(self):
        self.destroy()
        InformationCompletion26(self.master)

    def create_widgets3(self):
        for i in range(len(self.labels)):
            dict[self.labels[i]] = self.entries[i].get()

        self.pack_forget()  # 隐藏第一页的部件
        self.destroy()
        InformationCompletion30(self.master)

    def getdata(self, labels):
        for i in labels:
            if i in dict.keys():
                print()


class InformationCompletion30(Frame):
    def __init__(self, master=None):
        super().__init__(master)
        self.master.title("第30页")
        self.master = master
        self.pack()
        self.current_page = 0
        self.create_widgets()

    def create_widgets(self):

        # Labels and input boxes
        self.labels = ["工程名称", "建设单位", "梯号", "放线日期", "分包单位"]
        self.entries = []
        for i in range(len(self.labels)):
            label = Label(self, text=self.labels[i], font=('Arial', 14))
            label.grid(row=i, column=0, padx=5, pady=5, sticky='w')
            entry = Entry(self, font=('Arial', 14), width=30)
            if self.labels[i] in dict:
                entry.insert(0, dict[self.labels[i]])
            entry.grid(row=i, column=1, padx=5, pady=5)
            self.entries.append(entry)

        # Save button

        # Next button
        self.next_button = Button(self, text="下一页", font=('Arial', 14), command=self.create_widgets3)
        self.next_button.grid(row=len(self.labels) + 1, column=0, columnspan=2, pady=5)

        self.agrens_button = Button(self, text="上一页", font=('Arial', 14), command=self.agrens_create)
        self.agrens_button.grid(row=len(self.labels) + 2, column=0, columnspan=2, pady=5)

    def agrens_create(self):
        self.destroy()
        InformationCompletion28(self.master)

    def create_widgets3(self):
        for i in range(len(self.labels)):
            dict[self.labels[i]] = self.entries[i].get()

        self.pack_forget()  # 隐藏第一页的部件
        self.destroy()
        InformationCompletion44(self.master)

    def getdata(self, labels):
        for i in labels:
            if i in dict.keys():
                print()


class InformationCompletion44(Frame):
    def __init__(self, master=None):
        super().__init__(master)
        self.master.title("第44页")
        self.master = master
        self.pack()
        self.current_page = 0
        self.create_widgets()

    def create_widgets(self):

        # Labels and input boxes
        self.labels = ["工程名称", "自检部位", "自检项目", "操作日期",
                       "完成日期"]
        self.entries = []
        for i in range(len(self.labels)):
            label = Label(self, text=self.labels[i], font=('Arial', 14))
            label.grid(row=i, column=0, padx=5, pady=5, sticky='w')
            entry = Entry(self, font=('Arial', 14), width=30)
            if self.labels[i] in dict:
                entry.insert(0, dict[self.labels[i]])
            entry.grid(row=i, column=1, padx=5, pady=5)
            self.entries.append(entry)

        # Save button

        # Next button
        self.next_button = Button(self, text="下一页", font=('Arial', 14), command=self.create_widgets3)
        self.next_button.grid(row=len(self.labels) + 1, column=0, columnspan=2, pady=5)

        self.agrens_button = Button(self, text="上一页", font=('Arial', 14), command=self.agrens_create)
        self.agrens_button.grid(row=len(self.labels) + 2, column=0, columnspan=2, pady=5)

    def agrens_create(self):
        self.destroy()
        InformationCompletion30(self.master)

    def create_widgets3(self):
        for i in range(len(self.labels)):
            dict[self.labels[i]] = self.entries[i].get()

        self.pack_forget()  # 隐藏第一页的部件
        self.destroy()
        InformationCompletion45(self.master)

    def getdata(self, labels):
        for i in labels:
            if i in dict.keys():
                print()


class InformationCompletion45(Frame):
    def __init__(self, master=None):
        super().__init__(master)
        self.master.title("第45页")
        self.master = master
        self.pack()
        self.current_page = 0
        self.create_widgets()

    def create_widgets(self):

        # Labels and input boxes
        self.labels = ["工程名称", "移交部门名称", "接收部门名称", "交接部位",
                       "检查日期"]
        self.entries = []
        for i in range(len(self.labels)):
            label = Label(self, text=self.labels[i], font=('Arial', 14))
            label.grid(row=i, column=0, padx=5, pady=5, sticky='w')
            entry = Entry(self, font=('Arial', 14), width=30)
            if self.labels[i] in dict:
                entry.insert(0, dict[self.labels[i]])
            entry.grid(row=i, column=1, padx=5, pady=5)
            self.entries.append(entry)

        # Save button

        # Next button
        self.next_button = Button(self, text="下一页", font=('Arial', 14), command=self.create_widgets3)
        self.next_button.grid(row=len(self.labels) + 1, column=0, columnspan=2, pady=5)

        self.agrens_button = Button(self, text="上一页", font=('Arial', 14), command=self.agrens_create)
        self.agrens_button.grid(row=len(self.labels) + 2, column=0, columnspan=2, pady=5)

    def agrens_create(self):
        self.destroy()
        InformationCompletion44(self.master)

    def create_widgets3(self):
        for i in range(len(self.labels)):
            dict[self.labels[i]] = self.entries[i].get()

        self.pack_forget()  # 隐藏第一页的部件
        self.destroy()
        InformationCompletion46(self.master)

    def getdata(self, labels):
        for i in labels:
            if i in dict.keys():
                print()


class InformationCompletion46(Frame):
    def __init__(self, master=None):
        super().__init__(master)
        self.master.title("第46页")
        self.master = master
        self.pack()
        self.current_page = 0
        self.create_widgets()

    def create_widgets(self):

        # Labels and input boxes
        self.labels = ["工程名称", "复核项目", "复核部位", "复查日期",
                       "分包单位"]
        self.entries = []
        for i in range(len(self.labels)):
            label = Label(self, text=self.labels[i], font=('Arial', 14))
            label.grid(row=i, column=0, padx=5, pady=5, sticky='w')
            entry = Entry(self, font=('Arial', 14), width=30)
            if self.labels[i] in dict:
                entry.insert(0, dict[self.labels[i]])
            entry.grid(row=i, column=1, padx=5, pady=5)
            self.entries.append(entry)

        # Save button

        # Next button
        self.next_button = Button(self, text="下一页", font=('Arial', 14), command=self.create_widgets3)
        self.next_button.grid(row=len(self.labels) + 1, column=0, columnspan=2, pady=5)

        self.agrens_button = Button(self, text="上一页", font=('Arial', 14), command=self.agrens_create)
        self.agrens_button.grid(row=len(self.labels) + 2, column=0, columnspan=2, pady=5)

    def agrens_create(self):
        self.destroy()
        InformationCompletion45(self.master)

    def create_widgets3(self):
        for i in range(len(self.labels)):
            dict[self.labels[i]] = self.entries[i].get()

        self.pack_forget()  # 隐藏第一页的部件
        self.destroy()
        InformationCompletion47(self.master)

    def getdata(self, labels):
        for i in labels:
            if i in dict.keys():
                print()


class InformationCompletion47(Frame):
    def __init__(self, master=None):
        super().__init__(master)
        self.master.title("第47页")
        self.master = master
        self.pack()
        self.current_page = 0
        self.create_widgets()

    def create_widgets(self):

        # Labels and input boxes
        self.labels = ["工程名称"]
        self.entries = []
        for i in range(len(self.labels)):
            label = Label(self, text=self.labels[i], font=('Arial', 14))
            label.grid(row=i, column=0, padx=5, pady=5, sticky='w')
            entry = Entry(self, font=('Arial', 14), width=30)
            if self.labels[i] in dict:
                entry.insert(0, dict[self.labels[i]])
            entry.grid(row=i, column=1, padx=5, pady=5)
            self.entries.append(entry)

        # Save button

        # Next button
        self.next_button = Button(self, text="下一页", font=('Arial', 14), command=self.create_widgets3)
        self.next_button.grid(row=len(self.labels) + 1, column=0, columnspan=2, pady=5)

        self.agrens_button = Button(self, text="上一页", font=('Arial', 14), command=self.agrens_create)
        self.agrens_button.grid(row=len(self.labels) + 2, column=0, columnspan=2, pady=5)

    def agrens_create(self):
        self.destroy()
        InformationCompletion46(self.master)

    def create_widgets3(self):
        for i in range(len(self.labels)):
            dict[self.labels[i]] = self.entries[i].get()

        self.pack_forget()  # 隐藏第一页的部件
        self.destroy()
        InformationCompletion48(self.master)

    def getdata(self, labels):
        for i in labels:
            if i in dict.keys():
                print()


class InformationCompletion48(Frame):
    def __init__(self, master=None):
        super().__init__(master)
        self.master.title("第48页")
        self.master = master
        self.pack()
        self.current_page = 0
        self.create_widgets()

    def create_widgets(self):

        # Labels and input boxes
        self.labels = ["工程名称", "建设单位", "分包单位", "施工部位"]
        self.entries = []
        for i in range(len(self.labels)):
            label = Label(self, text=self.labels[i], font=('Arial', 14))
            label.grid(row=i, column=0, padx=5, pady=5, sticky='w')
            entry = Entry(self, font=('Arial', 14), width=30)
            if self.labels[i] in dict:
                entry.insert(0, dict[self.labels[i]])
            entry.grid(row=i, column=1, padx=5, pady=5)
            self.entries.append(entry)

        # Save button

        # Next button
        self.next_button = Button(self, text="下一页", font=('Arial', 14), command=self.create_widgets3)
        self.next_button.grid(row=len(self.labels) + 1, column=0, columnspan=2, pady=5)

        self.agrens_button = Button(self, text="上一页", font=('Arial', 14), command=self.agrens_create)
        self.agrens_button.grid(row=len(self.labels) + 2, column=0, columnspan=2, pady=5)

    def agrens_create(self):
        self.destroy()
        InformationCompletion47(self.master)

    def create_widgets3(self):
        for i in range(len(self.labels)):
            dict[self.labels[i]] = self.entries[i].get()

        self.pack_forget()  # 隐藏第一页的部件
        self.destroy()
        InformationCompletion49(self.master)

    def getdata(self, labels):
        for i in labels:
            if i in dict.keys():
                print()


class InformationCompletion49(Frame):
    def __init__(self, master=None):
        super().__init__(master)
        self.master.title("第49页")
        self.master = master
        self.pack()
        self.current_page = 0
        self.create_widgets()

    def create_widgets(self):

        # Labels and input boxes
        self.labels = ["子分部工程名称", "施工日期",
                       "分包单位", "施工部位"]
        self.entries = []
        for i in range(len(self.labels)):
            label = Label(self, text=self.labels[i], font=('Arial', 14))
            label.grid(row=i, column=0, padx=5, pady=5, sticky='w')
            entry = Entry(self, font=('Arial', 14), width=30)
            if self.labels[i] in dict:
                entry.insert(0, dict[self.labels[i]])
            entry.grid(row=i, column=1, padx=5, pady=5)
            self.entries.append(entry)

        # Save button

        # Next button
        self.next_button = Button(self, text="下一页", font=('Arial', 14), command=self.create_widgets3)
        self.next_button.grid(row=len(self.labels) + 1, column=0, columnspan=2, pady=5)

        self.agrens_button = Button(self, text="上一页", font=('Arial', 14), command=self.agrens_create)
        self.agrens_button.grid(row=len(self.labels) + 2, column=0, columnspan=2, pady=5)

    def agrens_create(self):
        self.destroy()
        InformationCompletion48(self.master)

    def create_widgets3(self):
        for i in range(len(self.labels)):
            dict[self.labels[i]] = self.entries[i].get()

        self.pack_forget()  # 隐藏第一页的部件
        self.destroy()
        InformationCompletion50(self.master)

    def getdata(self, labels):
        for i in labels:
            if i in dict.keys():
                print()


class InformationCompletion50(Frame):
    def __init__(self, master=None):
        super().__init__(master)
        self.master.title("第50页")
        self.master = master
        self.pack()
        self.current_page = 0
        self.create_widgets()

    def create_widgets(self):

        # Labels and input boxes
        self.labels = ["工程名称"]
        self.entries = []
        for i in range(len(self.labels)):
            label = Label(self, text=self.labels[i], font=('Arial', 14))
            label.grid(row=i, column=0, padx=5, pady=5, sticky='w')
            entry = Entry(self, font=('Arial', 14), width=30)
            if self.labels[i] in dict:
                entry.insert(0, dict[self.labels[i]])
            entry.grid(row=i, column=1, padx=5, pady=5)
            self.entries.append(entry)

        # Save button

        # Next button
        self.next_button = Button(self, text="下一页", font=('Arial', 14), command=self.create_widgets3)
        self.next_button.grid(row=len(self.labels) + 1, column=0, columnspan=2, pady=5)

        self.agrens_button = Button(self, text="上一页", font=('Arial', 14), command=self.agrens_create)
        self.agrens_button.grid(row=len(self.labels) + 2, column=0, columnspan=2, pady=5)

    def agrens_create(self):
        self.destroy()
        InformationCompletion49(self.master)

    def create_widgets3(self):
        for i in range(len(self.labels)):
            dict[self.labels[i]] = self.entries[i].get()

        self.pack_forget()  # 隐藏第一页的部件
        self.destroy()
        InformationCompletion51(self.master)

    def getdata(self, labels):
        for i in labels:
            if i in dict.keys():
                print()


class InformationCompletion51(Frame):
    def __init__(self, master=None):
        super().__init__(master)
        self.master.title("第51页")
        self.master = master
        self.pack()
        self.current_page = 0
        self.create_widgets()

    def create_widgets(self):

        # Labels and input boxes
        self.labels = ["子分部工程名称", "子分部工程数量", "分项工程数量", "总包单位",
                       "总包单位项目负责人", "分包单位", "分包单位负责人",
                       "分包内容"]
        self.entries = []
        for i in range(len(self.labels)):
            label = Label(self, text=self.labels[i], font=('Arial', 14))
            label.grid(row=i, column=0, padx=5, pady=5, sticky='w')
            entry = Entry(self, font=('Arial', 14), width=30)
            if self.labels[i] in dict:
                entry.insert(0, dict[self.labels[i]])
            entry.grid(row=i, column=1, padx=5, pady=5)
            self.entries.append(entry)

        # Save button

        # Next button
        self.next_button = Button(self, text="下一页", font=('Arial', 14), command=self.create_widgets3)
        self.next_button.grid(row=len(self.labels) + 1, column=0, columnspan=2, pady=5)

        self.agrens_button = Button(self, text="上一页", font=('Arial', 14), command=self.agrens_create)
        self.agrens_button.grid(row=len(self.labels) + 2, column=0, columnspan=2, pady=5)

    def agrens_create(self):
        self.destroy()
        InformationCompletion50(self.master)

    def create_widgets3(self):
        for i in range(len(self.labels)):
            dict[self.labels[i]] = self.entries[i].get()

        self.pack_forget()  # 隐藏第一页的部件
        self.destroy()
        InformationCompletion52(self.master)

    def getdata(self, labels):
        for i in labels:
            if i in dict.keys():
                print()


class InformationCompletion52(Frame):
    def __init__(self, master=None):
        super().__init__(master)
        self.master.title("第52页")
        self.master = master
        self.pack()
        self.current_page = 0
        self.create_widgets()

    def create_widgets(self):

        # Labels and input boxes
        self.labels = ["工程名称", "分包单位", "总包单位项目负责人", "总监理工程师"]
        self.entries = []
        for i in range(len(self.labels)):
            label = Label(self, text=self.labels[i], font=('Arial', 14))
            label.grid(row=i, column=0, padx=5, pady=5, sticky='w')
            entry = Entry(self, font=('Arial', 14), width=30)
            if self.labels[i] in dict:
                entry.insert(0, dict[self.labels[i]])
            entry.grid(row=i, column=1, padx=5, pady=5)
            self.entries.append(entry)

        # Save button

        # Next button
        self.next_button = Button(self, text="下一页", font=('Arial', 14), command=self.create_widgets3)
        self.next_button.grid(row=len(self.labels) + 1, column=0, columnspan=2, pady=5)

        self.agrens_button = Button(self, text="上一页", font=('Arial', 14), command=self.agrens_create)
        self.agrens_button.grid(row=len(self.labels) + 2, column=0, columnspan=2, pady=5)

    def agrens_create(self):
        self.destroy()
        InformationCompletion51(self.master)

    def create_widgets3(self):
        for i in range(len(self.labels)):
            dict[self.labels[i]] = self.entries[i].get()

        self.pack_forget()  # 隐藏第一页的部件
        self.destroy()
        InformationCompletion53(self.master)

    def getdata(self, labels):
        for i in labels:
            if i in dict.keys():
                print()


class InformationCompletion53(Frame):
    def __init__(self, master=None):
        super().__init__(master)
        self.master.title("第53页")
        self.master = master
        self.pack()
        self.current_page = 0
        self.create_widgets()

    def create_widgets(self):

        # Labels and input boxes
        self.labels = ["工程名称", "分包单位", "总包单位项目负责人", "总监理工程师"]
        self.entries = []
        for i in range(len(self.labels)):
            label = Label(self, text=self.labels[i], font=('Arial', 14))
            label.grid(row=i, column=0, padx=5, pady=5, sticky='w')
            entry = Entry(self, font=('Arial', 14), width=30)
            if self.labels[i] in dict:
                entry.insert(0, dict[self.labels[i]])
            entry.grid(row=i, column=1, padx=5, pady=5)
            self.entries.append(entry)

        # Save button

        # Next button
        self.next_button = Button(self, text="下一页", font=('Arial', 14), command=self.create_widgets3)
        self.next_button.grid(row=len(self.labels) + 1, column=0, columnspan=2, pady=5)

        self.agrens_button = Button(self, text="上一页", font=('Arial', 14), command=self.agrens_create)
        self.agrens_button.grid(row=len(self.labels) + 2, column=0, columnspan=2, pady=5)

    def agrens_create(self):
        self.destroy()
        InformationCompletion52(self.master)

    def create_widgets3(self):
        for i in range(len(self.labels)):
            dict[self.labels[i]] = self.entries[i].get()

        self.pack_forget()  # 隐藏第一页的部件
        self.destroy()
        InformationCompletion54(self.master)

    def getdata(self, labels):
        for i in labels:
            if i in dict.keys():
                print()


class InformationCompletion54(Frame):
    def __init__(self, master=None):
        super().__init__(master)
        self.master.title("第54页")
        self.master = master
        self.pack()
        self.current_page = 0
        self.create_widgets()

    def create_widgets(self):

        # Labels and input boxes
        self.labels = ["工程名称", "分包单位", "总包单位项目负责人", "总监理工程师"]
        self.entries = []
        for i in range(len(self.labels)):
            label = Label(self, text=self.labels[i], font=('Arial', 14))
            label.grid(row=i, column=0, padx=5, pady=5, sticky='w')
            entry = Entry(self, font=('Arial', 14), width=30)
            if self.labels[i] in dict:
                entry.insert(0, dict[self.labels[i]])
            entry.grid(row=i, column=1, padx=5, pady=5)
            self.entries.append(entry)

        # Save button

        # Next button
        self.next_button = Button(self, text="下一页", font=('Arial', 14), command=self.create_widgets3)
        self.next_button.grid(row=len(self.labels) + 1, column=0, columnspan=2, pady=5)

        self.agrens_button = Button(self, text="上一页", font=('Arial', 14), command=self.agrens_create)
        self.agrens_button.grid(row=len(self.labels) + 2, column=0, columnspan=2, pady=5)

    def agrens_create(self):
        self.destroy()
        InformationCompletion53(self.master)

    def create_widgets3(self):
        for i in range(len(self.labels)):
            dict[self.labels[i]] = self.entries[i].get()

        self.pack_forget()  # 隐藏第一页的部件
        self.destroy()
        InformationCompletion58(self.master)

    def getdata(self, labels):
        for i in labels:
            if i in dict.keys():
                print()


class InformationCompletion58(Frame):
    def __init__(self, master=None):
        super().__init__(master)
        self.master.title("第58页")
        self.master = master
        self.pack()
        self.current_page = 0
        self.create_widgets()

    def create_widgets(self):

        # Labels and input boxes
        self.labels = ["检验名字", "子分部工程名称", "检验批名称", "检验批编号", "建设单位项目负责人",
                       "检查日期"]
        self.entries = []
        for i in range(len(self.labels)):
            label = Label(self, text=self.labels[i], font=('Arial', 14))
            label.grid(row=i, column=0, padx=5, pady=5, sticky='w')
            entry = Entry(self, font=('Arial', 14), width=30)
            if self.labels[i] in dict:
                entry.insert(0, dict[self.labels[i]])
            entry.grid(row=i, column=1, padx=5, pady=5)
            self.entries.append(entry)

        # Next button
        self.next_button = Button(self, text="下一页", font=('Arial', 14), command=self.create_widgets3)
        self.next_button.grid(row=len(self.labels) + 1, column=0, columnspan=2, pady=5)

        self.agrens_button = Button(self, text="上一页", font=('Arial', 14), command=self.agrens_create)
        self.agrens_button.grid(row=len(self.labels) + 2, column=0, columnspan=2, pady=5)

    def agrens_create(self):
        self.destroy()
        InformationCompletion54(self.master)

    def create_widgets3(self):
        for i in range(len(self.labels)):
            dict[self.labels[i]] = self.entries[i].get()

        self.pack_forget()  # 隐藏第一页的部件
        self.destroy()
        InformationCompletion60(self.master)

    def getdata(self, labels):
        for i in labels:
            if i in dict.keys():
                print()


class InformationCompletion60(Frame):
    def __init__(self, master=None):
        super().__init__(master)
        self.master.title("第60页")
        self.master = master
        self.pack()
        self.current_page = 0
        self.create_widgets()

    def create_widgets(self):

        # Labels and input boxes
        self.labels = ["子分部工程名称"]
        self.entries = []
        for i in range(len(self.labels)):
            label = Label(self, text=self.labels[i], font=('Arial', 14))
            label.grid(row=i, column=0, padx=5, pady=5, sticky='w')
            entry = Entry(self, font=('Arial', 14), width=30)
            if self.labels[i] in dict:
                entry.insert(0, dict[self.labels[i]])
            entry.grid(row=i, column=1, padx=5, pady=5)
            self.entries.append(entry)

        # Next button
        self.next_button = Button(self, text="下一页", font=('Arial', 14), command=self.create_widgets3)
        self.next_button.grid(row=len(self.labels) + 1, column=0, columnspan=2, pady=5)

        self.agrens_button = Button(self, text="上一页", font=('Arial', 14), command=self.agrens_create)
        self.agrens_button.grid(row=len(self.labels) + 2, column=0, columnspan=2, pady=5)

    def agrens_create(self):
        self.destroy()
        InformationCompletion58(self.master)

    def create_widgets3(self):
        for i in range(len(self.labels)):
            dict[self.labels[i]] = self.entries[i].get()

        self.pack_forget()  # 隐藏第一页的部件
        self.destroy()
        InformationCompletion71(self.master)

    def getdata(self, labels):
        for i in labels:
            if i in dict.keys():
                print()


class InformationCompletion71(Frame):
    def __init__(self, master=None):
        super().__init__(master)
        self.master.title("第71页")
        self.master = master
        self.pack()
        self.current_page = 0
        self.create_widgets()

    def create_widgets(self):

        # Labels and input boxes
        self.labels = ["工程名称", "监理单位"]
        self.entries = []
        for i in range(len(self.labels)):
            label = Label(self, text=self.labels[i], font=('Arial', 14))
            label.grid(row=i, column=0, padx=5, pady=5, sticky='w')
            entry = Entry(self, font=('Arial', 14), width=30)
            if self.labels[i] in dict:
                entry.insert(0, dict[self.labels[i]])
            entry.grid(row=i, column=1, padx=5, pady=5)
            self.entries.append(entry)

        # Save button
        self.save_button = Button(self, text="保存", font=('Arial', 14), command=self.save_info)
        self.save_button.grid(row=len(self.labels), column=0, columnspan=2, pady=5)

        self.agrens_button = Button(self, text="再生成一份", font=('Arial', 14), command=self.agrens_create)
        self.agrens_button.grid(row=len(self.labels) + 1, column=0, columnspan=2, pady=5)

    def agrens_create(self):
        self.destroy()
        InformationCompletion60(self.master)

    def save_info(self):
        global numtest1

        async def copy_sheet(source_sheet, target_sheet):
            # Tab color
            global numtest1
            target_sheet.sheet_properties.tabColor = source_sheet.sheet_properties.tabColor

            # 复制列宽度
            for column in source_sheet.column_dimensions:
                target_sheet.column_dimensions[column].width = source_sheet.column_dimensions[column].width

            # 复制行高
            for row in source_sheet.row_dimensions:
                target_sheet.row_dimensions[row].height = source_sheet.row_dimensions[row].height

            # 处理合并单元格
            merged_cells = list(source_sheet.merged_cells)
            if len(merged_cells) > 0:
                for merged_cell in merged_cells:
                    cell2 = str(merged_cell).replace('(<CellRange ', '').replace('>,)', '')
                    target_sheet.merge_cells(cell2)

            for i, row in enumerate(source_sheet.iter_rows(max_row=60)):  # 复制前60行
                if i >= 60:
                    break
                for j, cell in enumerate(row):
                    if j >= 100:  # 仅复制前100列
                        break
                    target_sheet.cell(row=i + 1, column=j + 1, value=cell.value)

                    # 如果是合并单元格的一部分，则设置单元格格式和尺寸
                    if cell.coordinate in source_sheet.merged_cells:
                        for merged_cell in merged_cells:
                            if cell.coordinate in merged_cell:
                                first_cell = source_sheet.cell(merged_cell.min_row, merged_cell.min_col)
                                target_cell = target_sheet.cell(i + 1, j + 1)
                                target_sheet.row_dimensions[i + 1].height = source_sheet.row_dimensions[
                                    first_cell.row].height
                                target_sheet.column_dimensions[get_column_letter(j + 1)].width = \
                                    source_sheet.column_dimensions[get_column_letter(first_cell.column)].width
                                break

                    # 设置单元格格式
                    source_cell = source_sheet.cell(i + 1, j + 1)
                    target_cell = target_sheet.cell(i + 1, j + 1)

                    # 创建新样式对象
                    font = Font(name=source_cell.font.name, size=source_cell.font.size,
                                bold=source_cell.font.bold, italic=source_cell.font.italic,
                                vertAlign=source_cell.font.vertAlign, underline=source_cell.font.underline,
                                strike=source_cell.font.strike, color='FF000000')  # Set font color to black

                    black_side = Side(color='FF000000')
                    border = Border(
                        left=Side(color=black_side.color, border_style=source_cell.border.left.border_style),
                        right=Side(color=black_side.color, border_style=source_cell.border.right.border_style),
                        top=Side(color=black_side.color, border_style=source_cell.border.top.border_style),
                        bottom=Side(color=black_side.color, border_style=source_cell.border.bottom.border_style))

                    fill = PatternFill(fill_type=source_cell.fill.fill_type, fgColor=source_cell.fill.fgColor,
                                       bgColor=source_cell.fill.bgColor, patternType=source_cell.fill.patternType)

                    alignment = Alignment(horizontal=source_cell.alignment.horizontal,
                                          vertical=source_cell.alignment.vertical,
                                          text_rotation=source_cell.alignment.text_rotation,
                                          wrap_text=source_cell.alignment.wrap_text,
                                          shrink_to_fit=source_cell.alignment.shrink_to_fit,
                                          indent=source_cell.alignment.indent)

                    protection = Protection(locked=source_cell.protection.locked,
                                            hidden=source_cell.protection.hidden)

                    # 将新样式对象分配给目标单元格
                    target_cell.font = font
                    target_cell.border = border
                    target_cell.fill = fill
                    target_cell.alignment = alignment
                    target_cell.protection = protection
                    target_cell.number_format = source_cell.number_format


        try:
            async def main():
                global numtest1
                path = r"./直梯/01-【第二十一卷】电梯工程施工技术资料.xlsx"
                save_path = "直梯.xlsx"
                wb2 = openpyxl.Workbook()
                wb2.save(save_path)
                wb = openpyxl.load_workbook(path)
                wb2 = openpyxl.load_workbook(save_path)
                sheetnames = wb.sheetnames

                tasks = []
                for numtest, sheetname in enumerate(sheetnames, start=1):
                    sheet = wb[sheetname]
                    sheet2 = wb2.create_sheet(f"第{numtest1}页")
                    print(f"第{numtest1}页")
                    numtest1 += 1

                    task = asyncio.create_task(copy_sheet(sheet, sheet2))
                    tasks.append(task)

                # 等待所有任务完成
                await asyncio.gather(*tasks)

                if 'Sheet' in wb2.sheetnames:
                    del wb2['Sheet']

                wb2.save(save_path)
                wb.close()
                wb2.close()
                workbook = openpyxl.load_workbook('直梯.xlsx')
                sheet = workbook[f'第{numtest1 - 3}页']
                cell_g41 = sheet['T16']
                cell_g41.font = Font(name='宋体', size=16)
                cell_g41.alignment = Alignment(horizontal='center', vertical='center')
                cell_g41.value = dict["工程名称"]
                cell_g42 = sheet['T17']
                cell_g42.value = dict["建设单位"]
                cell_g43 = sheet['T18']
                cell_g43.value = dict["监理单位"]
                cell_g44 = sheet['W19']
                cell_g44.value = dict["总包单位"]
                cell_g45 = sheet['T20']
                cell_g45.value = dict["分包单位"]
                cell_g46 = sheet['AC21']
                cell_g46.value = dict["分包单位技术负责人"]
                cell_g47 = sheet['T22']
                cell_g47.value = dict["编制人"]
                cell_g48 = sheet['T23']
                if len(str(dict["竣工日期"]).split('-')) >= 3:
                    cell_g48.value = str(dict["竣工日期"]).split('-')[0] + "年" + str(dict["竣工日期"]).split('-')[
                        1] + "月" + \
                                     str(dict["竣工日期"]).split('-')[2] + "日"

                sheet2 = workbook[f'第{numtest1 - 2}页']
                cell_g49 = sheet2['K4']
                cell_g49.value = dict["工程名称"]
                sheet3 = workbook[f'第{numtest1 - 1}页']
                cell_g410 = sheet3['K4']
                cell_g410.value = dict["工程名称"]

                workbook.save('直梯.xlsx')
                workbook.close()

            # 运行异步主程序
            asyncio.run(main())
        except Exception as e:
            print(numtest1,e)

        try:
            async def main():
                global numtest1
                path = r"./直梯/02-表B.0.2 工程开工报审表（后附开工告知）.xlsx"
                save_path = "直梯.xlsx"
                wb = openpyxl.load_workbook(path)
                wb2 = openpyxl.load_workbook(save_path)
                sheetnames = wb.sheetnames

                tasks = []
                for numtest, sheetname in enumerate(sheetnames, start=1):
                    sheet = wb[sheetname]
                    sheet2 = wb2.create_sheet(f"第{numtest1}页")
                    print(numtest1)
                    numtest1 += 1

                    task = asyncio.create_task(copy_sheet(sheet, sheet2))
                    tasks.append(task)

                # 等待所有任务完成
                await asyncio.gather(*tasks)

                if 'Sheet' in wb2.sheetnames:
                    del wb2['Sheet']

                wb2.save(save_path)
                wb.close()
                wb2.close()

                workbook = openpyxl.load_workbook('直梯.xlsx')
                print(numtest1)
                sheet = workbook[f'第{numtest1 - 1}页']

                cell_g4 = sheet['E3']
                cell_g4.value = dict["工程名称"]
                cell_g4 = sheet['D5']
                cell_g4.value = dict["建设单位"]
                cell_g4 = sheet['L13']
                cell_g4.value = dict["分包单位"]
                cell_g4 = sheet['L29']
                cell_g4.value = dict["建设单位"]
                cell_g4 = sheet['D6']
                cell_g4.value = dict["监理单位"]
                cell_g4 = sheet['K21']
                cell_g4.value = dict["监理单位"]
                cell_g4 = sheet['F7']
                cell_g4.value = dict["工程名称"]
                if len(str(dict["开工日期"]).split('-')) >= 3:
                    cell_g4 = sheet['F8']
                    cell_g4.value = str(dict["开工日期"]).split('-')[0]
                    cell_g4 = sheet['I8']
                    cell_g4.value = str(dict["开工日期"]).split('-')[1]
                    cell_g4 = sheet['K8']
                    cell_g4.value = str(dict["开工日期"]).split('-')[2]
                workbook.save('直梯.xlsx')
                workbook.close()
                folder_path = r'.\开工告知'
                workbook = openpyxl.load_workbook('直梯.xlsx')
                # 遍历文件夹中的所有文件
                for filename in os.listdir(folder_path):
                    if filename.endswith('.jpg') or filename.endswith('.png'):
                        file_path = os.path.join(folder_path, filename)
                        # 创建一个新的工作表，并使用图片名称作为工作表名称
                        sheet = workbook.create_sheet(f"第{numtest1}页")
                        # 在工作表中插入图片
                        img = Image(file_path)
                        sheet.add_image(img, 'A1')
                        numtest1 = numtest1 + 1

                # 保存 Excel 文件
                workbook.save('直梯.xlsx')
                workbook.close()

                # 运行异步主程序

            asyncio.run(main())
        except Exception as e:
            print(numtest1, e)

        try:
            async def main():
                global numtest1
                path = r"./直梯/03-鲁DT-001施工现场质量管理检查记录.xlsx"
                save_path = "直梯.xlsx"
                wb = openpyxl.load_workbook(path)
                wb2 = openpyxl.load_workbook(save_path)
                sheetnames = wb.sheetnames

                tasks = []
                for numtest, sheetname in enumerate(sheetnames, start=1):
                    sheet = wb[sheetname]
                    sheet2 = wb2.create_sheet(f"第{numtest1}页")
                    print(numtest1)
                    numtest1 += 1
                    task = asyncio.create_task(copy_sheet(sheet, sheet2))
                    tasks.append(task)

                # 等待所有任务完成
                await asyncio.gather(*tasks)

                wb2.save(save_path)
                wb.close()
                wb2.close()
                workbook = openpyxl.load_workbook('直梯.xlsx')
                sheet = workbook[f'第{numtest1 - 1}页']
                cell_g4 = sheet['I4']
                if len(str(dict["开工日期"]).split('-')) >= 3:
                    cell_g4.value = str(dict["开工日期"]).split('-')[0] + "年" + str(dict["开工日期"]).split('-')[
                        1] + "月" + \
                                    str(dict["开工日期"]).split('-')[2] + "日"
                cell_g4 = sheet['L6']
                cell_g4.value = dict["工程名称"]
                cell_g4 = sheet['AR6']
                cell_g4.value = dict["施工许可证号"]
                cell_g4 = sheet['L7']
                cell_g4.value = dict["建设单位"]
                cell_g4 = sheet['AR7']
                cell_g4.value = dict["建设单位项目负责人"]
                cell_g4 = sheet['L8']
                cell_g4.value = dict["设计单位"]
                cell_g4 = sheet['AR8']
                cell_g4.value = dict["设计单位项目负责人"]
                cell_g4 = sheet['L9']
                cell_g4.value = dict["监理单位"]
                cell_g4 = sheet['AR9']
                cell_g4.value = dict["总监理工程师"]
                cell_g4 = sheet['L10']
                cell_g4.value = dict["分包单位"]
                cell_g4 = sheet['AH10']
                cell_g4.value = dict["分包单位项目负责人"]
                cell_g4 = sheet['BA10']
                cell_g4.value = dict["分包单位技术负责人"]
                workbook.save('直梯.xlsx')
                workbook.close()

                # 运行异步主程序

            asyncio.run(main())
        except Exception as e:
            print(numtest1, e)

        try:
            async def main():
                global numtest1
                path = r"./直梯/04-鲁DT-002工程参建各方签字签章存样表.xlsx"
                save_path = "直梯.xlsx"
                wb = openpyxl.load_workbook(path)
                wb2 = openpyxl.load_workbook(save_path)
                sheetnames = wb.sheetnames

                tasks = []
                for numtest, sheetname in enumerate(sheetnames, start=1):
                    sheet = wb[sheetname]
                    sheet2 = wb2.create_sheet(f"第{numtest1}页")
                    print(numtest1)
                    numtest1 += 1
                    task = asyncio.create_task(copy_sheet(sheet, sheet2))
                    tasks.append(task)

                # 等待所有任务完成
                await asyncio.gather(*tasks)

                # 保存并关闭工作簿
                wb2.save(save_path)
                wb.close()
                wb2.close()
                workbook = openpyxl.load_workbook('直梯.xlsx')
                sheet = workbook[f'第{numtest1 - 2}页']
                cell_g4 = sheet['K6']
                cell_g4.value = dict["工程名称"]
                sheet = workbook[f'第{numtest1 - 1}页']
                cell_g4 = sheet['K6']
                cell_g4.value = dict["工程名称"]
                workbook.save('直梯.xlsx')
                workbook.close()

                # 运行异步主程序

            asyncio.run(main())
        except Exception as e:
            print(numtest1, e)

        try:
            async def main():
                global numtest1
                path = r"./直梯/05-鲁DT-003工程项目管理人员名单（后附人员证件）.xlsx"
                save_path = "直梯.xlsx"
                wb = openpyxl.load_workbook(path)
                wb2 = openpyxl.load_workbook(save_path)
                sheetnames = wb.sheetnames

                tasks = []
                for numtest, sheetname in enumerate(sheetnames, start=1):
                    sheet = wb[sheetname]
                    sheet2 = wb2.create_sheet(f"第{numtest1}页")
                    print(numtest1)
                    numtest1 += 1
                    task = asyncio.create_task(copy_sheet(sheet, sheet2))
                    tasks.append(task)

                # 等待所有任务完成
                await asyncio.gather(*tasks)

                # 保存并关闭工作簿
                wb2.save(save_path)
                wb.close()
                wb2.close()
                workbook = openpyxl.load_workbook('直梯.xlsx')
                sheet = workbook[f'第{numtest1 - 1}页']
                cell_g4 = sheet['K6']
                cell_g4.value = dict["工程名称"]
                cell_g4 = sheet['AM6']
                cell_g4.value = dict["分包单位"]
                cell_g4 = sheet['S18']
                cell_g4.value = dict["工程名称"]
                exls = pd.read_excel(r'./测量值/人员名单.xlsx', keep_default_na=False)
                demo = 8
                for data in exls.values:
                    cell_g4 = sheet[f'B{demo}']
                    cell_g4.value = data[0]
                    cell_g4 = sheet[f'K{demo}']
                    cell_g4.value = data[1]
                    cell_g4 = sheet[f'X{demo}']
                    cell_g4.value = data[2]
                    cell_g4 = sheet[f'AT{demo}']
                    cell_g4.value = data[3]
                    demo = demo + 1

                workbook.save('直梯.xlsx')
                workbook.close()
                folder_path = r'.\人员证件'
                workbook = openpyxl.load_workbook('直梯.xlsx')
                # 遍历文件夹中的所有文件
                for filename in os.listdir(folder_path):
                    if filename.endswith('.jpg') or filename.endswith('.png'):
                        file_path = os.path.join(folder_path, filename)
                        # 创建一个新的工作表，并使用图片名称作为工作表名称
                        sheet = workbook.create_sheet(f"第{numtest1}页")
                        # 在工作表中插入图片
                        img = Image(file_path)
                        sheet.add_image(img, 'A1')
                        numtest1 = numtest1 + 1

                # 保存 Excel 文件
                workbook.save('直梯.xlsx')
                workbook.close()

                # 运行异步主程序

            asyncio.run(main())
        except Exception as e:
            print(numtest1, e)

        try:
            async def main():
                global numtest1
                path = r"./直梯/06-鲁DT-004工程参建各方人员及签章变更备案表.xlsx"
                save_path = "直梯.xlsx"
                wb = openpyxl.load_workbook(path)
                wb2 = openpyxl.load_workbook(save_path)
                sheetnames = wb.sheetnames

                tasks = []
                for numtest, sheetname in enumerate(sheetnames, start=1):
                    sheet = wb[sheetname]
                    sheet2 = wb2.create_sheet(f"第{numtest1}页")
                    print(numtest1)
                    numtest1 += 1
                    task = asyncio.create_task(copy_sheet(sheet, sheet2))
                    tasks.append(task)

                # 等待所有任务完成
                await asyncio.gather(*tasks)

                # 保存并关闭工作簿
                wb2.save(save_path)
                wb.close()
                wb2.close()
                workbook = openpyxl.load_workbook('直梯.xlsx')
                sheet = workbook[f'第{numtest1 - 1}页']
                cell_g4 = sheet['AW7']
                if len(str(dict["开工日期"]).split('-')) >= 3:
                    cell_g4.value = str(dict["开工日期"]).split('-')[0] + "年" + str(dict["开工日期"]).split('-')[
                        1] + "月" + \
                                    str(dict["开工日期"]).split('-')[2] + "日"

                cell_g4 = sheet['K6']
                cell_g4.value = dict["工程名称"]
                cell_g4 = sheet['K7']
                cell_g4.value = dict["分包单位"]

                workbook.save('直梯.xlsx')
                workbook.close()

                # 运行异步主程序

            asyncio.run(main())
        except Exception as e:
            print(numtest1, e)

        try:
            async def main():
                global numtest1
                path = r"./直梯/07-表B.0.4 分包单位资格报审表.xlsx"
                save_path = "直梯.xlsx"
                wb = openpyxl.load_workbook(path)
                wb2 = openpyxl.load_workbook(save_path)
                sheetnames = wb.sheetnames

                tasks = []
                for numtest, sheetname in enumerate(sheetnames, start=1):
                    sheet = wb[sheetname]
                    sheet2 = wb2.create_sheet(f"第{numtest1}页")
                    print(numtest1)
                    numtest1 += 1
                    task = asyncio.create_task(copy_sheet(sheet, sheet2))
                    tasks.append(task)

                # 等待所有任务完成
                await asyncio.gather(*tasks)

                # 保存并关闭工作簿
                wb2.save(save_path)
                wb.close()
                wb2.close()
                workbook = openpyxl.load_workbook('直梯.xlsx')
                sheet = workbook[f'第{numtest1 - 1}页']
                cell_g4 = sheet['E3']
                cell_g4.value = dict["工程名称"]
                cell_g4 = sheet['D5']
                cell_g4.value = dict["监理单位"]
                cell_g4 = sheet['G6']
                cell_g4.value = dict["分包单位"]
                cell_g4 = sheet['L34']
                cell_g4.value = dict["监理单位"]
                workbook.save('直梯.xlsx')
                workbook.close()

                # 运行异步主程序

            asyncio.run(main())
        except Exception as e:
            print(numtest1, e)

        try:
            async def main():
                global numtest1
                path = r"./直梯/08-鲁DT-005分包单位资质报审表（后附公司资质）.xlsx"
                save_path = "直梯.xlsx"
                wb = openpyxl.load_workbook(path)
                wb2 = openpyxl.load_workbook(save_path)
                sheetnames = wb.sheetnames

                tasks = []
                for numtest, sheetname in enumerate(sheetnames, start=1):
                    sheet = wb[sheetname]
                    sheet2 = wb2.create_sheet(f"第{numtest1}页")
                    print(numtest1)
                    numtest1 += 1
                    task = asyncio.create_task(copy_sheet(sheet, sheet2))
                    tasks.append(task)

                # 等待所有任务完成
                await asyncio.gather(*tasks)

                # 保存并关闭工作簿
                wb2.save(save_path)
                wb.close()
                wb2.close()
                workbook = openpyxl.load_workbook('直梯.xlsx')
                sheet = workbook[f'第{numtest1 - 1}页']
                if len(str(dict["开工日期"]).split('-')) >= 3:
                    cell_g4 = sheet['AY8']
                    cell_g4.value = str(dict["开工日期"]).split('-')[0] + "年" + str(dict["开工日期"]).split('-')[
                        1] + "月" + \
                                    str(dict["开工日期"]).split('-')[2] + "日"
                cell_g4 = sheet['N6']
                cell_g4.value = dict["工程名称"]
                cell_g4 = sheet['N7']
                cell_g4.value = dict["总包单位"]
                cell_g4 = sheet['N8']
                cell_g4.value = dict["分包单位"]
                cell_g4 = sheet['F9']
                cell_g4.value = dict["监理单位"]
                cell_g4 = sheet['V11']
                cell_g4.value = dict["分包单位"]
                exls = pd.read_excel(r'./测量值/分包项目.xlsx', keep_default_na=False)
                datanum = 19
                for data in exls.values:
                    cell_g4 = sheet[f'B{datanum}']
                    cell_g4.value = data[0]
                    cell_g4 = sheet[f'R{datanum}']
                    cell_g4.value = data[1]
                    cell_g4 = sheet[f'AH{datanum}']
                    cell_g4.value = data[2]
                    cell_g4 = sheet[f'AX{datanum}']
                    cell_g4.value = data[3]
                    datanum = datanum + 1

                workbook.save('直梯.xlsx')
                workbook.close()
                folder_path = r'.\公司资质'
                workbook = openpyxl.load_workbook('直梯.xlsx')
                # 遍历文件夹中的所有文件
                for filename in os.listdir(folder_path):
                    if filename.endswith('.jpg') or filename.endswith('.png'):
                        file_path = os.path.join(folder_path, filename)
                        # 创建一个新的工作表，并使用图片名称作为工作表名称
                        sheet = workbook.create_sheet(f"第{numtest1}页")
                        # 在工作表中插入图片
                        img = Image(file_path)
                        sheet.add_image(img, 'A1')
                        numtest1 = numtest1 + 1

                # 保存 Excel 文件
                workbook.save('直梯.xlsx')
                workbook.close()

            # 运行异步主程序
            asyncio.run(main())
        except Exception as e:
            print(numtest1, e)

        try:
            async def main():
                global numtest1
                path = r"./直梯/09-鲁DT-006工程质量事故调（勘）查记录.xlsx"
                save_path = "直梯.xlsx"
                wb = openpyxl.load_workbook(path)
                wb2 = openpyxl.load_workbook(save_path)
                sheetnames = wb.sheetnames

                tasks = []
                for numtest, sheetname in enumerate(sheetnames, start=1):
                    sheet = wb[sheetname]
                    sheet2 = wb2.create_sheet(f"第{numtest1}页")
                    print(numtest1)
                    numtest1 += 1
                    task = asyncio.create_task(copy_sheet(sheet, sheet2))
                    tasks.append(task)

                # 等待所有任务完成
                await asyncio.gather(*tasks)

                # 保存并关闭工作簿
                wb2.save(save_path)
                wb.close()
                wb2.close()
                workbook = openpyxl.load_workbook('直梯.xlsx')
                sheet = workbook[f'第{numtest1 - 1}页']
                cell_g4 = sheet['AV6']
                if len(str(dict["调查日期"]).split('-')) >= 3:
                    cell_g4.value = str(dict["调查日期"]).split('-')[0] + "年" + str(dict["调查日期"]).split('-')[
                        1] + "月" + \
                                    str(dict["调查日期"]).split('-')[2] + "日"

                    cell_g4 = sheet['K7']
                    cell_g4.value = str(dict["调查日期"]).split('-')[0] + "年" + str(dict["调查日期"]).split('-')[
                        1] + "月" + \
                                    str(dict["调查日期"]).split('-')[2] + "日"
                cell_g4 = sheet['K6']
                cell_g4.value = dict["工程名称"]
                cell_g4 = sheet['AH7']
                cell_g4.value = dict["调(勘)查时间起始时间"]
                cell_g4 = sheet['AQ7']
                cell_g4.value = dict["调(勘)查时间终止时间"]
                cell_g4 = sheet['K8']
                cell_g4.value = dict["调(勘)查地点"]
                workbook.save('直梯.xlsx')
                workbook.close()

            # 运行异步主程序
            asyncio.run(main())
        except Exception as e:
            print(numtest1, e)

        try:
            async def main():
                global numtest1
                path = r"./直梯/10-鲁DT-007建设工程质量事故报告.xlsx"
                save_path = "直梯.xlsx"
                wb = openpyxl.load_workbook(path)
                wb2 = openpyxl.load_workbook(save_path)
                sheetnames = wb.sheetnames

                tasks = []
                for numtest, sheetname in enumerate(sheetnames, start=1):
                    sheet = wb[sheetname]
                    sheet2 = wb2.create_sheet(f"第{numtest1}页")
                    print(numtest1)
                    numtest1 += 1
                    task = asyncio.create_task(copy_sheet(sheet, sheet2))
                    tasks.append(task)

                # 等待所有任务完成
                await asyncio.gather(*tasks)

                # 保存并关闭工作簿
                wb2.save(save_path)
                wb.close()
                wb2.close()
                workbook = openpyxl.load_workbook('直梯.xlsx')
                sheet = workbook[f'第{numtest1 - 1}页']
                cell_g4 = sheet['K6']
                cell_g4.value = dict["工程名称"]
                cell_g4 = sheet['K7']
                cell_g4.value = dict["建设单位"]
                cell_g4 = sheet['AT7']
                cell_g4.value = dict["分包单位"]
                cell_g4 = sheet['K8']
                cell_g4.value = dict["设计单位"]
                cell_g4 = sheet['K9']
                cell_g4.value = dict["工程地址"]

                workbook.save('直梯.xlsx')
                workbook.close()
                folder_path = r'.\施工日志'
                workbook = openpyxl.load_workbook('直梯.xlsx')
                # 遍历文件夹中的所有文件
                for filename in os.listdir(folder_path):
                    if filename.endswith('.jpg') or filename.endswith('.png'):
                        file_path = os.path.join(folder_path, filename)
                        # 创建一个新的工作表，并使用图片名称作为工作表名称
                        sheet = workbook.create_sheet(f"第{numtest1}页")
                        # 在工作表中插入图片
                        img = Image(file_path)
                        sheet.add_image(img, 'A1')
                        numtest1 = numtest1 + 1

                # 保存 Excel 文件
                workbook.save('直梯.xlsx')
                workbook.close()

            # 运行异步主程序
            asyncio.run(main())
        except Exception as e:
            print(numtest1, e)

        try:
            async def main():
                global numtest1
                path = r"./直梯/12-表B.0.1 施工组织设计（专项）施工方案报审表.xlsx"
                save_path = "直梯.xlsx"
                wb = openpyxl.load_workbook(path)
                wb2 = openpyxl.load_workbook(save_path)
                sheetnames = wb.sheetnames

                tasks = []
                for numtest, sheetname in enumerate(sheetnames, start=1):
                    sheet = wb[sheetname]
                    sheet2 = wb2.create_sheet(f"第{numtest1}页")
                    print(numtest1)
                    numtest1 += 1
                    task = asyncio.create_task(copy_sheet(sheet, sheet2))
                    tasks.append(task)

                # 等待所有任务完成
                await asyncio.gather(*tasks)

                # 保存并关闭工作簿
                wb2.save(save_path)
                wb.close()
                wb2.close()
                workbook = openpyxl.load_workbook('直梯.xlsx')
                sheet = workbook[f'第{numtest1 - 1}页']
                cell_g4 = sheet['D3']
                cell_g4.value = dict["工程名称"]
                cell_g4 = sheet['C5']
                cell_g4.value = dict["监理单位"]
                cell_g4 = sheet['E6']
                cell_g4.value = dict["工程名称"]
                cell_g4 = sheet['K12']
                cell_g4.value = dict["分包单位"]
                cell_g4 = sheet['J27']
                cell_g4.value = dict["监理单位"]
                cell_g4 = sheet['K33']
                cell_g4.value = dict["建设单位"]
                workbook.save('直梯.xlsx')
                workbook.close()

            # 运行异步主程序
            asyncio.run(main())
        except Exception as e:
            print(numtest1, e)

        try:
            async def main():
                global numtest1
                path = r"./直梯/13-鲁DT-009施工组织设计（施工方案）审批表（后附施工方案）.xlsx"
                save_path = "直梯.xlsx"
                wb = openpyxl.load_workbook(path)
                wb2 = openpyxl.load_workbook(save_path)
                sheetnames = wb.sheetnames

                tasks = []
                for numtest, sheetname in enumerate(sheetnames, start=1):
                    sheet = wb[sheetname]
                    sheet2 = wb2.create_sheet(f"第{numtest1}页")
                    print(numtest1)
                    numtest1 += 1
                    task = asyncio.create_task(copy_sheet(sheet, sheet2))
                    tasks.append(task)

                # 等待所有任务完成
                await asyncio.gather(*tasks)

                # 保存并关闭工作簿
                wb2.save(save_path)
                wb.close()
                wb2.close()
                workbook = openpyxl.load_workbook('直梯.xlsx')
                sheet = workbook[f'第{numtest1 - 1}页']
                cell_g4 = sheet['L6']
                cell_g4.value = dict["工程名称"]
                datetime_object = dt.strptime(dict["开工日期"], '%Y-%m-%d').date()
                datetime_object = datetime_object + datetime.timedelta(days=-1)
                cell_g4 = sheet['AS6']
                cell_g4.value = str(datetime_object).split('-')[0] + "年" + str(datetime_object).split('-')[
                    1] + "月" + \
                                str(datetime_object).split('-')[2] + "日"
                workbook.save('直梯.xlsx')
                workbook.close()
                folder_path = r'.\施工方案'
                workbook = openpyxl.load_workbook('直梯.xlsx')
                # 遍历文件夹中的所有文件
                for filename in os.listdir(folder_path):
                    if filename.endswith('.jpg') or filename.endswith('.png'):
                        file_path = os.path.join(folder_path, filename)
                        # 创建一个新的工作表，并使用图片名称作为工作表名称
                        sheet = workbook.create_sheet(f"第{numtest1}页")
                        # 在工作表中插入图片
                        img = Image(file_path)
                        sheet.add_image(img, 'A1')
                        numtest1 = numtest1 + 1

                # 保存 Excel 文件
                workbook.save('直梯.xlsx')
                workbook.close()

            # 运行异步主程序
            asyncio.run(main())
        except Exception as e:
            print(numtest1, e)

        try:
            async def main():
                global numtest1
                path = r"./直梯/14-鲁DT-010技术（安全）交底记录.xlsx"
                save_path = "直梯.xlsx"
                wb = openpyxl.load_workbook(path)
                wb2 = openpyxl.load_workbook(save_path)
                sheetnames = wb.sheetnames

                tasks = []
                for numtest, sheetname in enumerate(sheetnames, start=1):
                    sheet = wb[sheetname]
                    sheet2 = wb2.create_sheet(f"第{numtest1}页")
                    print(numtest1)
                    numtest1 += 1
                    task = asyncio.create_task(copy_sheet(sheet, sheet2))
                    tasks.append(task)

                # 等待所有任务完成
                await asyncio.gather(*tasks)

                # 保存并关闭工作簿
                wb2.save(save_path)
                wb.close()
                wb2.close()
                workbook = openpyxl.load_workbook('直梯.xlsx')
                sheet = workbook[f'第{numtest1 - 1}页']
                cell_g4 = sheet['L6']
                cell_g4.value = dict["工程名称"]
                cell_g4 = sheet['AS6']
                cell_g4.value = dict["分包单位"]
                cell_g4 = sheet['L7']
                cell_g4.value = dict["分项工程名称"]
                datetime_object = dt.strptime(dict["开工日期"], '%Y-%m-%d').date()
                datetime_object = datetime_object + datetime.timedelta(days=-1)
                cell_g4 = sheet['AS7']
                cell_g4.value = str(datetime_object).split('-')[0] + "年" + str(datetime_object).split('-')[
                    1] + "月" + \
                                str(datetime_object).split('-')[2] + "日"
                workbook.save('直梯.xlsx')
                workbook.close()

            # 运行异步主程序
            asyncio.run(main())
        except Exception as e:
            print(numtest1, e)

        try:
            async def main():
                global numtest1
                path = r"./直梯/15-鲁DT-011图纸会审、设计变更、洽商记录汇总表（后附图纸）.xlsx"
                save_path = "直梯.xlsx"
                wb = openpyxl.load_workbook(path)
                wb2 = openpyxl.load_workbook(save_path)
                sheetnames = wb.sheetnames

                tasks = []
                for numtest, sheetname in enumerate(sheetnames, start=1):
                    sheet = wb[sheetname]
                    sheet2 = wb2.create_sheet(f"第{numtest1}页")
                    print(numtest1)
                    numtest1 += 1
                    task = asyncio.create_task(copy_sheet(sheet, sheet2))
                    tasks.append(task)

                # 等待所有任务完成
                await asyncio.gather(*tasks)

                # 保存并关闭工作簿
                wb2.save(save_path)
                wb.close()
                wb2.close()
                workbook = openpyxl.load_workbook('直梯.xlsx')
                sheet = workbook[f'第{numtest1 - 1}页']
                cell_g4 = sheet['L6']
                cell_g4.value = dict["工程名称"]
                cell_g4 = sheet['AW6']
                if len(str(dict["会审日期"]).split('-')) >= 3:
                    cell_g4.value = str(dict["会审日期"]).split('-')[0] + "年" + str(dict["会审日期"]).split('-')[
                        1] + "月" + \
                                    str(dict["会审日期"]).split('-')[2] + "日"

                workbook.save('直梯.xlsx')
                workbook.close()
                folder_path = r'.\图纸'
                workbook = openpyxl.load_workbook('直梯.xlsx')
                # 遍历文件夹中的所有文件
                for filename in os.listdir(folder_path):
                    if filename.endswith('.jpg') or filename.endswith('.png'):
                        file_path = os.path.join(folder_path, filename)
                        # 创建一个新的工作表，并使用图片名称作为工作表名称
                        sheet = workbook.create_sheet(f"第{numtest1}页")
                        # 在工作表中插入图片
                        img = Image(file_path)
                        sheet.add_image(img, 'A1')
                        numtest1 = numtest1 + 1

                # 保存 Excel 文件
                workbook.save('直梯.xlsx')
                workbook.close()

            # 运行异步主程序
            asyncio.run(main())
        except Exception as e:
            print(numtest1, e)

        try:
            async def main():
                global numtest1
                path = r"./直梯/16-鲁DT-012图纸会审记录.xlsx"
                save_path = "直梯.xlsx"
                wb = openpyxl.load_workbook(path)
                wb2 = openpyxl.load_workbook(save_path)
                sheetnames = wb.sheetnames

                tasks = []
                for numtest, sheetname in enumerate(sheetnames, start=1):
                    sheet = wb[sheetname]
                    sheet2 = wb2.create_sheet(f"第{numtest1}页")
                    print(numtest1)
                    numtest1 += 1
                    task = asyncio.create_task(copy_sheet(sheet, sheet2))
                    tasks.append(task)

                # 等待所有任务完成
                await asyncio.gather(*tasks)

                # 保存并关闭工作簿
                wb2.save(save_path)
                wb.close()
                wb2.close()
                workbook = openpyxl.load_workbook('直梯.xlsx')
                sheet = workbook[f'第{numtest1 - 1}页']
                cell_g4 = sheet['AW7']
                if len(str(dict["会审日期"]).split('-')) >= 3:
                    cell_g4.value = str(dict["会审日期"]).split('-')[0] + "年" + str(dict["会审日期"]).split('-')[
                        1] + "月" + \
                                    str(dict["会审日期"]).split('-')[2] + "日"
                cell_g4 = sheet['L6']
                cell_g4.value = dict["工程名称"]
                cell_g4 = sheet['L7']
                cell_g4.value = dict["子分部工程名称"]
                cell_g4 = sheet['B19']
                cell_g4.value = dict["建设单位"]
                cell_g4 = sheet['AH19']
                cell_g4.value = dict["设计单位"]
                cell_g4 = sheet['R19']
                cell_g4.value = dict["监理单位"]
                cell_g4 = sheet['AX19']
                cell_g4.value = dict["分包单位"]
                workbook.save('直梯.xlsx')
                workbook.close()

            # 运行异步主程序
            asyncio.run(main())
        except Exception as e:
            print(numtest1, e)

        try:
            async def main():
                global numtest1
                path = r"./直梯/17-鲁DT-013设计交底记录.xlsx"
                save_path = "直梯.xlsx"
                wb = openpyxl.load_workbook(path)
                wb2 = openpyxl.load_workbook(save_path)
                sheetnames = wb.sheetnames

                tasks = []
                for numtest, sheetname in enumerate(sheetnames, start=1):
                    sheet = wb[sheetname]
                    sheet2 = wb2.create_sheet(f"第{numtest1}页")
                    print(numtest1)
                    numtest1 += 1
                    task = asyncio.create_task(copy_sheet(sheet, sheet2))
                    tasks.append(task)

                # 等待所有任务完成
                await asyncio.gather(*tasks)

                # 保存并关闭工作簿
                wb2.save(save_path)
                wb.close()
                wb2.close()
                workbook = openpyxl.load_workbook('直梯.xlsx')
                sheet = workbook[f'第{numtest1 - 1}页']
                if len(str(dict["开工日期"]).split('-')) >= 3:
                    datetime_object = dt.strptime(dict["开工日期"], '%Y-%m-%d').date()
                    datetime_object = datetime_object + datetime.timedelta(days=-1)
                    cell_g4 = sheet['AW7']
                    cell_g4.value = str(datetime_object).split('-')[0] + "年" + str(datetime_object).split('-')[
                        1] + "月" + \
                                    str(datetime_object).split('-')[2] + "日"
                cell_g4 = sheet['L6']
                cell_g4.value = dict["工程名称"]
                cell_g4 = sheet['L7']
                cell_g4.value = dict["子分部工程名称"]
                cell_g4 = sheet['L8']
                cell_g4.value = dict["交底人"]
                cell_g4 = sheet['AW8']
                cell_g4.value = dict["交底地点"]
                cell_g4 = sheet['B28']
                cell_g4.value = dict["建设单位"]
                cell_g4 = sheet['R28']
                cell_g4.value = dict["监理单位"]
                cell_g4 = sheet['AH28']
                cell_g4.value = dict["设计单位"]
                cell_g4 = sheet['AX28']
                cell_g4.value = dict["分包单位"]
                workbook.save('直梯.xlsx')
                workbook.close()

            # 运行异步主程序
            asyncio.run(main())
        except Exception as e:
            print(numtest1, e)

        try:
            async def main():
                global numtest1
                path = r"./直梯/18-鲁DT-014设计变更通知单.xlsx"
                save_path = "直梯.xlsx"
                wb = openpyxl.load_workbook(path)
                wb2 = openpyxl.load_workbook(save_path)
                sheetnames = wb.sheetnames

                tasks = []
                for numtest, sheetname in enumerate(sheetnames, start=1):
                    sheet = wb[sheetname]
                    sheet2 = wb2.create_sheet(f"第{numtest1}页")
                    print(numtest1)
                    numtest1 += 1
                    task = asyncio.create_task(copy_sheet(sheet, sheet2))
                    tasks.append(task)

                # 等待所有任务完成
                await asyncio.gather(*tasks)

                # 保存并关闭工作簿
                wb2.save(save_path)
                wb.close()
                wb2.close()
                workbook = openpyxl.load_workbook('直梯.xlsx')
                sheet = workbook[f'第{numtest1 - 1}页']
                cell_g4 = sheet['AW7']
                if len(str(dict["变更日期"]).split('-')) >= 3:
                    cell_g4.value = str(dict["变更日期"]).split('-')[0] + "年" + str(dict["变更日期"]).split('-')[
                        1] + "月" + \
                                    str(dict["变更日期"]).split('-')[2] + "日"
                cell_g4 = sheet['L6']
                cell_g4.value = dict["工程名称"]
                cell_g4 = sheet['AW6']
                cell_g4.value = dict["子分部工程名称"]
                cell_g4 = sheet['L7']
                cell_g4.value = dict["设计单位"]
                cell_g4 = sheet['B18']
                cell_g4.value = dict["建设单位"]
                cell_g4 = sheet['R18']
                cell_g4.value = dict["监理单位"]
                cell_g4 = sheet['AH18']
                cell_g4.value = dict["设计单位"]
                cell_g4 = sheet['AX18']
                cell_g4.value = dict["分包单位"]
                workbook.save('直梯.xlsx')
                workbook.close()

            # 运行异步主程序
            asyncio.run(main())
        except Exception as e:
            print(numtest1, e)

        try:
            async def main():
                global numtest1
                path = r"./直梯/19-鲁DT-015工程洽商记录.xlsx"
                save_path = "直梯.xlsx"
                wb = openpyxl.load_workbook(path)
                wb2 = openpyxl.load_workbook(save_path)
                sheetnames = wb.sheetnames

                tasks = []
                for numtest, sheetname in enumerate(sheetnames, start=1):
                    sheet = wb[sheetname]
                    sheet2 = wb2.create_sheet(f"第{numtest1}页")
                    print(numtest1)
                    numtest1 += 1
                    task = asyncio.create_task(copy_sheet(sheet, sheet2))
                    tasks.append(task)

                # 等待所有任务完成
                await asyncio.gather(*tasks)

                # 保存并关闭工作簿
                wb2.save(save_path)
                wb.close()
                wb2.close()
                workbook = openpyxl.load_workbook('直梯.xlsx')
                sheet = workbook[f'第{numtest1 - 1}页']
                datetime_object = dt.strptime(dict["开工日期"], '%Y-%m-%d').date()
                datetime_object = datetime_object + datetime.timedelta(days=-1)
                cell_g4 = sheet['AW7']
                cell_g4.value = str(datetime_object).split('-')[0] + "年" + str(datetime_object).split('-')[
                    1] + "月" + \
                                str(datetime_object).split('-')[2] + "日"
                cell_g4 = sheet['L6']
                cell_g4.value = dict["工程名称"]
                cell_g4 = sheet['AW6']
                cell_g4.value = dict["子分部工程名称"]
                cell_g4 = sheet['L7']
                cell_g4.value = dict["分包单位"]
                cell_g4 = sheet['B19']
                cell_g4.value = dict["建设单位"]
                cell_g4 = sheet['R19']
                cell_g4.value = dict["监理单位"]
                cell_g4 = sheet['AH19']
                cell_g4.value = dict["设计单位"]
                cell_g4 = sheet['AX19']
                cell_g4.value = dict["分包单位"]
                workbook.save('直梯.xlsx')
                workbook.close()

            # 运行异步主程序
            asyncio.run(main())
        except Exception as e:
            print(numtest1, e)

        try:
            async def main():
                global numtest1
                path = r"./直梯/20-表B.0.6 工程材料、构配件、设备报审表.xlsx"
                save_path = "直梯.xlsx"
                wb = openpyxl.load_workbook(path)
                wb2 = openpyxl.load_workbook(save_path)
                sheetnames = wb.sheetnames

                tasks = []
                for numtest, sheetname in enumerate(sheetnames, start=1):
                    sheet = wb[sheetname]
                    sheet2 = wb2.create_sheet(f"第{numtest1}页")
                    print(numtest1)
                    numtest1 += 1
                    task = asyncio.create_task(copy_sheet(sheet, sheet2))
                    tasks.append(task)

                # 等待所有任务完成
                await asyncio.gather(*tasks)

                # 保存并关闭工作簿
                wb2.save(save_path)
                wb.close()
                wb2.close()
                workbook = openpyxl.load_workbook('直梯.xlsx')
                sheet = workbook[f'第{numtest1 - 1}页']
                if len(str(dict["设备报审日期"]).split('-')) >= 3:
                    cell_g4 = sheet['G6']
                    cell_g4.value = str(dict["设备报审日期"]).split('-')[0]
                    cell_g4 = sheet['M6']
                    cell_g4.value = str(dict["设备报审日期"]).split('-')[1]
                    cell_g4 = sheet['Q6']
                    cell_g4.value = str(dict["设备报审日期"]).split('-')[2]
                cell_g4 = sheet['G3']
                cell_g4.value = dict["工程名称"]
                cell_g4 = sheet['E5']
                cell_g4.value = dict["监理单位"]
                cell_g4 = sheet['AB26']
                cell_g4.value = dict["监理单位"]
                workbook.save('直梯.xlsx')
                workbook.close()

            # 运行异步主程序
            asyncio.run(main())
        except Exception as e:
            print(numtest1, e)

        try:
            async def main():
                global numtest1
                path = r"./直梯/21-鲁DT-016材料、构配件进场检验记录（后附型式试验）.xlsx"
                save_path = "直梯.xlsx"
                wb = openpyxl.load_workbook(path)
                wb2 = openpyxl.load_workbook(save_path)
                sheetnames = wb.sheetnames

                tasks = []
                for numtest, sheetname in enumerate(sheetnames, start=1):
                    sheet = wb[sheetname]
                    sheet2 = wb2.create_sheet(f"第{numtest1}页")
                    print(numtest1)
                    numtest1 += 1
                    task = asyncio.create_task(copy_sheet(sheet, sheet2))
                    tasks.append(task)

                # 等待所有任务完成
                await asyncio.gather(*tasks)

                # 保存并关闭工作簿
                wb2.save(save_path)
                wb.close()
                wb2.close()
                workbook = openpyxl.load_workbook('直梯.xlsx')
                sheet = workbook[f'第{numtest1 - 1}页']
                if len(str(dict["检验日期"]).split('-')) >= 3:
                    cell_g4 = sheet['AY6']
                    cell_g4.value = str(dict["检验日期"]).split('-')[0] + "年" + str(dict["检验日期"]).split('-')[
                        1] + "月" + \
                                    str(dict["检验日期"]).split('-')[2] + "日"
                cell_g4 = sheet['L6']
                cell_g4.value = dict["工程名称"]
                exls = pd.read_excel(r'./测量值/合格证信息.xlsx', keep_default_na=False)
                testDemo = 1
                demoNum = 9
                for data in exls.values:
                    cell_g4 = sheet[f'B{demoNum}']
                    cell_g4.value = data[0]
                    cell_g4 = sheet[f'E{demoNum}']
                    cell_g4.value = data[1]
                    cell_g4 = sheet[f'V{demoNum}']
                    cell_g4.value = data[2]
                    cell_g4 = sheet[f'AB{demoNum}']
                    cell_g4.value = data[3]
                    cell_g4 = sheet[f'AH{demoNum}']
                    cell_g4.value = data[4]
                    demoNum = demoNum + 1
                    testDemo = testDemo + 1

                    if testDemo > 13:
                        workbook.save('直梯.xlsx')
                        workbook.close()
                        testDemo = 1
                        demoNum = 9
                        path = r"./直梯/21-鲁DT-016材料、构配件进场检验记录（后附型式试验）.xlsx"
                        save_path = "直梯.xlsx"
                        wb = openpyxl.load_workbook(path)
                        wb2 = openpyxl.load_workbook(save_path)
                        sheetnames = wb.sheetnames
                        for sheetname in sheetnames:
                            print(f"第{numtest1}页")
                            sheet = wb[sheetname]
                            sheet2 = wb2.create_sheet(f"第{numtest1}页")
                            # Tab color
                            sheet2.sheet_properties.tabColor = sheet.sheet_properties.tabColor

                            # 复制列宽度
                            for column in sheet.column_dimensions:
                                sheet2.column_dimensions[column].width = sheet.column_dimensions[column].width

                            # 复制行高
                            for row in sheet.row_dimensions:
                                sheet2.row_dimensions[row].height = sheet.row_dimensions[row].height

                            # 处理合并单元格
                            wm = list(sheet.merged_cells)
                            if len(wm) > 0:
                                for i in range(0, len(wm)):
                                    cell2 = str(wm[i]).replace('(<CellRange ', '').replace('>,)', '')
                                    sheet2.merge_cells(cell2)

                            # 复制行、列和单元格值
                            for i, row in enumerate(sheet.iter_rows()):
                                for j, cell in enumerate(row):
                                    # 复制单元格值
                                    sheet2.cell(row=i + 1, column=j + 1, value=cell.value)

                                    # 如果是合并单元格的一部分，则设置单元格格式和尺寸
                                    if cell.coordinate in sheet.merged_cells:
                                        for merged_cell in wm:
                                            if cell.coordinate in merged_cell:
                                                first_cell = sheet.cell(merged_cell.min_row, merged_cell.min_col)
                                                target_cell = sheet2.cell(i + 1, j + 1)
                                                sheet2.row_dimensions[i + 1].height = sheet.row_dimensions[
                                                    first_cell.row].height
                                                sheet2.column_dimensions[get_column_letter(j + 1)].width = \
                                                    sheet.column_dimensions[
                                                        get_column_letter(first_cell.column)].width
                                                break

                                    # 设置单元格格式
                                    source_cell = sheet.cell(i + 1, j + 1)
                                    target_cell = sheet2.cell(i + 1, j + 1)

                                    # 创建新样式对象
                                    font = Font(name=source_cell.font.name, size=source_cell.font.size,
                                                bold=source_cell.font.bold, italic=source_cell.font.italic,
                                                vertAlign=source_cell.font.vertAlign,
                                                underline=source_cell.font.underline,
                                                strike=source_cell.font.strike,
                                                color='FF000000')  # Set font color to black

                                    black_side = Side(color='FF000000')
                                    border = Border(
                                        left=Side(color=black_side.color,
                                                  border_style=source_cell.border.left.border_style),
                                        right=Side(color=black_side.color,
                                                   border_style=source_cell.border.right.border_style),
                                        top=Side(color=black_side.color,
                                                 border_style=source_cell.border.top.border_style),
                                        bottom=Side(color=black_side.color,
                                                    border_style=source_cell.border.bottom.border_style))

                                    fill = PatternFill(fill_type=source_cell.fill.fill_type,
                                                       fgColor=source_cell.fill.fgColor,
                                                       bgColor=source_cell.fill.bgColor,
                                                       patternType=source_cell.fill.patternType)

                                    alignment = Alignment(horizontal=source_cell.alignment.horizontal,
                                                          vertical=source_cell.alignment.vertical,
                                                          text_rotation=source_cell.alignment.text_rotation,
                                                          wrap_text=source_cell.alignment.wrap_text,
                                                          shrink_to_fit=source_cell.alignment.shrink_to_fit,
                                                          indent=source_cell.alignment.indent)

                                    protection = Protection(locked=source_cell.protection.locked,
                                                            hidden=source_cell.protection.hidden)

                                    # 将新样式对象分配给目标单元格
                                    target_cell.font = font
                                    target_cell.border = border
                                    target_cell.fill = fill
                                    target_cell.alignment = alignment
                                    target_cell.protection = protection
                                    target_cell.number_format = source_cell.number_format
                            numtest1 = numtest1 + 1
                        # 保存并关闭工作簿
                        wb2.save(save_path)
                        wb.close()
                        wb2.close()
                        workbook = openpyxl.load_workbook('直梯.xlsx')
                        sheet = workbook[f'第{numtest1 - 1}页']
                        if len(str(dict["检验日期"]).split('-')) >= 3:
                            cell_g4 = sheet['AY6']
                            cell_g4.value = str(dict["检验日期"]).split('-')[0] + "年" + \
                                            str(dict["检验日期"]).split('-')[
                                                1] + "月" + \
                                            str(dict["检验日期"]).split('-')[2] + "日"
                        cell_g4 = sheet['L6']
                        cell_g4.value = dict["工程名称"]
                workbook.save('直梯.xlsx')
                workbook.close()

                folder_path = r'.\型式试验'
                workbook = openpyxl.load_workbook('直梯.xlsx')
                # 遍历文件夹中的所有文件
                for filename in os.listdir(folder_path):
                    if filename.endswith('.jpg') or filename.endswith('.png'):
                        file_path = os.path.join(folder_path, filename)
                        # 创建一个新的工作表，并使用图片名称作为工作表名称
                        sheet = workbook.create_sheet(f"第{numtest1}页")
                        # 在工作表中插入图片
                        img = Image(file_path)
                        sheet.add_image(img, 'A1')
                        numtest1 = numtest1 + 1

                # 保存 Excel 文件
                workbook.save('直梯.xlsx')
                workbook.close()
                workbook.close()

            # 运行异步主程序
            asyncio.run(main())
        except Exception as e:
            print(numtest1, e)

        exls = pd.read_excel(r'./测量值/电梯竣工资料信息表.xlsx', sheet_name='电梯配置信息', keep_default_na=False)
        datas_exls = exls.iloc[6:, 1:]

        try:
            for num in range(0, dict['电梯数量']):
                first_column = datas_exls.columns[num]
                dict["设备梯号"] = datas_exls[first_column].iloc[0]
                dict["装箱单号"] = datas_exls[first_column].iloc[0]
                dict["电梯型号规格"] = datas_exls[first_column].iloc[29]

                async def main():
                    global numtest1
                    path = r"./直梯/22-鲁DT-017设备（开箱）进场检验记录（每台一份）.xlsx"
                    save_path = "直梯.xlsx"
                    wb = openpyxl.load_workbook(path)
                    wb2 = openpyxl.load_workbook(save_path)
                    sheetnames = wb.sheetnames

                    tasks = []
                    for numtest, sheetname in enumerate(sheetnames, start=1):
                        sheet = wb[sheetname]
                        sheet2 = wb2.create_sheet(f"第{numtest1}页")
                        print(numtest1)
                        numtest1 += 1
                        task = asyncio.create_task(copy_sheet(sheet, sheet2))
                        tasks.append(task)

                    # 等待所有任务完成
                    await asyncio.gather(*tasks)

                    # 保存并关闭工作簿
                    wb2.save(save_path)
                    wb.close()
                    wb2.close()
                    workbook = openpyxl.load_workbook('直梯.xlsx')
                    sheet = workbook[f'第{numtest1 - 1}页']
                    cell_g4 = sheet['BF4']
                    cell_g4.value = num
                    cell_g4 = sheet['N6']
                    cell_g4.value = dict["工程名称"]
                    cell_g4 = sheet['AS6']
                    cell_g4.value = dict["建设单位"]
                    cell_g4 = sheet['N7']
                    cell_g4.value = dict["设备梯号"]
                    cell_g4 = sheet['AS7']
                    cell_g4.value = dict["监理单位"]
                    cell_g4 = sheet['N8']
                    cell_g4.value = dict["装箱单号"]
                    cell_g4 = sheet['AS8']
                    cell_g4.value = dict["分包单位"]
                    cell_g4 = sheet['N9']
                    cell_g4.value = dict["电梯型号规格"]
                    cell_g4 = sheet['AS9']
                    cell_g4.value = dict["供货商"]
                    workbook.save('直梯.xlsx')
                    workbook.close()

                # 运行异步主程序
                asyncio.run(main())
        except Exception as e:
            print(numtest1, e)

        try:
            async def main():
                global numtest1
                path = r"./直梯/23-鲁DT-018材料合格证、复试报告汇总表（后附合格证）.xlsx"
                save_path = "直梯.xlsx"
                wb = openpyxl.load_workbook(path)
                wb2 = openpyxl.load_workbook(save_path)
                sheetnames = wb.sheetnames

                tasks = []
                for numtest, sheetname in enumerate(sheetnames, start=1):
                    sheet = wb[sheetname]
                    sheet2 = wb2.create_sheet(f"第{numtest1}页")
                    print(numtest1)
                    numtest1 += 1
                    task = asyncio.create_task(copy_sheet(sheet, sheet2))
                    tasks.append(task)

                # 等待所有任务完成
                await asyncio.gather(*tasks)

                # 保存并关闭工作簿
                wb2.save(save_path)
                wb.close()
                wb2.close()
                workbook = openpyxl.load_workbook('直梯.xlsx')
                sheet = workbook[f'第{numtest1 - 1}页']
                cell_g4 = sheet['I4']
                cell_g4.value = dict["工程名称"]
                workbook.save('直梯.xlsx')
                workbook.close()
                workbook.close()

                folder_path = r'.\合格证'
                workbook = openpyxl.load_workbook('直梯.xlsx')
                # 遍历文件夹中的所有文件
                for filename in os.listdir(folder_path):
                    if filename.endswith('.jpg') or filename.endswith('.png'):
                        file_path = os.path.join(folder_path, filename)
                        # 创建一个新的工作表，并使用图片名称作为工作表名称
                        sheet = workbook.create_sheet(f"第{numtest1}页")
                        # 在工作表中插入图片
                        img = Image(file_path)
                        sheet.add_image(img, 'A1')
                        numtest1 = numtest1 + 1

                # 保存 Excel 文件
                workbook.save('直梯.xlsx')
                workbook.close()
                workbook.close()

            # 运行异步主程序
            asyncio.run(main())
        except Exception as e:
            print(numtest1, e)

        try:
            async def main():
                global numtest1
                path = r"./直梯/24-鲁DT-019合格证【复印件（或抄件）】贴条.xlsx"
                save_path = "直梯.xlsx"
                wb = openpyxl.load_workbook(path)
                wb2 = openpyxl.load_workbook(save_path)
                sheetnames = wb.sheetnames

                tasks = []
                for numtest, sheetname in enumerate(sheetnames, start=1):
                    sheet = wb[sheetname]
                    sheet2 = wb2.create_sheet(f"第{numtest1}页")
                    print(numtest1)
                    numtest1 += 1
                    task = asyncio.create_task(copy_sheet(sheet, sheet2))
                    tasks.append(task)

                # 等待所有任务完成
                await asyncio.gather(*tasks)

                # 保存并关闭工作簿
                wb2.save(save_path)
                wb.close()
                wb2.close()
                workbook = openpyxl.load_workbook('直梯.xlsx')
                sheet = workbook[f'第{numtest1 - 1}页']
                cell_g4 = sheet['AH15']
                if len(str(dict["到货日期"]).split('-')) >= 3:
                    cell_g4.value = str(dict["到货日期"]).split('-')[0] + "年" + str(dict["到货日期"]).split('-')[
                        1] + "月" + \
                                    str(dict["到货日期"]).split('-')[2] + "日"

                cell_g4 = sheet['AH6']
                cell_g4.value = dict["材料名称"]
                cell_g4 = sheet['AH7']
                cell_g4.value = dict["合格证（原件）编号"]
                cell_g4 = sheet['AH9']
                cell_g4.value = dict["进货数量"]
                cell_g4 = sheet['AH10']
                cell_g4.value = dict["工程总需要量"]
                cell_g4 = sheet['AH14']
                cell_g4.value = dict["供货单位"]
                cell_g4 = sheet['AH18']
                cell_g4.value = dict["合格证原件存放单位"]
                workbook.save('直梯.xlsx')
                workbook.close()

            # 运行异步主程序
            asyncio.run(main())
        except Exception as e:
            print(numtest1, e)

        try:
            async def main():
                global numtest1
                path = r"./直梯/25-鲁DT-020材料见证取样检测汇总表.xlsx"
                save_path = "直梯.xlsx"
                wb = openpyxl.load_workbook(path)
                wb2 = openpyxl.load_workbook(save_path)
                sheetnames = wb.sheetnames

                tasks = []
                for numtest, sheetname in enumerate(sheetnames, start=1):
                    sheet = wb[sheetname]
                    sheet2 = wb2.create_sheet(f"第{numtest1}页")
                    print(numtest1)
                    numtest1 += 1
                    task = asyncio.create_task(copy_sheet(sheet, sheet2))
                    tasks.append(task)

                # 等待所有任务完成
                await asyncio.gather(*tasks)

                # 保存并关闭工作簿
                wb2.save(save_path)
                wb.close()
                wb2.close()
                workbook = openpyxl.load_workbook('直梯.xlsx')
                sheet = workbook[f'第{numtest1 - 1}页']
                cell_g4 = sheet['O6']
                cell_g4.value = dict["工程名称"]
                cell_g4 = sheet['O7']
                cell_g4.value = dict["监理单位"]
                cell_g4 = sheet['AU7']
                cell_g4.value = dict["分包单位"]
                workbook.save('直梯.xlsx')
                workbook.close()

            # 运行异步主程序
            asyncio.run(main())
        except Exception as e:
            print(numtest1, e)

        try:
            async def main():
                global numtest1
                path = r"./直梯/26-鲁DT-021取样送样试验见证记录.xlsx"
                save_path = "直梯.xlsx"
                wb = openpyxl.load_workbook(path)
                wb2 = openpyxl.load_workbook(save_path)
                sheetnames = wb.sheetnames

                tasks = []
                for numtest, sheetname in enumerate(sheetnames, start=1):
                    sheet = wb[sheetname]
                    sheet2 = wb2.create_sheet(f"第{numtest1}页")
                    print(numtest1)
                    numtest1 += 1
                    task = asyncio.create_task(copy_sheet(sheet, sheet2))
                    tasks.append(task)

                # 等待所有任务完成
                await asyncio.gather(*tasks)

                # 保存并关闭工作簿
                wb2.save(save_path)
                wb.close()
                wb2.close()
                workbook = openpyxl.load_workbook('直梯.xlsx')
                sheet = workbook[f'第{numtest1 - 1}页']
                cell_g4 = sheet['AU8']
                if len(str(dict["取样日期"]).split('-')) >= 3:
                    cell_g4.value = str(dict["取样日期"]).split('-')[0] + "年" + str(dict["取样日期"]).split('-')[
                        1] + "月" + \
                                    str(dict["取样日期"]).split('-')[2] + "日"

                cell_g4 = sheet['O6']
                cell_g4.value = dict["工程名称"]
                cell_g4 = sheet['AU6']
                cell_g4.value = dict["取样部位"]
                cell_g4 = sheet['O7']
                cell_g4.value = dict["样品名称"]
                cell_g4 = sheet['AU7']
                cell_g4.value = dict["取样数量"]
                cell_g4 = sheet['O8']
                cell_g4.value = dict["取样地点"]
                cell_g4 = sheet['O9']
                cell_g4.value = dict["执行标准、规范"]
                cell_g4 = sheet['O10']
                cell_g4.value = dict["试验项目"]
                workbook.save('直梯.xlsx')
                workbook.close()

            # 运行异步主程序
            asyncio.run(main())
        except Exception as e:
            print(numtest1, e)

        try:
            for num in range(0, dict['电梯数量']):
                first_column = datas_exls.columns[num]
                dict["检查部位"] = datas_exls[first_column].iloc[56]
                dict["隐检项目"] = datas_exls[first_column].iloc[57]
                date_value = datas_exls[first_column].iloc[58]
                if date_value is not None:
                    # 将日期格式化为字符串，例如"2023-01-13"
                    date_str = date_value.strftime("%Y-%m-%d")
                    # 将格式化后的日期字符串存储到字典中
                    dict["隐检日期"] = date_str
                else:
                    dict["隐检日期"] = ""
                dict["隐检部位"] = datas_exls[first_column].iloc[56]
                dict["施工图图号"] = datas_exls[first_column].iloc[55]
                dict["隐检材料名称及规格/型号"] = datas_exls[first_column].iloc[59]

                async def main():
                    global numtest1
                    path = r"./直梯/27-鲁DT-022隐蔽工程验收记录（每台一份）.xlsx"
                    save_path = "直梯.xlsx"
                    wb = openpyxl.load_workbook(path)
                    wb2 = openpyxl.load_workbook(save_path)
                    sheetnames = wb.sheetnames

                    tasks = []
                    for numtest, sheetname in enumerate(sheetnames, start=1):
                        sheet = wb[sheetname]
                        sheet2 = wb2.create_sheet(f"第{numtest1}页")
                        print(numtest1)
                        numtest1 += 1
                        task = asyncio.create_task(copy_sheet(sheet, sheet2))
                        tasks.append(task)

                    # 等待所有任务完成
                    await asyncio.gather(*tasks)

                    # 保存并关闭工作簿
                    wb2.save(save_path)
                    wb.close()
                    wb2.close()
                    workbook = openpyxl.load_workbook('直梯.xlsx')
                    sheet = workbook[f'第{numtest1 - 1}页']
                    cell_g4 = sheet['BF4']
                    cell_g4.value = num
                    cell_g4 = sheet['N6']
                    cell_g4.value = dict["工程名称"]
                    cell_g4 = sheet['AS6']
                    cell_g4.value = dict["检查部位"]
                    cell_g4 = sheet['N7']
                    cell_g4.value = dict["隐检项目"]
                    if len(str(dict["隐检日期"]).split('-')) >= 3:
                        cell_g4 = sheet['AS7']
                        cell_g4.value = str(dict["隐检日期"]).split('-')[0] + "年" + \
                                        str(dict["隐检日期"]).split('-')[
                                            1] + "月" + \
                                        str(dict["隐检日期"]).split('-')[2] + "日"
                    cell_g4 = sheet['N8']
                    cell_g4.value = dict["隐检部位"]
                    cell_g4 = sheet['Q9']
                    cell_g4.value = dict["施工图图号"]
                    cell_g4 = sheet['T11']
                    cell_g4.value = dict["隐检材料名称及规格/型号"]
                    cell_g4 = sheet['Q23']
                    cell_g4.value = dict["监理单位"]
                    cell_g4 = sheet['AQ23']
                    cell_g4.value = dict["分包单位"]
                    cell_g4 = sheet['Q25']
                    cell_g4.value = dict["总监理工程师"]
                    cell_g4 = sheet['AW25']
                    cell_g4.value = dict["分包单位技术负责人"]
                    workbook.save('直梯.xlsx')
                    workbook.close()

                # 运行异步主程序
                asyncio.run(main())
        except Exception as e:
            print(numtest1, e)

        try:
            async def main():
                global numtest1
                path = r"./直梯/28-鲁DT-023施工检查记录.xlsx"
                save_path = "直梯.xlsx"
                wb = openpyxl.load_workbook(path)
                wb2 = openpyxl.load_workbook(save_path)
                sheetnames = wb.sheetnames

                tasks = []
                for numtest, sheetname in enumerate(sheetnames, start=1):
                    sheet = wb[sheetname]
                    sheet2 = wb2.create_sheet(f"第{numtest1}页")
                    print(numtest1)
                    numtest1 += 1
                    task = asyncio.create_task(copy_sheet(sheet, sheet2))
                    tasks.append(task)

                # 等待所有任务完成
                await asyncio.gather(*tasks)

                # 保存并关闭工作簿
                wb2.save(save_path)
                wb.close()
                wb2.close()
                workbook = openpyxl.load_workbook('直梯.xlsx')
                sheet = workbook[f'第{numtest1 - 1}页']
                if len(str(dict["检查日期"]).split('-')) >= 3:
                    cell_g4 = sheet['AQ7']
                    cell_g4.value = str(dict["检查日期"]).split('-')[0] + "年" + str(dict["检查日期"]).split('-')[
                        1] + "月" + \
                                    str(dict["检查日期"]).split('-')[2] + "日"
                cell_g4 = sheet['M6']
                cell_g4.value = dict["工程名称"]
                cell_g4 = sheet['AQ6']
                cell_g4.value = dict["检查项目"]
                cell_g4 = sheet['M7']
                cell_g4.value = dict["检查部位"]
                cell_g4 = sheet['Z21']
                cell_g4.value = dict["分包单位"]
                workbook.save('直梯.xlsx')
                workbook.close()

            # 运行异步主程序
            asyncio.run(main())
        except Exception as e:
            print(numtest1, e)

        try:
            for num in range(0, dict['电梯数量']):
                first_column = datas_exls.columns[num]
                dict["电梯厂设计图号"] = datas_exls[first_column].iloc[0]
                dict["同机房电梯数"] = datas_exls[first_column].iloc[31]
                date_value = datas_exls[first_column].iloc[38]
                if date_value is not None:
                    # 将日期格式化为字符串，例如"2023-01-13"
                    date_str = date_value.strftime("%Y-%m-%d")
                    # 将格式化后的日期字符串存储到字典中
                    dict["井道测量检查日期"] = date_str
                else:
                    dict["井道测量检查日期"] = ""
                dict["楼层数"] = datas_exls[first_column].iloc[8]
                dict["同机房电梯数"] = datas_exls[first_column].iloc[31]
                dict["同井道电梯数"] = datas_exls[first_column].iloc[32]
                dict["土建设计图号"] = datas_exls[first_column].iloc[33]

                async def main():
                    global numtest1
                    path = r"./直梯/29-鲁DT-024电梯机房、井道测量交接检查记录（每台一份）.xlsx"
                    save_path = "直梯.xlsx"
                    wb = openpyxl.load_workbook(path)
                    wb2 = openpyxl.load_workbook(save_path)
                    sheetnames = wb.sheetnames

                    tasks = []
                    for numtest, sheetname in enumerate(sheetnames, start=1):
                        sheet = wb[sheetname]
                        sheet2 = wb2.create_sheet(f"第{numtest1}页")
                        print(numtest1)
                        numtest1 += 1
                        task = asyncio.create_task(copy_sheet(sheet, sheet2))
                        tasks.append(task)

                    # 等待所有任务完成
                    await asyncio.gather(*tasks)

                    # 保存并关闭工作簿
                    wb2.save(save_path)
                    wb.close()
                    wb2.close()
                    workbook = openpyxl.load_workbook('直梯.xlsx')
                    sheet = workbook[f'第{numtest1 - 1}页']
                    if len(str(dict["井道测量检查日期"]).split('-')) >= 3:
                        cell_g4 = sheet['AY9']
                        cell_g4.value = str(dict["井道测量检查日期"]).split('-')[0] + "年" + \
                                        str(dict["井道测量检查日期"]).split('-')[1] + "月" + \
                                        str(dict["井道测量检查日期"]).split('-')[2] + "日"
                    cell_g4 = sheet['L6']
                    cell_g4.value = dict["工程名称"]
                    cell_g4 = sheet['BF4']
                    cell_g4.value = num
                    cell_g4 = sheet['L7']
                    cell_g4.value = dict["土建设计图号"]
                    cell_g4 = sheet['AY7']
                    cell_g4.value = dict["电梯厂设计图号"]
                    cell_g4 = sheet['L8']
                    cell_g4.value = dict["同机房电梯数"]
                    cell_g4 = sheet['AY8']
                    cell_g4.value = dict["楼层数"]
                    cell_g4 = sheet['L9']
                    cell_g4.value = dict["同井道电梯数"]
                    exls1 = pd.read_excel(r'./测量值/测量数据.xlsx', sheet_name=f'Sheet{num + 1}',
                                          keep_default_na=False)
                    datanum = 11
                    for data in exls1.values:
                        cell_g4 = sheet[f'L{datanum}']
                        cell_g4.value = data[2]
                        cell_g4 = sheet[f'Y{datanum}']
                        cell_g4.value = data[3]
                        cell_g4 = sheet[f'AL{datanum}']
                        cell_g4.value = data[4]
                        cell_g4 = sheet[f'AY{datanum}']
                        cell_g4.value = data[5]
                        datanum = datanum + 1

                    workbook.save('直梯.xlsx')
                    workbook.close()

                # 运行异步主程序
                asyncio.run(main())
        except Exception as e:
            print(numtest1, e)

        try:
            for num in range(0, dict['电梯数量']):
                first_column = datas_exls.columns[num]
                dict["设备梯号"] = datas_exls[first_column].iloc[0]
                date_value = datas_exls[first_column].iloc[41]
                if date_value is not None:
                    # 将日期格式化为字符串，例如"2023-01-13"
                    date_str = date_value.strftime("%Y-%m-%d")
                    # 将格式化后的日期字符串存储到字典中
                    dict["放线日期"] = date_str
                else:
                    dict["放线日期"] = ""

                async def main():
                    global numtest1
                    path = r"./直梯/30-鲁DT-025电梯安装样板放线记录（每台一份）.xlsx"
                    save_path = "直梯.xlsx"
                    wb = openpyxl.load_workbook(path)
                    wb2 = openpyxl.load_workbook(save_path)
                    sheetnames = wb.sheetnames

                    tasks = []
                    for numtest, sheetname in enumerate(sheetnames, start=1):
                        sheet = wb[sheetname]
                        sheet2 = wb2.create_sheet(f"第{numtest1}页")
                        print(numtest1)
                        numtest1 += 1
                        task = asyncio.create_task(copy_sheet(sheet, sheet2))
                        tasks.append(task)

                    # 等待所有任务完成
                    await asyncio.gather(*tasks)

                    # 保存并关闭工作簿
                    wb2.save(save_path)
                    wb.close()
                    wb2.close()
                    workbook = openpyxl.load_workbook('直梯.xlsx')
                    sheet = workbook[f'第{numtest1 - 1}页']
                    cell_g4 = sheet['BF4']
                    cell_g4.value = num
                    if len(str(dict["放线日期"]).split('-')) >= 3:
                        cell_g4 = sheet['AQ7']
                        cell_g4.value = str(dict["放线日期"]).split('-')[0] + "年" + \
                                        str(dict["放线日期"]).split('-')[1] + "月" + \
                                        str(dict["放线日期"]).split('-')[2] + "日"

                    cell_g4 = sheet['L6']
                    cell_g4.value = dict["工程名称"]
                    cell_g4 = sheet['AQ6']
                    cell_g4.value = dict["建设单位"]
                    cell_g4 = sheet['L7']
                    cell_g4.value = dict["设备梯号"]
                    cell_g4 = sheet['R25']
                    cell_g4.value = dict["分包单位"]

                    workbook.save('直梯.xlsx')
                    workbook.close()

                # 运行异步主程序
                asyncio.run(main())
        except Exception as e:
            print(numtest1, e)

        try:
            for num in range(0, dict['电梯数量']):
                first_column = datas_exls.columns[num]
                dict["检查区域"] = datas_exls[first_column].iloc[60]
                dict["有机房/无机房"] = datas_exls[first_column].iloc[5]
                date_value = datas_exls[first_column].iloc[42]
                if date_value is not None:
                    # 将日期格式化为字符串，例如"2023-01-13"
                    date_str = date_value.strftime("%Y-%m-%d")
                    # 将格式化后的日期字符串存储到字典中
                    dict["电气检查日期"] = date_str
                else:
                    dict["电气检查日期"] = ""

                async def main():
                    global numtest1
                    path = r"./直梯/31-鲁DT-026.1电梯电气装置安装检查记录（一）（每台一份）.xlsx"
                    save_path = "直梯.xlsx"
                    wb = openpyxl.load_workbook(path)
                    wb2 = openpyxl.load_workbook(save_path)
                    sheetnames = wb.sheetnames
                    tasks = []
                    for numtest, sheetname in enumerate(sheetnames, start=1):
                        sheet = wb[sheetname]
                        sheet2 = wb2.create_sheet(f"第{numtest1}页")
                        print(numtest1)
                        numtest1 += 1
                        task = asyncio.create_task(copy_sheet(sheet, sheet2))
                        tasks.append(task)

                    # 等待所有任务完成
                    await asyncio.gather(*tasks)

                    # 保存并关闭工作簿
                    wb2.save(save_path)
                    wb.close()
                    wb2.close()
                    workbook = openpyxl.load_workbook('直梯.xlsx')
                    sheet = workbook[f'第{numtest1 - 1}页']
                    if len(str(dict["电气检查日期"]).split('-')) >= 3:
                        cell_g4 = sheet['AQ7']
                        cell_g4.value = str(dict["电气检查日期"]).split('-')[0] + "年" + \
                                        str(dict["电气检查日期"]).split('-')[
                                            1] + "月" + \
                                        str(dict["电气检查日期"]).split('-')[2] + "日"
                    cell_g4 = sheet['BF4']
                    cell_g4.value = num
                    cell_g4 = sheet['L6']
                    cell_g4.value = dict["工程名称"]
                    cell_g4 = sheet['AQ6']
                    cell_g4.value = dict["建设单位"]
                    cell_g4 = sheet['L7']
                    cell_g4.value = dict["检查区域"]
                    cell_g4 = sheet['BB11']
                    cell_g4.value = "合格" if "有机房" in dict["有机房/无机房"] else "/"
                    cell_g4 = sheet['BB26']
                    cell_g4.value = "合格" if "有机房" in dict["有机房/无机房"] else "/"
                    cell_g4 = sheet['BB28']
                    cell_g4.value = "合格" if "有机房" in dict["有机房/无机房"] else "/"
                    cell_g4 = sheet['AI29']
                    cell_g4.value = dict["分包单位"]

                    workbook.save('直梯.xlsx')
                    workbook.close()

                # 运行异步主程序
                asyncio.run(main())
        except Exception as e:
            print(numtest1, e)

        try:
            for num in range(0, dict['电梯数量']):
                first_column = datas_exls.columns[num]
                dict["检查区域"] = datas_exls[first_column].iloc[60]
                dict["有机房/无机房"] = datas_exls[first_column].iloc[5]
                date_value = datas_exls[first_column].iloc[42]
                if date_value is not None:
                    # 将日期格式化为字符串，例如"2023-01-13"
                    date_str = date_value.strftime("%Y-%m-%d")
                    # 将格式化后的日期字符串存储到字典中
                    dict["电气检查日期"] = date_str
                else:
                    dict["电气检查日期"] = ""

                async def main():
                    global numtest1
                    path = r"./直梯/32-鲁DT-026.2电梯电气装置安装检查记录（二）（每台一份）.xlsx"
                    save_path = "直梯.xlsx"
                    wb = openpyxl.load_workbook(path)
                    wb2 = openpyxl.load_workbook(save_path)
                    sheetnames = wb.sheetnames

                    tasks = []
                    for numtest, sheetname in enumerate(sheetnames, start=1):
                        sheet = wb[sheetname]
                        sheet2 = wb2.create_sheet(f"第{numtest1}页")
                        print(numtest1)
                        numtest1 += 1
                        task = asyncio.create_task(copy_sheet(sheet, sheet2))
                        tasks.append(task)

                    # 等待所有任务完成
                    await asyncio.gather(*tasks)

                    # 保存并关闭工作簿
                    wb2.save(save_path)
                    wb.close()
                    wb2.close()
                    workbook = openpyxl.load_workbook('直梯.xlsx')
                    sheet = workbook[f'第{numtest1 - 1}页']
                    if len(str(dict["电气检查日期"]).split('-')) >= 3:
                        cell_g4 = sheet['AQ7']
                        cell_g4.value = str(dict["电气检查日期"]).split('-')[0] + "年" + \
                                        str(dict["电气检查日期"]).split('-')[
                                            1] + "月" + \
                                        str(dict["电气检查日期"]).split('-')[2] + "日"

                    cell_g4 = sheet['L6']
                    cell_g4.value = dict["工程名称"]
                    cell_g4 = sheet['BF4']
                    cell_g4.value = num
                    cell_g4 = sheet['AQ6']
                    cell_g4.value = dict["建设单位"]
                    cell_g4 = sheet['L7']
                    cell_g4.value = dict["检查区域"]
                    cell_g4 = sheet['BB10']
                    cell_g4.value = "合格" if "有机房" in dict["有机房/无机房"] else "/"
                    cell_g4 = sheet['AI28']
                    cell_g4.value = dict["分包单位"]
                    workbook.save('直梯.xlsx')
                    workbook.close()

                # 运行异步主程序
                asyncio.run(main())
        except Exception as e:
            print(numtest1, e)

        try:
            for num in range(0, dict['电梯数量']):
                first_column = datas_exls.columns[num]
                dict["检查区域"] = datas_exls[first_column].iloc[60]
                date_value = datas_exls[first_column].iloc[42]
                if date_value is not None:
                    # 将日期格式化为字符串，例如"2023-01-13"
                    date_str = date_value.strftime("%Y-%m-%d")
                    # 将格式化后的日期字符串存储到字典中
                    dict["电气检查日期"] = date_str
                else:
                    dict["电气检查日期"] = ""

                async def main():
                    global numtest1
                    path = r"./直梯/33-鲁DT-026.3电梯电气装置安装检查记录（三）（每台一份）.xlsx"
                    save_path = "直梯.xlsx"
                    wb = openpyxl.load_workbook(path)
                    wb2 = openpyxl.load_workbook(save_path)
                    sheetnames = wb.sheetnames

                    tasks = []
                    for numtest, sheetname in enumerate(sheetnames, start=1):
                        sheet = wb[sheetname]
                        sheet2 = wb2.create_sheet(f"第{numtest1}页")
                        print(numtest1)
                        numtest1 += 1
                        task = asyncio.create_task(copy_sheet(sheet, sheet2))
                        tasks.append(task)

                    # 等待所有任务完成
                    await asyncio.gather(*tasks)

                    # 保存并关闭工作簿
                    wb2.save(save_path)
                    wb.close()
                    wb2.close()
                    workbook = openpyxl.load_workbook('直梯.xlsx')
                    sheet = workbook[f'第{numtest1 - 1}页']
                    if len(str(dict["电气检查日期"]).split('-')) >= 3:
                        cell_g4 = sheet['AQ7']
                        cell_g4.value = str(dict["电气检查日期"]).split('-')[0] + "年" + \
                                        str(dict["电气检查日期"]).split('-')[
                                            1] + "月" + \
                                        str(dict["电气检查日期"]).split('-')[2] + "日"
                    cell_g4 = sheet['BF4']
                    cell_g4.value = num
                    cell_g4 = sheet['L6']
                    cell_g4.value = dict["工程名称"]
                    cell_g4 = sheet['AQ6']
                    cell_g4.value = dict["建设单位"]
                    cell_g4 = sheet['L7']
                    cell_g4.value = dict["检查区域"]
                    cell_g4 = sheet['AI24']
                    cell_g4.value = dict["分包单位"]
                    workbook.save('直梯.xlsx')
                    workbook.close()

                # 运行异步主程序
                asyncio.run(main())
        except Exception as e:
            print(numtest1, e)

        try:
            for num in range(0, dict['电梯数量']):
                first_column = datas_exls.columns[num]
                dict["绝缘电阻仪表型号"] = datas_exls[first_column].iloc[62]
                date_value = datas_exls[first_column].iloc[43]
                if date_value is not None:
                    # 将日期格式化为字符串，例如"2023-01-13"
                    date_str = date_value.strftime("%Y-%m-%d")
                    # 将格式化后的日期字符串存储到字典中
                    dict["绝缘电阻测试日期"] = date_str
                else:
                    dict["绝缘电阻测试日期"] = ""
                dict["环境温度"] = datas_exls[first_column].iloc[64]

                async def main():
                    global numtest1
                    path = r"./直梯/34-鲁DT-031绝缘电阻测试记录（每台一份）.xlsx"
                    save_path = "直梯.xlsx"
                    wb = openpyxl.load_workbook(path)
                    wb2 = openpyxl.load_workbook(save_path)
                    sheetnames = wb.sheetnames

                    tasks = []
                    for numtest, sheetname in enumerate(sheetnames, start=1):
                        sheet = wb[sheetname]
                        sheet2 = wb2.create_sheet(f"第{numtest1}页")
                        print(numtest1)
                        numtest1 += 1
                        task = asyncio.create_task(copy_sheet(sheet, sheet2))
                        tasks.append(task)

                    # 等待所有任务完成
                    await asyncio.gather(*tasks)

                    # 保存并关闭工作簿
                    wb2.save(save_path)
                    wb.close()
                    wb2.close()
                    workbook = openpyxl.load_workbook('直梯.xlsx')
                    sheet = workbook[f'第{numtest1 - 1}页']
                    if len(str(dict["绝缘电阻测试日期"]).split('-')) >= 3:
                        cell_g4 = sheet['AN8']
                        cell_g4.value = str(dict["绝缘电阻测试日期"]).split('-')[0] + "年" + \
                                        str(dict["绝缘电阻测试日期"]).split('-')[1] + "月" + \
                                        str(dict["绝缘电阻测试日期"]).split('-')[2] + "日"
                    cell_g4 = sheet['L6']
                    cell_g4.value = dict["工程名称"]
                    cell_g4 = sheet['BF4']
                    cell_g4.value = num
                    cell_g4 = sheet['AN6']
                    cell_g4.value = dict["建设单位"]
                    cell_g4 = sheet['L7']
                    cell_g4.value = dict["绝缘电阻仪表型号"]
                    cell_g4 = sheet['AN7']
                    cell_g4.value = dict["环境温度"]
                    cell_g4 = sheet['AJ27']
                    cell_g4.value = dict["分包单位"]
                    workbook.save('直梯.xlsx')
                    workbook.close()

                # 运行异步主程序
                asyncio.run(main())
        except Exception as e:
            print(numtest1, e)

        try:
            for num in range(0, dict['电梯数量']):
                first_column = datas_exls.columns[num]
                dict["绝缘电阻仪表型号"] = datas_exls[first_column].iloc[63]
                date_value = datas_exls[first_column].iloc[44]
                if date_value is not None:
                    # 将日期格式化为字符串，例如"2023-01-13"
                    date_str = date_value.strftime("%Y-%m-%d")
                    # 将格式化后的日期字符串存储到字典中
                    dict["接地电阻测试日期"] = date_str
                else:
                    dict["接地电阻测试日期"] = ""
                dict["环境温度"] = datas_exls[first_column].iloc[64]

                async def main():
                    global numtest1
                    path = r"./直梯/35-鲁DT-032接地电阻测试记录（每台一份）.xlsx"
                    save_path = "直梯.xlsx"
                    wb = openpyxl.load_workbook(path)
                    wb2 = openpyxl.load_workbook(save_path)
                    sheetnames = wb.sheetnames

                    tasks = []
                    for numtest, sheetname in enumerate(sheetnames, start=1):
                        sheet = wb[sheetname]
                        sheet2 = wb2.create_sheet(f"第{numtest1}页")
                        print(numtest1)
                        numtest1 += 1
                        task = asyncio.create_task(copy_sheet(sheet, sheet2))
                        tasks.append(task)

                    # 等待所有任务完成
                    await asyncio.gather(*tasks)

                    # 保存并关闭工作簿
                    wb2.save(save_path)
                    wb.close()
                    wb2.close()
                    workbook = openpyxl.load_workbook('直梯.xlsx')
                    sheet = workbook[f'第{numtest1 - 1}页']
                    if len(str(dict["接地电阻测试日期"]).split('-')) >= 3:
                        cell_g4 = sheet['AR8']
                        cell_g4.value = str(dict["接地电阻测试日期"]).split('-')[0] + "年" + \
                                        str(dict["接地电阻测试日期"]).split('-')[1] + "月" + \
                                        str(dict["接地电阻测试日期"]).split('-')[2] + "日"
                    cell_g4 = sheet['L6']
                    cell_g4.value = dict["工程名称"]
                    cell_g4 = sheet['BF4']
                    cell_g4.value = num
                    cell_g4 = sheet['AR6']
                    cell_g4.value = dict["建设单位"]
                    cell_g4 = sheet['L7']
                    cell_g4.value = dict["绝缘电阻仪表型号"]
                    cell_g4 = sheet['AR7']
                    cell_g4.value = dict["环境温度"]
                    cell_g4 = sheet['AJ21']
                    cell_g4.value = dict["分包单位"]
                    workbook.save('直梯.xlsx')
                    workbook.close()

                # 运行异步主程序
                asyncio.run(main())
        except Exception as e:
            print(numtest1, e)

        try:
            for num in range(0, dict['电梯数量']):
                first_column = datas_exls.columns[num]
                dict["电梯型号规格"] = datas_exls[first_column].iloc[29]
                dict["设备梯号"] = datas_exls[first_column].iloc[0]
                dict["载重"] = datas_exls[first_column].iloc[6]
                dict["驱动方式"] = datas_exls[first_column].iloc[30]
                dict["层站"] = datas_exls[first_column].iloc[8]
                dict["层高"] = datas_exls[first_column].iloc[11]
                dict["速度"] = datas_exls[first_column].iloc[7]
                date_value = datas_exls[first_column].iloc[47]
                if date_value is not None:
                    # 将日期格式化为字符串，例如"2023-01-13"
                    date_str = date_value.strftime("%Y-%m-%d")
                    # 将格式化后的日期字符串存储到字典中
                    dict["轿厢平层测试日期"] = date_str
                else:
                    dict["轿厢平层测试日期"] = ""
                dict["轿厢平层测量工具"] = datas_exls[first_column].iloc[65]

                async def main():
                    global numtest1
                    path = r"./直梯/36-鲁DT-033轿厢平层准确度测量记录（每台一份）.xlsx"
                    save_path = "直梯.xlsx"
                    wb = openpyxl.load_workbook(path)
                    wb2 = openpyxl.load_workbook(save_path)
                    sheetnames = wb.sheetnames

                    tasks = []
                    for numtest, sheetname in enumerate(sheetnames, start=1):
                        sheet = wb[sheetname]
                        sheet2 = wb2.create_sheet(f"第{numtest1}页")
                        print(numtest1)
                        numtest1 += 1
                        task = asyncio.create_task(copy_sheet(sheet, sheet2))
                        tasks.append(task)

                    # 等待所有任务完成
                    await asyncio.gather(*tasks)

                    # 保存并关闭工作簿
                    wb2.save(save_path)
                    wb.close()
                    wb2.close()
                    workbook = openpyxl.load_workbook('直梯.xlsx')
                    sheet = workbook[f'第{numtest1 - 1}页']
                    if len(str(dict["轿厢平层测试日期"]).split('-')) >= 3:
                        cell_g4 = sheet['AQ6']
                        cell_g4.value = str(dict["轿厢平层测试日期"]).split('-')[0] + "年" + \
                                        str(dict["轿厢平层测试日期"]).split('-')[1] + "月" + \
                                        str(dict["轿厢平层测试日期"]).split('-')[2] + "日"
                    cell_g4 = sheet['L6']
                    cell_g4.value = dict["工程名称"]
                    cell_g4 = sheet['BF4']
                    cell_g4.value = num
                    cell_g4 = sheet['L7']
                    cell_g4.value = dict["电梯型号规格"]
                    cell_g4 = sheet['AQ7']
                    cell_g4.value = dict["设备梯号"]
                    cell_g4 = sheet['L8']
                    cell_g4.value = dict["载重"]
                    cell_g4 = sheet['AQ8']
                    cell_g4.value = dict["驱动方式"]
                    cell_g4 = sheet['L9']
                    if len(dict["层站"].rsplit('/', 1)[0]) > 2:
                        cell_g4.value = dict["层站"].rsplit('/', 1)[0]
                    cell_g4 = sheet['AH9']
                    cell_g4.value = dict["层高"]
                    cell_g4 = sheet['BA9']
                    cell_g4.value = dict["速度"]
                    cell_g4 = sheet['AH10']
                    cell_g4.value = dict["轿厢平层测量工具"]
                    cell_g4 = sheet['AQ38']
                    cell_g4.value = dict["分包单位"]
                    workbook.save('直梯.xlsx')
                    workbook.close()

                # 运行异步主程序
                asyncio.run(main())
        except Exception as e:
            print(numtest1, e)

        try:
            for num in range(0, dict['电梯数量']):
                first_column = datas_exls.columns[num]
                dict["电梯型号规格"] = datas_exls[first_column].iloc[29]
                dict["设备梯号"] = datas_exls[first_column].iloc[0]
                dict["层站门"] = datas_exls[first_column].iloc[8]
                dict["起止层"] = datas_exls[first_column].iloc[9]
                dict["盲层"] = datas_exls[first_column].iloc[10]
                dict["初始层"] = dict["起止层"].split('/')[0]
                dict["终止层"] = dict["起止层"].split('/')[1]
                dict["开门方式"] = datas_exls[first_column].iloc[12]
                dict["开门宽度"] = datas_exls[first_column].iloc[13]
                dict["门扇数"] = datas_exls[first_column].iloc[14]
                dict["门锁装置厂家"] = datas_exls[first_column].iloc[15]
                dict["门锁型式试验单位"] = datas_exls[first_column].iloc[16]
                date_value = datas_exls[first_column].iloc[48]
                if date_value is not None:
                    # 将日期格式化为字符串，例如"2023-01-13"
                    date_str = date_value.strftime("%Y-%m-%d")
                    # 将格式化后的日期字符串存储到字典中
                    dict["层门安全装置测试日期"] = date_str
                else:
                    dict["层门安全装置测试日期"] = ""
                date_value = datas_exls[first_column].iloc[17]
                if date_value is not None:
                    # 将日期格式化为字符串，例如"2023-01-13"
                    date_str = date_value.strftime("%Y-%m-%d")
                    # 将格式化后的日期字符串存储到字典中
                    dict["门锁型式试验有效期"] = date_str
                else:
                    dict["门锁型式试验有效期"] = ""

                async def main():
                    global numtest1
                    path = r"./直梯/37-鲁DT-034电梯层门安全装置检验记录（每台一份）.xlsx"
                    save_path = "直梯.xlsx"
                    wb = openpyxl.load_workbook(path)
                    wb2 = openpyxl.load_workbook(save_path)
                    sheetnames = wb.sheetnames

                    tasks = []
                    for numtest, sheetname in enumerate(sheetnames, start=1):
                        sheet = wb[sheetname]
                        sheet2 = wb2.create_sheet(f"第{numtest1}页")
                        print(numtest1)
                        numtest1 += 1
                        task = asyncio.create_task(copy_sheet(sheet, sheet2))
                        tasks.append(task)

                    # 等待所有任务完成
                    await asyncio.gather(*tasks)

                    # 保存并关闭工作簿
                    wb2.save(save_path)
                    wb.close()
                    wb2.close()
                    workbook = openpyxl.load_workbook('直梯.xlsx')
                    sheet = workbook[f'第{numtest1 - 1}页']
                    if len(str(dict["层门安全装置测试日期"]).split('-')) >= 3:
                        cell_g4 = sheet['AQ6']
                        cell_g4.value = str(dict["层门安全装置测试日期"]).split('-')[0] + "年" + \
                                        str(dict["层门安全装置测试日期"]).split('-')[1] + "月" + \
                                        str(dict["层门安全装置测试日期"]).split('-')[2] + "日"
                    cell_g4 = sheet['H6']
                    cell_g4.value = dict["工程名称"]
                    cell_g4 = sheet['BF4']
                    cell_g4.value = num
                    cell_g4 = sheet['H7']
                    cell_g4.value = dict["电梯型号规格"]
                    cell_g4 = sheet['AQ7']
                    cell_g4.value = dict["设备梯号"]
                    cell_g4 = sheet['H8']
                    cell_g4.value = dict["层站门"]
                    cell_g4 = sheet['W8']
                    cell_g4.value = dict["开门方式"]
                    cell_g4 = sheet['AQ8']
                    cell_g4.value = dict["开门宽度"]
                    cell_g4 = sheet['BF8']
                    cell_g4.value = dict["门扇数"]
                    cell_g4 = sheet['W9']
                    cell_g4.value = dict["门锁装置厂家"]
                    cell_g4 = sheet['W10']
                    cell_g4.value = dict["门锁型式试验单位"]
                    cell_g4 = sheet['AY9']
                    cell_g4.value = dict["门锁型式试验有效期"]
                    cell_g4 = sheet['AI39']
                    cell_g4.value = dict["分包单位"]
                    datanum = 13
                    for data in range(int(dict["初始层"]), int(dict["终止层"]) + 1):
                        if data == 0:
                            continue
                        if dict["盲层"] != "":
                            mc_values = [int(mc) for mc in dict["盲层"].split(',')]
                            if data in mc_values:
                                continue
                        cell_g4 = sheet[f'B{datanum}']
                        cell_g4.value = data
                        cell_g4 = sheet[f'E{datanum}']
                        random_float = round(random.uniform(3.3, 3.5), 1)
                        formatted_float = "{:.1f}".format(random_float)
                        cell_g4.value = (formatted_float + "s")
                        cell_g4 = sheet[f'H{datanum}']
                        random_float = round(random.uniform(3.7, 3.9), 1)
                        formatted_float = "{:.1f}".format(random_float)
                        cell_g4.value = (formatted_float + "s")
                        cell_g4 = sheet[f'K{datanum}']
                        cell_g4.value = '√'
                        cell_g4 = sheet[f'Q{datanum}']
                        cell_g4.value = '√'
                        random_float = int(random.uniform(8, 10))
                        cell_g4 = sheet[f'W{datanum}']
                        cell_g4.value = str(random_float) + "mm"
                        cell_g4 = sheet[f'Z{datanum}']
                        cell_g4.value = str(random_float) + "mm"
                        cell_g4 = sheet[f'AC{datanum}']
                        cell_g4.value = '√'
                        cell_g4 = sheet[f'AF{datanum}']
                        cell_g4.value = '√'
                        cell_g4 = sheet[f'AQ{datanum}']
                        cell_g4.value = '√'
                        cell_g4 = sheet[f'BE{datanum}']
                        cell_g4.value = '√'
                        cell_g4 = sheet[f'AI{datanum}']
                        cell_g4.value = '132N'
                        datanum = datanum + 1
                        if datanum > 29:
                            workbook.save('直梯.xlsx')
                            workbook.close()
                            path = r"./直梯/37-鲁DT-034电梯层门安全装置检验记录（每台一份）.xlsx"
                            save_path = "直梯.xlsx"
                            wb = openpyxl.load_workbook(path)
                            wb2 = openpyxl.load_workbook(save_path)
                            sheetnames = wb.sheetnames
                            for sheetname in sheetnames:
                                print(f"第{numtest1}页")
                                sheet = wb[sheetname]
                                sheet2 = wb2.create_sheet(f"第{numtest1}页")
                                # Tab color
                                sheet2.sheet_properties.tabColor = sheet.sheet_properties.tabColor

                                # 复制列宽度
                                for column in sheet.column_dimensions:
                                    sheet2.column_dimensions[column].width = sheet.column_dimensions[column].width

                                # 复制行高
                                for row in sheet.row_dimensions:
                                    sheet2.row_dimensions[row].height = sheet.row_dimensions[row].height

                                # 处理合并单元格
                                wm = list(sheet.merged_cells)
                                if len(wm) > 0:
                                    for i in range(0, len(wm)):
                                        cell2 = str(wm[i]).replace('(<CellRange ', '').replace('>,)', '')
                                        sheet2.merge_cells(cell2)

                                # 复制行、列和单元格值
                                for i, row in enumerate(sheet.iter_rows()):
                                    for j, cell in enumerate(row):
                                        # 复制单元格值
                                        sheet2.cell(row=i + 1, column=j + 1, value=cell.value)

                                        # 如果是合并单元格的一部分，则设置单元格格式和尺寸
                                        if cell.coordinate in sheet.merged_cells:
                                            for merged_cell in wm:
                                                if cell.coordinate in merged_cell:
                                                    first_cell = sheet.cell(merged_cell.min_row,
                                                                            merged_cell.min_col)
                                                    target_cell = sheet2.cell(i + 1, j + 1)
                                                    sheet2.row_dimensions[i + 1].height = sheet.row_dimensions[
                                                        first_cell.row].height
                                                    sheet2.column_dimensions[get_column_letter(j + 1)].width = \
                                                        sheet.column_dimensions[
                                                            get_column_letter(first_cell.column)].width
                                                    break

                                        # 设置单元格格式
                                        source_cell = sheet.cell(i + 1, j + 1)
                                        target_cell = sheet2.cell(i + 1, j + 1)

                                        # 创建新样式对象
                                        font = Font(name=source_cell.font.name, size=source_cell.font.size,
                                                    bold=source_cell.font.bold, italic=source_cell.font.italic,
                                                    vertAlign=source_cell.font.vertAlign,
                                                    underline=source_cell.font.underline,
                                                    strike=source_cell.font.strike,
                                                    color='FF000000')  # Set font color to black

                                        black_side = Side(color='FF000000')
                                        border = Border(
                                            left=Side(color=black_side.color,
                                                      border_style=source_cell.border.left.border_style),
                                            right=Side(color=black_side.color,
                                                       border_style=source_cell.border.right.border_style),
                                            top=Side(color=black_side.color,
                                                     border_style=source_cell.border.top.border_style),
                                            bottom=Side(color=black_side.color,
                                                        border_style=source_cell.border.bottom.border_style))

                                        fill = PatternFill(fill_type=source_cell.fill.fill_type,
                                                           fgColor=source_cell.fill.fgColor,
                                                           bgColor=source_cell.fill.bgColor,
                                                           patternType=source_cell.fill.patternType)

                                        alignment = Alignment(horizontal=source_cell.alignment.horizontal,
                                                              vertical=source_cell.alignment.vertical,
                                                              text_rotation=source_cell.alignment.text_rotation,
                                                              wrap_text=source_cell.alignment.wrap_text,
                                                              shrink_to_fit=source_cell.alignment.shrink_to_fit,
                                                              indent=source_cell.alignment.indent)

                                        protection = Protection(locked=source_cell.protection.locked,
                                                                hidden=source_cell.protection.hidden)

                                        # 将新样式对象分配给目标单元格
                                        target_cell.font = font
                                        target_cell.border = border
                                        target_cell.fill = fill
                                        target_cell.alignment = alignment
                                        target_cell.protection = protection
                                        target_cell.number_format = source_cell.number_format
                                numtest1 = numtest1 + 1
                            # 保存并关闭工作簿
                            wb2.save(save_path)
                            wb.close()
                            wb2.close()
                            workbook = openpyxl.load_workbook('直梯.xlsx')
                            sheet = workbook[f'第{numtest1 - 1}页']
                            cell_g4 = sheet['BF4']
                            cell_g4.value = num
                            if len(str(dict["测试日期"]).split('-')) >= 3:
                                cell_g4 = sheet['AB6']
                                cell_g4.value = str(dict["测试日期"]).split('-')[0] + "年" + \
                                                str(dict["测试日期"]).split('-')[1] + "月" + \
                                                str(dict["测试日期"]).split('-')[2] + "日"
                            cell_g4 = sheet['H6']
                            cell_g4.value = dict["工程名称"]
                            cell_g4 = sheet['H7']
                            cell_g4.value = dict["电梯型号"]
                            cell_g4 = sheet['AQ7']
                            cell_g4.value = dict["梯号"]
                            cell_g4 = sheet['AQ8']
                            cell_g4.value = dict["开门宽度B(mm)"]
                            cell_g4 = sheet['BF8']
                            cell_g4.value = dict["门扇数"]
                            cell_g4 = sheet['W9']
                            cell_g4.value = dict["门锁装置铭牌制造厂名称"]
                            cell_g4 = sheet['W10']
                            cell_g4.value = dict["型式试验标志及试验单位"]
                            cell_g4 = sheet['AY9']
                            cell_g4.value = dict["有效期"]
                            cell_g4 = sheet['AI39']
                            cell_g4.value = dict["分包单位"]
                            datanum = 13
                    workbook.save('直梯.xlsx')
                    workbook.close()

                # 运行异步主程序
                asyncio.run(main())
        except Exception as e:
            print(numtest1, e)

        try:
            for num in range(0, dict['电梯数量']):
                first_column = datas_exls.columns[num]
                dict["电梯型号规格"] = datas_exls[first_column].iloc[29]
                dict["设备梯号"] = datas_exls[first_column].iloc[0]
                dict["缓冲器形式（耗能/蓄能）"] = datas_exls[first_column].iloc[18]
                dict["补偿链配备（有/无））"] = datas_exls[first_column].iloc[19]
                date_value = datas_exls[first_column].iloc[49]
                if date_value is not None:
                    # 将日期格式化为字符串，例如"2023-01-13"
                    date_str = date_value.strftime("%Y-%m-%d")
                    # 将格式化后的日期字符串存储到字典中
                    dict["电气安全装置测试日期"] = date_str
                else:
                    dict["电气安全装置测试日期"] = ""

                async def main():
                    global numtest1
                    path = r"./直梯/38-鲁DT-035电梯电气安全装置检验记录（每台一份）.xlsx"
                    save_path = "直梯.xlsx"
                    wb = openpyxl.load_workbook(path)
                    wb2 = openpyxl.load_workbook(save_path)
                    sheetnames = wb.sheetnames

                    tasks = []
                    for numtest, sheetname in enumerate(sheetnames, start=1):
                        sheet = wb[sheetname]
                        sheet2 = wb2.create_sheet(f"第{numtest1}页")
                        print(numtest1)
                        numtest1 += 1
                        task = asyncio.create_task(copy_sheet(sheet, sheet2))
                        tasks.append(task)

                    # 等待所有任务完成
                    await asyncio.gather(*tasks)

                    # 保存并关闭工作簿
                    wb2.save(save_path)
                    wb.close()
                    wb2.close()
                    workbook = openpyxl.load_workbook('直梯.xlsx')
                    sheet = workbook[f'第{numtest1 - 1}页']
                    cell_g4 = sheet['BF4']
                    cell_g4.value = num
                    if len(str(dict["电气安全装置测试日期"]).split('-')) >= 3:
                        cell_g4 = sheet['AQ6']
                        cell_g4.value = str(dict["电气安全装置测试日期"]).split('-')[0] + "年" + \
                                        str(dict["电气安全装置测试日期"]).split('-')[1] + "月" + \
                                        str(dict["电气安全装置测试日期"]).split('-')[2] + "日"
                    cell_g4 = sheet['L6']
                    cell_g4.value = dict["工程名称"]
                    cell_g4 = sheet['BF4']
                    cell_g4.value = num
                    cell_g4 = sheet['L7']
                    cell_g4.value = dict["电梯型号规格"]
                    cell_g4 = sheet['BB11']
                    cell_g4.value = "合格" if "耗能" in dict["缓冲器形式（耗能/蓄能）"] else "/"
                    cell_g4 = sheet['BB27']
                    cell_g4.value = "合格" if "有" in dict["补偿链配备（有/无））"] else "/"
                    cell_g4 = sheet['AQ7']
                    cell_g4.value = dict["设备梯号"]
                    cell_g4 = sheet['AI30']
                    cell_g4.value = dict["分包单位"]
                    workbook.save('直梯.xlsx')
                    workbook.close()

                # 运行异步主程序
                asyncio.run(main())
        except Exception as e:
            print(numtest1, e)

        try:
            for num in range(0, dict['电梯数量']):
                first_column = datas_exls.columns[num]
                dict["检查区域"] = datas_exls[first_column].iloc[60]
                dict["缓冲器形式（耗能/蓄能）"] = datas_exls[first_column].iloc[18]
                dict["安全钳形式"] = datas_exls[first_column].iloc[66]
                date_value = datas_exls[first_column].iloc[50]
                if date_value is not None:
                    # 将日期格式化为字符串，例如"2023-01-13"
                    date_str = date_value.strftime("%Y-%m-%d")
                    # 将格式化后的日期字符串存储到字典中
                    dict["电梯整机检查日期"] = date_str
                else:
                    dict["电梯整机检查日期"] = ""

                async def main():
                    global numtest1
                    path = r"./直梯/39鲁DT-036电梯整机功能检验记录（每台一份）.xlsx"
                    save_path = "直梯.xlsx"
                    wb = openpyxl.load_workbook(path)
                    wb2 = openpyxl.load_workbook(save_path)
                    sheetnames = wb.sheetnames

                    tasks = []
                    for numtest, sheetname in enumerate(sheetnames, start=1):
                        sheet = wb[sheetname]
                        sheet2 = wb2.create_sheet(f"第{numtest1}页")
                        print(numtest1)
                        numtest1 += 1
                        task = asyncio.create_task(copy_sheet(sheet, sheet2))
                        tasks.append(task)

                    # 等待所有任务完成
                    await asyncio.gather(*tasks)

                    # 保存并关闭工作簿
                    wb2.save(save_path)
                    wb.close()
                    wb2.close()
                    workbook = openpyxl.load_workbook('直梯.xlsx')
                    sheet = workbook[f'第{numtest1 - 1}页']
                    if len(str(dict["电梯整机检查日期"]).split('-')) >= 3:
                        cell_g4 = sheet['AQ7']
                        cell_g4.value = str(dict["电梯整机检查日期"]).split('-')[0] + "年" + \
                                        str(dict["电梯整机检查日期"]).split('-')[1] + "月" + \
                                        str(dict["电梯整机检查日期"]).split('-')[2] + "日"
                    cell_g4 = sheet['L6']
                    cell_g4.value = dict["工程名称"]
                    cell_g4 = sheet['BF4']
                    cell_g4.value = num
                    cell_g4 = sheet['AQ6']
                    cell_g4.value = dict["建设单位"]
                    cell_g4 = sheet['L7']
                    cell_g4.value = dict["检查区域"]
                    cell_g4 = sheet['BB16']
                    cell_g4.value = "符合规范要求" if "瞬时" in dict["安全钳形式"] else "/"
                    cell_g4 = sheet['BB17']
                    cell_g4.value = "符合规范要求" if "渐进" in dict["安全钳形式"] else "/"
                    cell_g4 = sheet['BB18']
                    cell_g4.value = "符合规范要求" if "蓄能" in dict["缓冲器形式（耗能/蓄能）"] else "/"
                    cell_g4 = sheet['BB19']
                    cell_g4.value = "符合规范要求" if "耗能" in dict["缓冲器形式（耗能/蓄能）"] else "/"

                    cell_g4 = sheet['AI20']
                    cell_g4.value = dict["分包单位"]
                    workbook.save('直梯.xlsx')
                    workbook.close()

                # 运行异步主程序
                asyncio.run(main())
        except Exception as e:
            print(numtest1, e)

        try:
            for num in range(0, dict['电梯数量']):
                first_column = datas_exls.columns[num]
                dict["检验区域"] = datas_exls[first_column].iloc[60]
                date_value = datas_exls[first_column].iloc[51]
                if date_value is not None:
                    # 将日期格式化为字符串，例如"2023-01-13"
                    date_str = date_value.strftime("%Y-%m-%d")
                    # 将格式化后的日期字符串存储到字典中
                    dict["电梯主要功能检验日期"] = date_str
                else:
                    dict["电梯主要功能检验日期"] = ""

                async def main():
                    global numtest1
                    path = r"./直梯/40-鲁DT-037电梯主要功能检验记录（每台一份）.xlsx"
                    save_path = "直梯.xlsx"
                    wb = openpyxl.load_workbook(path)
                    wb2 = openpyxl.load_workbook(save_path)
                    sheetnames = wb.sheetnames

                    tasks = []
                    for numtest, sheetname in enumerate(sheetnames, start=1):
                        sheet = wb[sheetname]
                        sheet2 = wb2.create_sheet(f"第{numtest1}页")
                        print(numtest1)
                        numtest1 += 1
                        task = asyncio.create_task(copy_sheet(sheet, sheet2))
                        tasks.append(task)

                    # 等待所有任务完成
                    await asyncio.gather(*tasks)

                    # 保存并关闭工作簿
                    wb2.save(save_path)
                    wb.close()
                    wb2.close()
                    workbook = openpyxl.load_workbook('直梯.xlsx')
                    sheet = workbook[f'第{numtest1 - 1}页']
                    if len(str(dict["电梯主要功能检验日期"]).split('-')) >= 3:
                        cell_g4 = sheet['AQ7']
                        cell_g4.value = str(dict["电梯主要功能检验日期"]).split('-')[0] + "年" + \
                                        str(dict["电梯主要功能检验日期"]).split('-')[1] + "月" + \
                                        str(dict["电梯主要功能检验日期"]).split('-')[2] + "日"
                    cell_g4 = sheet['L6']
                    cell_g4.value = dict["工程名称"]
                    cell_g4 = sheet['BF4']
                    cell_g4.value = num
                    cell_g4 = sheet['AQ6']
                    cell_g4.value = dict["建设单位"]
                    cell_g4 = sheet['L7']
                    cell_g4.value = dict["检验区域"]
                    cell_g4 = sheet['AJ31']
                    cell_g4.value = dict["分包单位"]
                    workbook.save('直梯.xlsx')
                    workbook.close()

                # 运行异步主程序
                asyncio.run(main())
        except Exception as e:
            print(numtest1, e)

        try:
            for num in range(0, dict['电梯数量']):
                first_column = datas_exls.columns[num]
                date_value = datas_exls[first_column].iloc[52]
                if date_value is not None:
                    # 将日期格式化为字符串，例如"2023-01-13"
                    date_str = date_value.strftime("%Y-%m-%d")
                    # 将格式化后的日期字符串存储到字典中
                    dict["电梯负荷运行试验日期"] = date_str
                else:
                    dict["电梯负荷运行试验日期"] = ""
                dict["电梯型号规格"] = datas_exls[first_column].iloc[29]
                dict["设备梯号"] = datas_exls[first_column].iloc[0]
                dict["速度"] = datas_exls[first_column].iloc[7]
                dict["载重"] = datas_exls[first_column].iloc[6]
                dict["层站门"] = datas_exls[first_column].iloc[8]
                dict["负荷运行电机功率"] = datas_exls[first_column].iloc[22]
                dict["负荷运行额定转速"] = datas_exls[first_column].iloc[23]
                dict["负荷运行电压"] = datas_exls[first_column].iloc[24]
                dict["负荷运行电流"] = datas_exls[first_column].iloc[25]

                async def main():
                    global numtest1
                    path = r"./直梯/41-鲁DT-038电梯负荷运行试验记录（每台一份）.xlsx"
                    save_path = "直梯.xlsx"
                    wb = openpyxl.load_workbook(path)
                    wb2 = openpyxl.load_workbook(save_path)
                    sheetnames = wb.sheetnames

                    tasks = []
                    for numtest, sheetname in enumerate(sheetnames, start=1):
                        sheet = wb[sheetname]
                        sheet2 = wb2.create_sheet(f"第{numtest1}页")
                        print(numtest1)
                        numtest1 += 1
                        task = asyncio.create_task(copy_sheet(sheet, sheet2))
                        tasks.append(task)

                    # 等待所有任务完成
                    await asyncio.gather(*tasks)

                    # 保存并关闭工作簿
                    wb2.save(save_path)
                    wb.close()
                    wb2.close()
                    workbook = openpyxl.load_workbook('直梯.xlsx')
                    sheet = workbook[f'第{numtest1 - 1}页']
                    if len(str(dict["电梯负荷运行试验日期"]).split('-')) >= 3:
                        cell_g4 = sheet['AP6']
                        cell_g4.value = str(dict["电梯负荷运行试验日期"]).split('-')[0] + "年" + \
                                        str(dict["电梯负荷运行试验日期"]).split('-')[1] + "月" + \
                                        str(dict["电梯负荷运行试验日期"]).split('-')[2] + "日"
                    cell_g4 = sheet['I6']
                    cell_g4.value = dict["工程名称"]
                    cell_g4 = sheet['BE4']
                    cell_g4.value = num
                    cell_g4 = sheet['I7']
                    cell_g4.value = dict["电梯型号规格"]
                    cell_g4 = sheet['AP7']
                    cell_g4.value = dict["设备梯号"]
                    cell_g4 = sheet['I8']
                    cell_g4.value = dict["速度"]
                    cell_g4 = sheet['V8']
                    cell_g4.value = dict["载重"]
                    cell_g4 = sheet['AP8']
                    cell_g4.value = dict["层站门"].rsplit('/', 1)[0]
                    cell_g4 = sheet['I9']
                    cell_g4.value = dict["负荷运行电机功率"]
                    cell_g4 = sheet['V9']
                    cell_g4.value = dict["负荷运行额定转速"]
                    cell_g4 = sheet['AM9']
                    cell_g4.value = dict["负荷运行电压"]
                    cell_g4 = sheet['BE9']
                    cell_g4.value = dict["负荷运行电流"]
                    cell_g4 = sheet['AI28']
                    cell_g4.value = dict["分包单位"]
                    # 打开工作簿
                    workbook1 = openpyxl.load_workbook('./测量值/平衡系数表.xlsx')
                    # 选择工作表
                    worksheet = workbook1[f'Sheet{num + 1}']
                    # 获取某个单元格的值
                    cell_value = worksheet['B1'].value
                    dict["额定载荷"] = cell_value
                    cell_g4 = sheet['I13']
                    cell_g4.value = worksheet['B2'].value
                    cell_g4 = sheet['I15']
                    cell_g4.value = worksheet['C2'].value
                    cell_g4 = sheet['I17']
                    cell_g4.value = worksheet['D2'].value
                    cell_g4 = sheet['I19']
                    cell_g4.value = worksheet['F2'].value
                    cell_g4 = sheet['I21']
                    cell_g4.value = worksheet['G2'].value
                    cell_g4 = sheet['I23']
                    cell_g4.value = worksheet['H2'].value
                    listzm = ["B", "C", "D", "F", "G", "H"]
                    sum_lis = 0
                    for i in range(13, 24, 2):
                        cell_g4 = sheet[f'AG{i}']
                        cell_g4.value = worksheet[f'{listzm[sum_lis]}4'].value
                        sum_lis += 1
                    sum_lis = 0
                    for i in range(14, 25, 2):
                        cell_g4 = sheet[f'AG{i}']
                        cell_g4.value = worksheet[f'{listzm[sum_lis]}5'].value
                        sum_lis += 1

                    workbook.save('直梯.xlsx')
                    workbook.close()

                # 运行异步主程序
                asyncio.run(main())
        except Exception as e:
            print(numtest1, e)

        # 支持中文
        plt.rcParams['font.sans-serif'] = ['SimHei']  # 用来正常显示中文标签
        plt.rcParams['axes.unicode_minus'] = False  # 用来正常显示负号

        try:
            for num in range(0, dict['电梯数量']):

                # 创建一个图形对象
                exls1 = pd.read_excel('./测量值/平衡系数表.xlsx', sheet_name=f'Sheet{num + 1}',
                                      keep_default_na=False)
                # 去掉前两列和最后一列
                datas1 = exls1.drop(exls1.columns[:2], axis=1)  # 去掉前两列
                datas1 = datas1.drop(datas1.columns[-1], axis=1)  # 去掉最后一列
                x = datas1.values[1]
                y1 = datas1.values[2]
                y2 = datas1.values[3]
                # 绘制折线图
                plt.plot(x, y1, label='上行电流', marker='s')
                plt.plot(x, y2, label='下行电流', marker='s')
                # 获取交点坐标
                intersections = []
                for data in range(len(x) - 1):
                    if y1[data] < y2[data] and y1[data + 1] > y2[data + 1] or y1[data] > y2[data] and y1[data +
                                                                                                         1] < \
                            y2[
                                data + 1]:
                        slope1 = (y1[data + 1] - y1[data]) / (x[data + 1] - x[data])
                        slope2 = (y2[data + 1] - y2[data]) / (x[data + 1] - x[data])
                        intersect_x = (y2[data] - y1[data] + slope1 * x[data] - slope2 * x[data]) / (
                                    slope1 - slope2)
                        intersect_y = y1[data] + slope1 * (intersect_x - x[data])
                        intersections.append((intersect_x, intersect_y))
                # 打印交点坐标
                for point in intersections:
                    intersect_x_formatted = "{:.1f}%".format(point[0] * 100)
                    intersect_y_formatted = "{:.0f}".format(point[1] * 100)
                    dict['intersect_x_formatted'] = intersect_x_formatted
                    dict['intersect_y_formatted'] = intersect_y_formatted
                    print("交点坐标:", intersect_x_formatted, intersect_y_formatted)
                # 设置x轴为百分比格式
                plt.gca().xaxis.set_major_formatter(ticker.PercentFormatter(xmax=1, decimals=1))
                # 显示图例和图形
                plt.legend()
                # 添加网格线
                plt.grid(True)
                # 显示图表
                plt.savefig(f'image{num}.png')
                # 清除当前图形对象，以便绘制下一次循环的图像
                plt.clf()

                first_column = datas_exls.columns[num]
                date_value = datas_exls[first_column].iloc[45]
                if date_value is not None:
                    # 将日期格式化为字符串，例如"2023-01-13"
                    date_str = date_value.strftime("%Y-%m-%d")
                    # 将格式化后的日期字符串存储到字典中
                    dict["负荷运行曲线图绘制日期"] = date_str
                else:
                    dict["负荷运行曲线图绘制日期"] = ""

                dict["电梯型号规格"] = datas_exls[first_column].iloc[29]
                dict["设备梯号"] = datas_exls[first_column].iloc[0]
                dict["层站门"] = datas_exls[first_column].iloc[8]
                dict["载重"] = datas_exls[first_column].iloc[6]
                dict["起止层"] = datas_exls[first_column].iloc[9]
                dict["盲层"] = datas_exls[first_column].iloc[10]
                dict["初始层"] = dict["起止层"].rsplit('/')[0]
                dict["终止层"] = dict["起止层"].rsplit('/')[1]

                async def main():
                    global numtest1
                    path = r"./直梯/42-鲁DT-039电梯负荷运行试验曲线图（每台一份）.xlsx"
                    save_path = "直梯.xlsx"
                    wb = openpyxl.load_workbook(path)
                    wb2 = openpyxl.load_workbook(save_path)
                    sheetnames = wb.sheetnames

                    tasks = []
                    for numtest, sheetname in enumerate(sheetnames, start=1):
                        sheet = wb[sheetname]
                        sheet2 = wb2.create_sheet(f"第{numtest1}页")
                        print(numtest1)
                        numtest1 += 1
                        task = asyncio.create_task(copy_sheet(sheet, sheet2))
                        tasks.append(task)

                    # 等待所有任务完成
                    await asyncio.gather(*tasks)

                    # 保存并关闭工作簿
                    wb2.save(save_path)
                    wb.close()
                    wb2.close()
                    workbook = openpyxl.load_workbook('直梯.xlsx')
                    sheet = workbook[f'第{numtest1 - 1}页']
                    cell_g4 = sheet['BA4']
                    cell_g4.value = num
                    cell_g4 = sheet['K6']
                    cell_g4.value = dict["工程名称"]
                    cell_g4 = sheet['K7']
                    cell_g4.value = dict["电梯型号规格"]
                    cell_g4 = sheet['AO7']
                    cell_g4.value = dict["设备梯号"]
                    cell_g4 = sheet['K8']
                    cell_g4.value = dict["载重"]
                    cell_g4 = sheet['L49']
                    cell_g4.value = dict["分包单位"]
                    cell_g4 = sheet['AO6']
                    cell_g4.value = dict["分包单位"]
                    cell_g4 = sheet['AF8']
                    cell_g4.value = dict["intersect_x_formatted"]
                    cell_g4 = sheet['AX8']
                    cell_g4.value = dict["intersect_y_formatted"]
                    if len(str(dict["负荷运行曲线图绘制日期"]).split('-')) >= 3:
                        cell_g4 = sheet['AP49']
                        cell_g4.value = str(dict["负荷运行曲线图绘制日期"]).split('-')[0] + "年" + \
                                        str(dict["负荷运行曲线图绘制日期"]).split('-')[1] + "月" + \
                                        str(dict["负荷运行曲线图绘制日期"]).split('-')[2] + "日"
                    # 加载图片
                    image = Image(f'image{num}.png')
                    # 将图片插入到目标单元格中
                    sheet.add_image(image, 'B9')

                    workbook.save('直梯.xlsx')
                    workbook.close()

                # 运行异步主程序
                asyncio.run(main())
        except Exception as e:
            print(numtest1, e)

        try:
            for num in range(0, dict['电梯数量']):
                first_column = datas_exls.columns[num]
                date_value = datas_exls[first_column].iloc[45]
                if date_value is not None:
                    # 将日期格式化为字符串，例如"2023-01-13"
                    date_str = date_value.strftime("%Y-%m-%d")
                    # 将格式化后的日期字符串存储到字典中
                    dict["电梯噪声测试日期"] = date_str
                else:
                    dict["电梯噪声测试日期"] = ""

                dict["电梯型号规格"] = datas_exls[first_column].iloc[29]
                dict["设备梯号"] = datas_exls[first_column].iloc[0]
                dict["层站门"] = datas_exls[first_column].iloc[8]
                dict["起止层"] = datas_exls[first_column].iloc[9]
                dict["盲层"] = datas_exls[first_column].iloc[10]
                dict["初始层"] = dict["起止层"].rsplit('/')[0]
                dict["终止层"] = dict["起止层"].rsplit('/')[1]

                async def main():
                    global numtest1
                    path = r"./直梯/43-鲁DT-040电梯噪声测试记录（每台一份）.xlsx"
                    save_path = "直梯.xlsx"
                    wb = openpyxl.load_workbook(path)
                    wb2 = openpyxl.load_workbook(save_path)
                    sheetnames = wb.sheetnames

                    tasks = []
                    for numtest, sheetname in enumerate(sheetnames, start=1):
                        sheet = wb[sheetname]
                        sheet2 = wb2.create_sheet(f"第{numtest1}页")
                        print(numtest1)
                        numtest1 += 1
                        task = asyncio.create_task(copy_sheet(sheet, sheet2))
                        tasks.append(task)

                    # 等待所有任务完成
                    await asyncio.gather(*tasks)

                    # 保存并关闭工作簿
                    wb2.save(save_path)
                    wb.close()
                    wb2.close()
                    workbook = openpyxl.load_workbook('直梯.xlsx')
                    sheet = workbook[f'第{numtest1 - 1}页']
                    cell_g4 = sheet['BG4']
                    cell_g4.value = num
                    if len(str(dict["电梯噪声测试日期"]).split('-')) >= 3:
                        cell_g4 = sheet['AP6']
                        cell_g4.value = str(dict["电梯噪声测试日期"]).split('-')[0] + "年" + \
                                        str(dict["电梯噪声测试日期"]).split('-')[1] + "月" + \
                                        str(dict["电梯噪声测试日期"]).split('-')[2] + "日"
                    cell_g4 = sheet['J6']
                    cell_g4.value = dict["工程名称"]
                    cell_g4 = sheet['J7']
                    cell_g4.value = dict["电梯型号规格"]
                    cell_g4 = sheet['AP7']
                    cell_g4.value = dict["设备梯号"]
                    cell_g4 = sheet['J8']
                    cell_g4.value = dict["层站门"].rsplit('/', 1)[0]
                    cell_g4 = sheet['N33']
                    cell_g4.value = dict["分包单位"]

                    datanum = 15
                    datanum2 = 15
                    for data in range(int(dict["初始层"]), int(dict["终止层"]) + 1):
                        if data == 0:
                            continue
                        if dict["盲层"] != "":
                            mc_values = [int(mc) for mc in dict["盲层"].split(',')]
                            if data in mc_values:
                                continue
                        if datanum2 < 31:
                            cell_g4 = sheet[f'B{datanum}']
                            cell_g4.value = data
                            cell_g4 = sheet[f'D{datanum}']
                            random_float = round(random.uniform(54.6, 61.5), 1)
                            cell_g4.value = random_float
                            cell_g4 = sheet[f'I{datanum}']
                            random_float = round(random.uniform(54.6, 61.5), 1)
                            cell_g4.value = random_float
                            cell_g4 = sheet[f'N{datanum}']
                            random_float = round(random.uniform(40.5, 43), 1)
                            cell_g4.value = random_float
                            cell_g4 = sheet[f'S{datanum}']
                            random_float = round(random.uniform(54.6, 61.5), 1)
                            cell_g4.value = random_float
                            cell_g4 = sheet[f'X{datanum}']
                            random_float = round(random.uniform(54.6, 61.5), 1)
                            cell_g4.value = random_float
                            cell_g4 = sheet[f'AC{datanum}']
                            random_float = round(random.uniform(40.5, 43), 1)
                            cell_g4.value = random_float
                            datanum = datanum + 1
                            datanum2 = datanum2 + 1
                        if datanum2 == 31:
                            datanum = 15
                            datanum2 += 1
                        if datanum2 > 31:
                            cell_g4 = sheet[f'AH{datanum}']
                            cell_g4.value = data
                            cell_g4 = sheet[f'AJ{datanum}']
                            random_float = round(random.uniform(54.6, 61.5), 1)
                            cell_g4.value = random_float
                            cell_g4 = sheet[f'AO{datanum}']
                            random_float = round(random.uniform(54.6, 61.5), 1)
                            cell_g4.value = random_float
                            cell_g4 = sheet[f'AT{datanum}']
                            random_float = round(random.uniform(40.5, 43), 1)
                            cell_g4.value = random_float
                            cell_g4 = sheet[f'AY{datanum}']
                            random_float = round(random.uniform(54.6, 61.5), 1)
                            cell_g4.value = random_float
                            cell_g4 = sheet[f'BD{datanum}']
                            random_float = round(random.uniform(54.6, 61.5), 1)
                            cell_g4.value = random_float
                            cell_g4 = sheet[f'BI{datanum}']
                            random_float = round(random.uniform(40.5, 43), 1)
                            cell_g4.value = random_float
                            datanum = datanum + 1
                            datanum2 = datanum2 + 1

                    workbook.save('直梯.xlsx')
                    workbook.close()

                # 运行异步主程序
                asyncio.run(main())
        except Exception as e:
            print(numtest1, e)

        try:
            async def main():
                global numtest1
                path = r"./直梯/44-鲁DT-043班组自检（互检）记录.xlsx"
                save_path = "直梯.xlsx"
                wb = openpyxl.load_workbook(path)
                wb2 = openpyxl.load_workbook(save_path)
                sheetnames = wb.sheetnames

                tasks = []
                for numtest, sheetname in enumerate(sheetnames, start=1):
                    sheet = wb[sheetname]
                    sheet2 = wb2.create_sheet(f"第{numtest1}页")
                    print(numtest1)
                    numtest1 += 1
                    task = asyncio.create_task(copy_sheet(sheet, sheet2))
                    tasks.append(task)

                # 等待所有任务完成
                await asyncio.gather(*tasks)

                # 保存并关闭工作簿
                wb2.save(save_path)
                wb.close()
                wb2.close()
                workbook = openpyxl.load_workbook('直梯.xlsx')
                sheet = workbook[f'第{numtest1 - 1}页']
                if len(str(dict["操作日期"]).split('-')) >= 3:
                    cell_g4 = sheet['N8']
                    cell_g4.value = str(dict["操作日期"]).split('-')[0] + "年" + \
                                    str(dict["操作日期"]).split('-')[1] + "月" + \
                                    str(dict["操作日期"]).split('-')[2] + "日"
                if len(str(dict["完成日期"]).split('-')) >= 3:
                    cell_g4 = sheet['AS8']
                    cell_g4.value = str(dict["完成日期"]).split('-')[0] + "年" + \
                                    str(dict["完成日期"]).split('-')[1] + "月" + \
                                    str(dict["完成日期"]).split('-')[2] + "日"
                cell_g4 = sheet['N6']
                cell_g4.value = dict["工程名称"]
                cell_g4 = sheet['N7']
                cell_g4.value = dict["自检部位"]
                cell_g4 = sheet['AS7']
                cell_g4.value = dict["自检项目"]
                workbook.save('直梯.xlsx')
                workbook.close()

            # 运行异步主程序
            asyncio.run(main())
        except Exception as e:
            print(numtest1, e)

        try:
            async def main():
                global numtest1
                path = r"./直梯/45-鲁DT-044工序交接检查记录.xlsx"
                save_path = "直梯.xlsx"
                wb = openpyxl.load_workbook(path)
                wb2 = openpyxl.load_workbook(save_path)
                sheetnames = wb.sheetnames

                tasks = []
                for numtest, sheetname in enumerate(sheetnames, start=1):
                    sheet = wb[sheetname]
                    sheet2 = wb2.create_sheet(f"第{numtest1}页")
                    print(numtest1)
                    numtest1 += 1
                    task = asyncio.create_task(copy_sheet(sheet, sheet2))
                    tasks.append(task)

                # 等待所有任务完成
                await asyncio.gather(*tasks)

                # 保存并关闭工作簿
                wb2.save(save_path)
                wb.close()
                wb2.close()
                workbook = openpyxl.load_workbook('直梯.xlsx')
                sheet = workbook[f'第{numtest1 - 1}页']
                if len(str(dict["检查日期"]).split('-')) >= 3:
                    cell_g4 = sheet['AS8']
                    cell_g4.value = str(dict["检查日期"]).split('-')[0] + "年" + \
                                    str(dict["检查日期"]).split('-')[1] + "月" + \
                                    str(dict["检查日期"]).split('-')[2] + "日"
                cell_g4 = sheet['N6']
                cell_g4.value = dict["工程名称"]
                cell_g4 = sheet['N7']
                cell_g4.value = dict["移交部门名称"]
                cell_g4 = sheet['AS7']
                cell_g4.value = dict["接收部门名称"]
                cell_g4 = sheet['N8']
                cell_g4.value = dict["交接部位"]
                workbook.save('直梯.xlsx')
                workbook.close()

            # 运行异步主程序
            asyncio.run(main())
        except Exception as e:
            print(numtest1, e)

        try:
            async def main():
                global numtest1
                path = r"./直梯/46-鲁DT-045技术复核（或预检）记录.xlsx"
                save_path = "直梯.xlsx"
                wb = openpyxl.load_workbook(path)
                wb2 = openpyxl.load_workbook(save_path)
                sheetnames = wb.sheetnames

                tasks = []
                for numtest, sheetname in enumerate(sheetnames, start=1):
                    sheet = wb[sheetname]
                    sheet2 = wb2.create_sheet(f"第{numtest1}页")
                    print(numtest1)
                    numtest1 += 1
                    task = asyncio.create_task(copy_sheet(sheet, sheet2))
                    tasks.append(task)

                # 等待所有任务完成
                await asyncio.gather(*tasks)

                # 保存并关闭工作簿
                wb2.save(save_path)
                wb.close()
                wb2.close()
                workbook = openpyxl.load_workbook('直梯.xlsx')
                sheet = workbook[f'第{numtest1 - 1}页']
                if len(str(dict["复查日期"]).split('-')) >= 3:
                    cell_g4 = sheet['AT7']
                    cell_g4.value = str(dict["复查日期"]).split('-')[0] + "年" + \
                                    str(dict["复查日期"]).split('-')[1] + "月" + \
                                    str(dict["复查日期"]).split('-')[2] + "日"
                if len(str(dict["复查日期"]).split('-')) >= 3:
                    cell_g4 = sheet['AT21']
                    cell_g4.value = str(dict["复查日期"]).split('-')[0] + "年" + \
                                    str(dict["复查日期"]).split('-')[1] + "月" + \
                                    str(dict["复查日期"]).split('-')[2] + "日"
                cell_g4 = sheet['M6']
                cell_g4.value = dict["工程名称"]
                cell_g4 = sheet['AT6']
                cell_g4.value = dict["复核项目"]
                cell_g4 = sheet['M7']
                cell_g4.value = dict["复核部位"]
                cell_g4 = sheet['U22']
                cell_g4.value = dict["分包单位"]
                workbook.save('直梯.xlsx')
                workbook.close()

            # 运行异步主程序
            asyncio.run(main())
        except Exception as e:
            print(numtest1, e)

        try:
            async def main():
                global numtest1
                path = r"./直梯/47-鲁DT-046不符合要求项处理记录.xlsx"
                save_path = "直梯.xlsx"
                wb = openpyxl.load_workbook(path)
                wb2 = openpyxl.load_workbook(save_path)
                sheetnames = wb.sheetnames

                tasks = []
                for numtest, sheetname in enumerate(sheetnames, start=1):
                    sheet = wb[sheetname]
                    sheet2 = wb2.create_sheet(f"第{numtest1}页")
                    print(numtest1)
                    numtest1 += 1
                    task = asyncio.create_task(copy_sheet(sheet, sheet2))
                    tasks.append(task)

                # 等待所有任务完成
                await asyncio.gather(*tasks)

                # 保存并关闭工作簿
                wb2.save(save_path)
                wb.close()
                wb2.close()
                workbook = openpyxl.load_workbook('直梯.xlsx')
                sheet = workbook[f'第{numtest1 - 1}页']
                cell_g4 = sheet['M6']
                cell_g4.value = dict["工程名称"]
                workbook.save('直梯.xlsx')
                workbook.close()

            # 运行异步主程序
            asyncio.run(main())
        except Exception as e:
            print(numtest1, e)

        try:
            async def main():
                global numtest1
                path = r"./直梯/48-鲁DT-047样板间（分项工程）质量检查记录.xlsx"
                save_path = "直梯.xlsx"
                wb = openpyxl.load_workbook(path)
                wb2 = openpyxl.load_workbook(save_path)
                sheetnames = wb.sheetnames

                tasks = []
                for numtest, sheetname in enumerate(sheetnames, start=1):
                    sheet = wb[sheetname]
                    sheet2 = wb2.create_sheet(f"第{numtest1}页")
                    print(numtest1)
                    numtest1 += 1
                    task = asyncio.create_task(copy_sheet(sheet, sheet2))
                    tasks.append(task)

                # 等待所有任务完成
                await asyncio.gather(*tasks)

                # 保存并关闭工作簿
                wb2.save(save_path)
                wb.close()
                wb2.close()
                workbook = openpyxl.load_workbook('直梯.xlsx')
                sheet = workbook[f'第{numtest1 - 1}页']

                cell_g4 = sheet['M6']
                cell_g4.value = dict["工程名称"]
                cell_g4 = sheet['AT6']
                cell_g4.value = dict["建设单位"]
                cell_g4 = sheet['M7']
                cell_g4.value = dict["分包单位"]
                cell_g4 = sheet['AT7']
                cell_g4.value = dict["施工部位"]

                workbook.save('直梯.xlsx')
                workbook.close()

            # 运行异步主程序
            asyncio.run(main())
        except Exception as e:
            print(numtest1, e)

        try:
            async def main():
                global numtest1
                path = r"./直梯/49-鲁DT-048新技术、新设备、新材料、新工艺施工验收记录.xlsx"
                save_path = "直梯.xlsx"
                wb = openpyxl.load_workbook(path)
                wb2 = openpyxl.load_workbook(save_path)
                sheetnames = wb.sheetnames

                tasks = []
                for numtest, sheetname in enumerate(sheetnames, start=1):
                    sheet = wb[sheetname]
                    sheet2 = wb2.create_sheet(f"第{numtest1}页")
                    print(numtest1)
                    numtest1 += 1
                    task = asyncio.create_task(copy_sheet(sheet, sheet2))
                    tasks.append(task)

                # 等待所有任务完成
                await asyncio.gather(*tasks)

                # 保存并关闭工作簿
                wb2.save(save_path)
                wb.close()
                wb2.close()
                workbook = openpyxl.load_workbook('直梯.xlsx')
                sheet = workbook[f'第{numtest1 - 1}页']
                if len(str(dict["施工日期"]).split('-')) >= 3:
                    cell_g4 = sheet['AN7']
                    cell_g4.value = str(dict["施工日期"]).split('-')[0] + "年" + \
                                    str(dict["施工日期"]).split('-')[1] + "月" + \
                                    str(dict["施工日期"]).split('-')[2] + "日"
                cell_g4 = sheet['N6']
                cell_g4.value = dict["子分部工程名称"]
                cell_g4 = sheet['N7']
                cell_g4.value = dict["施工部位"]
                cell_g4 = sheet['P15']
                cell_g4.value = dict["总包单位"]
                workbook.save('直梯.xlsx')
                workbook.close()

            # 运行异步主程序
            asyncio.run(main())
        except Exception as e:
            print(numtest1, e)

        try:
            async def main():
                global numtest1
                path = r"./直梯/50-【第二十二卷】电梯工程施工质量验收资料.xlsx"
                save_path = "直梯.xlsx"
                wb = openpyxl.load_workbook(path)
                wb2 = openpyxl.load_workbook(save_path)
                sheetnames = wb.sheetnames

                tasks = []
                for numtest, sheetname in enumerate(sheetnames, start=1):
                    sheet = wb[sheetname]
                    sheet2 = wb2.create_sheet(f"第{numtest1}页")
                    print(numtest1)
                    numtest1 += 1
                    task = asyncio.create_task(copy_sheet(sheet, sheet2))
                    tasks.append(task)

                # 等待所有任务完成
                await asyncio.gather(*tasks)

                # 保存并关闭工作簿
                wb2.save(save_path)
                wb.close()
                wb2.close()
                workbook = openpyxl.load_workbook('直梯.xlsx')
                sheet = workbook[f'第{numtest1 - 1}页']
                cell_g4 = sheet['K4']
                cell_g4.value = dict["工程名称"]
                workbook.save('直梯.xlsx')
                workbook.close()

            # 运行异步主程序
            asyncio.run(main())
        except Exception as e:
            print(numtest1, e)

        try:
            async def main():
                global numtest1
                path = r"./直梯/51-鲁DT-049_______分部（子分部）工程质量验收记录.xlsx"
                save_path = "直梯.xlsx"
                wb = openpyxl.load_workbook(path)
                wb2 = openpyxl.load_workbook(save_path)
                sheetnames = wb.sheetnames

                tasks = []
                for numtest, sheetname in enumerate(sheetnames, start=1):
                    sheet = wb[sheetname]
                    sheet2 = wb2.create_sheet(f"第{numtest1}页")
                    print(numtest1)
                    numtest1 += 1
                    task = asyncio.create_task(copy_sheet(sheet, sheet2))
                    tasks.append(task)

                # 等待所有任务完成
                await asyncio.gather(*tasks)

                # 保存并关闭工作簿
                wb2.save(save_path)
                wb.close()
                wb2.close()
                workbook = openpyxl.load_workbook('直梯.xlsx')
                sheet = workbook[f'第{numtest1 - 1}页']
                cell_g4 = sheet['L6']
                cell_g4.value = dict["子分部工程名称"]
                cell_g4 = sheet['AJ6']
                cell_g4.value = dict["子分部工程数量"]
                cell_g4 = sheet['BC6']
                cell_g4.value = dict["分项工程数量"]
                cell_g4 = sheet['L7']
                cell_g4.value = dict["总包单位"]
                cell_g4 = sheet['AJ7']
                cell_g4.value = dict["总包单位项目负责人"]
                cell_g4 = sheet['BC7']
                cell_g4.value = dict["总包单位技术负责人"]
                cell_g4 = sheet['L8']
                cell_g4.value = dict["分包单位"]
                cell_g4 = sheet['AJ8']
                cell_g4.value = dict["分包单位负责人"]
                cell_g4 = sheet['BC8']
                cell_g4.value = dict["分包内容"]
                workbook.save('直梯.xlsx')
                workbook.close()

            # 运行异步主程序
            asyncio.run(main())
        except Exception as e:
            print(numtest1, e)

        try:
            async def main():
                global numtest1
                path = r"./直梯/52-鲁DT-050电梯分部工程质量控制资料核查记录.xlsx"
                save_path = "直梯.xlsx"
                wb = openpyxl.load_workbook(path)
                wb2 = openpyxl.load_workbook(save_path)
                sheetnames = wb.sheetnames

                tasks = []
                for numtest, sheetname in enumerate(sheetnames, start=1):
                    sheet = wb[sheetname]
                    sheet2 = wb2.create_sheet(f"第{numtest1}页")
                    print(numtest1)
                    numtest1 += 1
                    task = asyncio.create_task(copy_sheet(sheet, sheet2))
                    tasks.append(task)

                # 等待所有任务完成
                await asyncio.gather(*tasks)

                # 保存并关闭工作簿
                wb2.save(save_path)
                wb.close()
                wb2.close()
                workbook = openpyxl.load_workbook('直梯.xlsx')
                sheet = workbook[f'第{numtest1 - 1}页']
                cell_g4 = sheet['L6']
                cell_g4.value = dict["工程名称"]
                cell_g4 = sheet['AO6']
                cell_g4.value = dict["总包单位"]
                cell_g4 = sheet['P27']
                cell_g4.value = dict["总包单位项目负责人"]
                cell_g4 = sheet['AQ27']
                cell_g4.value = dict["总监理工程师"]

                workbook.save('直梯.xlsx')
                workbook.close()

            # 运行异步主程序
            asyncio.run(main())
        except Exception as e:
            print(numtest1, e)

        try:
            async def main():
                global numtest1
                path = r"./直梯/53-鲁DT-051电梯分部工程安全和功能检验资料核查及主要功能抽查记录.xlsx"
                save_path = "直梯.xlsx"
                wb = openpyxl.load_workbook(path)
                wb2 = openpyxl.load_workbook(save_path)
                sheetnames = wb.sheetnames

                tasks = []
                for numtest, sheetname in enumerate(sheetnames, start=1):
                    sheet = wb[sheetname]
                    sheet2 = wb2.create_sheet(f"第{numtest1}页")
                    print(numtest1)
                    numtest1 += 1
                    task = asyncio.create_task(copy_sheet(sheet, sheet2))
                    tasks.append(task)

                # 等待所有任务完成
                await asyncio.gather(*tasks)

                # 保存并关闭工作簿
                wb2.save(save_path)
                wb.close()
                wb2.close()
                workbook = openpyxl.load_workbook('直梯.xlsx')
                sheet = workbook[f'第{numtest1 - 1}页']

                cell_g4 = sheet['L6']
                cell_g4.value = dict["工程名称"]
                cell_g4 = sheet['AP6']
                cell_g4.value = dict["总包单位"]
                cell_g4 = sheet['P32']
                cell_g4.value = dict["总包单位项目负责人"]
                cell_g4 = sheet['AQ32']
                cell_g4.value = dict["总监理工程师"]

                workbook.save('直梯.xlsx')
                workbook.close()

            # 运行异步主程序
            asyncio.run(main())
        except Exception as e:
            print(numtest1, e)

        try:
            async def main():
                global numtest1
                path = r"./直梯/54-鲁DT-052电梯分部工程观感质量检查记录.xlsx"
                save_path = "直梯.xlsx"
                wb = openpyxl.load_workbook(path)
                wb2 = openpyxl.load_workbook(save_path)
                sheetnames = wb.sheetnames

                tasks = []
                for numtest, sheetname in enumerate(sheetnames, start=1):
                    sheet = wb[sheetname]
                    sheet2 = wb2.create_sheet(f"第{numtest1}页")
                    print(numtest1)
                    numtest1 += 1
                    task = asyncio.create_task(copy_sheet(sheet, sheet2))
                    tasks.append(task)

                # 等待所有任务完成
                await asyncio.gather(*tasks)

                # 保存并关闭工作簿
                wb2.save(save_path)
                wb.close()
                wb2.close()
                workbook = openpyxl.load_workbook('直梯.xlsx')
                sheet = workbook[f'第{numtest1 - 1}页']

                cell_g4 = sheet['K6']
                cell_g4.value = dict["工程名称"]
                cell_g4 = sheet['AS6']
                cell_g4.value = dict["总包单位"]
                cell_g4 = sheet['P30']
                cell_g4.value = dict["总包单位项目负责人"]
                cell_g4 = sheet['AQ30']
                cell_g4.value = dict["总监理工程师"]

                workbook.save('直梯.xlsx')
                workbook.close()

            # 运行异步主程序
            asyncio.run(main())
        except Exception as e:
            print(numtest1, e)

        try:
            for num in range(0, dict['电梯数量']):
                first_column = datas_exls.columns[num]
                date_value = datas_exls[first_column].iloc[45]
                if date_value is not None:
                    # 将日期格式化为字符串，例如"2023-01-13"
                    date_str = date_value.strftime("%Y-%m-%d")
                    # 将格式化后的日期字符串存储到字典中
                    dict["电梯噪声测试日期"] = date_str
                else:
                    dict["电梯噪声测试日期"] = ""

                dict["电梯型号规格"] = datas_exls[first_column].iloc[29]
                dict["设备梯号"] = datas_exls[first_column].iloc[0]
                dict["层站门"] = datas_exls[first_column].iloc[8]
                dict["起止层"] = datas_exls[first_column].iloc[9]
                dict["盲层"] = datas_exls[first_column].iloc[10]
                dict["初始层"] = dict["起止层"].rsplit('/')[0]
                dict["终止层"] = dict["起止层"].rsplit('/')[1]

                async def main():
                    global numtest1
                    path = r"./直梯/55-鲁DT-053.1电梯工程规范强制性条文检查记录（一）（每台一份）.xlsx"
                    save_path = "直梯.xlsx"
                    wb = openpyxl.load_workbook(path)
                    wb2 = openpyxl.load_workbook(save_path)
                    sheetnames = wb.sheetnames

                    tasks = []
                    for numtest, sheetname in enumerate(sheetnames, start=1):
                        sheet = wb[sheetname]
                        sheet2 = wb2.create_sheet(f"第{numtest1}页")
                        print(numtest1)
                        numtest1 += 1
                        task = asyncio.create_task(copy_sheet(sheet, sheet2))
                        tasks.append(task)

                    # 等待所有任务完成
                    await asyncio.gather(*tasks)

                    # 保存并关闭工作簿
                    wb2.save(save_path)
                    wb.close()
                    wb2.close()
                    workbook = openpyxl.load_workbook('直梯.xlsx')
                    sheet = workbook[f'第{numtest1 - 1}页']
                    cell_g4 = sheet['BF4']
                    cell_g4.value = num
                    cell_g4 = sheet['N6']
                    cell_g4.value = dict["工程名称"]
                    cell_g4 = sheet['AT6']
                    cell_g4.value = dict["设备梯号"]
                    cell_g4 = sheet['N7']
                    cell_g4.value = dict["总包单位"]
                    cell_g4 = sheet['AT7']
                    cell_g4.value = dict["总包单位项目负责人"]
                    cell_g4 = sheet['N8']
                    cell_g4.value = dict["总包单位"]
                    cell_g4 = sheet['AT8']
                    cell_g4.value = dict["分包单位项目负责人"]
                    cell_g4 = sheet['N9']
                    cell_g4.value = dict["监理单位"]
                    cell_g4 = sheet['AT9']
                    cell_g4.value = dict["总监理工程师"]

                    workbook.save('直梯.xlsx')
                    workbook.close()

                # 运行异步主程序
                asyncio.run(main())
        except Exception as e:
            print(numtest1, e)

        try:
            for num in range(0, dict['电梯数量']):
                first_column = datas_exls.columns[num]
                date_value = datas_exls[first_column].iloc[45]
                if date_value is not None:
                    # 将日期格式化为字符串，例如"2023-01-13"
                    date_str = date_value.strftime("%Y-%m-%d")
                    # 将格式化后的日期字符串存储到字典中
                    dict["电梯噪声测试日期"] = date_str
                else:
                    dict["电梯噪声测试日期"] = ""

                dict["电梯型号规格"] = datas_exls[first_column].iloc[29]
                dict["设备梯号"] = datas_exls[first_column].iloc[0]
                dict["层站门"] = datas_exls[first_column].iloc[8]
                dict["起止层"] = datas_exls[first_column].iloc[9]
                dict["盲层"] = datas_exls[first_column].iloc[10]
                dict["初始层"] = dict["起止层"].rsplit('/')[0]
                dict["终止层"] = dict["起止层"].rsplit('/')[1]

                async def main():
                    global numtest1
                    path = r"./直梯/56-鲁DT-053.2电梯工程规范强制性条文检查记录（二)（每台一份）.xlsx"
                    save_path = "直梯.xlsx"
                    wb = openpyxl.load_workbook(path)
                    wb2 = openpyxl.load_workbook(save_path)
                    sheetnames = wb.sheetnames


                    tasks = []
                    for numtest, sheetname in enumerate(sheetnames, start=1):
                        sheet = wb[sheetname]
                        sheet2 = wb2.create_sheet(f"第{numtest1}页")
                        print(numtest1)
                        numtest1 += 1
                        task = asyncio.create_task(copy_sheet(sheet, sheet2))
                        tasks.append(task)

                    # 等待所有任务完成
                    await asyncio.gather(*tasks)

                    # 保存并关闭工作簿
                    wb2.save(save_path)
                    wb.close()
                    wb2.close()
                    workbook = openpyxl.load_workbook('直梯.xlsx')
                    sheet = workbook[f'第{numtest1 - 1}页']
                    cell_g4 = sheet['BF4']
                    cell_g4.value = num
                    cell_g4 = sheet['N6']
                    cell_g4.value = dict["工程名称"]
                    cell_g4 = sheet['AT6']
                    cell_g4.value = dict["设备梯号"]
                    cell_g4 = sheet['N7']
                    cell_g4.value = dict["总包单位"]
                    cell_g4 = sheet['AT7']
                    cell_g4.value = dict["总包单位项目负责人"]
                    cell_g4 = sheet['N8']
                    cell_g4.value = dict["分包单位"]
                    cell_g4 = sheet['AT8']
                    cell_g4.value = dict["分包单位项目负责人"]
                    cell_g4 = sheet['N9']
                    cell_g4.value = dict["监理单位"]
                    cell_g4 = sheet['AT9']
                    cell_g4.value = dict["总监理工程师"]

                    workbook.save('直梯.xlsx')
                    workbook.close()

                # 运行异步主程序
                asyncio.run(main())
        except Exception as e:
            print(numtest1, e)

        try:
            for num in range(0, dict['电梯数量']):
                first_column = datas_exls.columns[num]
                date_value = datas_exls[first_column].iloc[45]
                if date_value is not None:
                    # 将日期格式化为字符串，例如"2023-01-13"
                    date_str = date_value.strftime("%Y-%m-%d")
                    # 将格式化后的日期字符串存储到字典中
                    dict["电梯噪声测试日期"] = date_str
                else:
                    dict["电梯噪声测试日期"] = ""

                dict["电梯型号规格"] = datas_exls[first_column].iloc[29]
                dict["设备梯号"] = datas_exls[first_column].iloc[0]
                dict["层站门"] = datas_exls[first_column].iloc[8]
                dict["起止层"] = datas_exls[first_column].iloc[9]
                dict["盲层"] = datas_exls[first_column].iloc[10]
                dict["初始层"] = dict["起止层"].rsplit('/')[0]
                dict["终止层"] = dict["起止层"].rsplit('/')[1]

                async def main():
                    global numtest1
                    path = r"./直梯/57-鲁DT-_______分项工程质量验收记录（每台一份）.xlsx"
                    save_path = "直梯.xlsx"
                    wb = openpyxl.load_workbook(path)
                    wb2 = openpyxl.load_workbook(save_path)
                    sheetnames = wb.sheetnames

                    tasks = []
                    for numtest, sheetname in enumerate(sheetnames, start=1):
                        sheet = wb[sheetname]
                        sheet2 = wb2.create_sheet(f"第{numtest1}页")
                        print(numtest1)
                        numtest1 += 1
                        task = asyncio.create_task(copy_sheet(sheet, sheet2))
                        tasks.append(task)

                    # 等待所有任务完成
                    await asyncio.gather(*tasks)

                    # 保存并关闭工作簿
                    wb2.save(save_path)
                    wb.close()
                    wb2.close()
                    workbook = openpyxl.load_workbook('直梯.xlsx')
                    sheet = workbook[f'第{numtest1 - 1}页']
                    cell_g4 = sheet['BF4']
                    cell_g4.value = num
                    cell_g4 = sheet['L2']
                    cell_g4.value = dict["分项工程名称"]
                    cell_g4 = sheet['L6']
                    cell_g4.value = dict["子分部工程名称"]
                    cell_g4 = sheet['AS6']
                    cell_g4.value = dict["子分部工程名称"]
                    cell_g4 = sheet['L7']
                    cell_g4.value = dict["分项工程数量"]

                    cell_g4 = sheet['L8']
                    cell_g4.value = dict["总包单位"]
                    cell_g4 = sheet['AM8']
                    cell_g4.value = dict["总包单位项目负责人"]
                    cell_g4 = sheet['BE8']
                    cell_g4.value = dict["总包单位技术负责人"]
                    cell_g4 = sheet['L9']
                    cell_g4.value = dict["分包单位"]
                    cell_g4 = sheet['AM9']
                    cell_g4.value = dict["分包单位项目负责人"]
                    cell_g4 = sheet['BE9']
                    cell_g4.value = dict["分包内容"]
                    cell_g4 = sheet['AZ27']
                    cell_g4.value = dict["建设单位项目负责人"]
                    workbook.save('直梯.xlsx')
                    workbook.close()

                # 运行异步主程序
                asyncio.run(main())
        except Exception as e:
            print(numtest1, e)

        try:
            async def main():
                global numtest1
                path = r"./直梯/58-鲁DT-_______检验批现场验收检查原始记录.xlsx"
                save_path = "直梯.xlsx"
                wb = openpyxl.load_workbook(path)
                wb2 = openpyxl.load_workbook(save_path)
                sheetnames = wb.sheetnames

                tasks = []
                for numtest, sheetname in enumerate(sheetnames, start=1):
                    sheet = wb[sheetname]
                    sheet2 = wb2.create_sheet(f"第{numtest1}页")
                    print(numtest1)
                    numtest1 += 1
                    task = asyncio.create_task(copy_sheet(sheet, sheet2))
                    tasks.append(task)

                # 等待所有任务完成
                await asyncio.gather(*tasks)

                # 保存并关闭工作簿
                wb2.save(save_path)
                wb.close()
                wb2.close()
                workbook = openpyxl.load_workbook('直梯.xlsx')
                sheet = workbook[f'第{numtest1 - 1}页']
                if len(str(dict["检查日期"]).split('-')) >= 3:
                    cell_g4 = sheet['AV23']
                    cell_g4.value = str(dict["检查日期"]).split('-')[0] + "年" + \
                                    str(dict["检查日期"]).split('-')[1] + "月" + \
                                    str(dict["检查日期"]).split('-')[2] + "日"

                cell_g4 = sheet['N2']
                cell_g4.value = dict["检验名字"]
                cell_g4 = sheet['L6']
                cell_g4.value = dict["子分部工程名称"]
                cell_g4 = sheet['L7']
                cell_g4.value = dict["检验批名称"]
                cell_g4 = sheet['AX7']
                cell_g4.value = dict["检验批编号"]
                cell_g4 = sheet['T23']
                cell_g4.value = dict["建设单位项目负责人"]

                workbook.save('直梯.xlsx')
                workbook.close()

            # 运行异步主程序
            asyncio.run(main())
        except Exception as e:
            print(numtest1, e)

        try:
            for num in range(0, dict['电梯数量']):
                first_column = datas_exls.columns[num]
                date_value = datas_exls[first_column].iloc[45]
                if date_value is not None:
                    # 将日期格式化为字符串，例如"2023-01-13"
                    date_str = date_value.strftime("%Y-%m-%d")
                    # 将格式化后的日期字符串存储到字典中
                    dict["电梯噪声测试日期"] = date_str
                else:
                    dict["电梯噪声测试日期"] = ""
                dict["检验区域"] = datas_exls[first_column].iloc[60]
                dict["电梯型号规格"] = datas_exls[first_column].iloc[29]
                dict["设备梯号"] = datas_exls[first_column].iloc[0]
                dict["补偿链配备（有/无））"] = datas_exls[first_column].iloc[19]
                dict["有机房/无机房"] = datas_exls[first_column].iloc[5]
                dict["层站门"] = datas_exls[first_column].iloc[8]
                dict["起止层"] = datas_exls[first_column].iloc[9]
                dict["盲层"] = datas_exls[first_column].iloc[10]
                dict["初始层"] = dict["起止层"].rsplit('/')[0]
                dict["终止层"] = dict["起止层"].rsplit('/')[1]

                async def main():
                    global numtest1
                    path = r"./直梯/59.1-鲁DT-_  电气检验批质量验收记录（每台一份）.xlsx"
                    save_path = "直梯.xlsx"
                    wb = openpyxl.load_workbook(path)
                    wb2 = openpyxl.Workbook()
                    sheetnames = wb.sheetnames

                    tasks = []
                    for numtest, sheetname in enumerate(sheetnames, start=1):
                        sheet = wb[sheetname]
                        sheet2 = wb2.create_sheet(f"第{numtest1}页")
                        print(numtest1)
                        numtest1 += 1
                        task = asyncio.create_task(copy_sheet(sheet, sheet2))
                        tasks.append(task)

                    # 等待所有任务完成
                    await asyncio.gather(*tasks)

                    # 保存并关闭工作簿
                    wb2.save(save_path)
                    wb.close()
                    wb2.close()
                    workbook = openpyxl.load_workbook('直梯.xlsx')
                    sheet = workbook[f'第{numtest1 - 1}页']
                    cell_g4 = sheet['BF4']
                    cell_g4.value = num
                    cell_g4 = sheet['L6']
                    cell_g4.value = dict["子分部工程名称"]
                    cell_g4 = sheet['AG6']
                    cell_g4.value = dict["子分部工程名称"]
                    cell_g4 = sheet['AZ6']
                    cell_g4.value = dict["分项工程名称"]
                    cell_g4 = sheet['L7']
                    cell_g4.value = dict["总包单位"]
                    cell_g4 = sheet['AG7']
                    cell_g4.value = dict["总包单位项目负责人"]
                    cell_g4 = sheet['L8']
                    cell_g4.value = dict["分包单位"]
                    cell_g4 = sheet['AG8']
                    cell_g4.value = dict["分包单位项目负责人"]
                    cell_g4 = sheet['AZ8']
                    cell_g4.value = dict["检验区域"]
                    cell_g4 = sheet['AZ32']
                    cell_g4.value = dict["建设单位项目负责人"]
                    workbook.save('直梯.xlsx')
                    workbook.close()

                # 运行异步主程序
                asyncio.run(main())

                async def main():
                    global numtest1
                    path = r"./直梯/59.2-鲁DT-_  悬挂装置检验批质量验收记录（每台一份）.xlsx"
                    save_path = "直梯.xlsx"
                    wb = openpyxl.load_workbook(path)
                    wb2 = openpyxl.Workbook()
                    sheetnames = wb.sheetnames

                    tasks = []
                    for numtest, sheetname in enumerate(sheetnames, start=1):
                        sheet = wb[sheetname]
                        sheet2 = wb2.create_sheet(f"第{numtest1}页")
                        print(numtest1)
                        numtest1 += 1
                        task = asyncio.create_task(copy_sheet(sheet, sheet2))
                        tasks.append(task)

                    # 等待所有任务完成
                    await asyncio.gather(*tasks)

                    # 保存并关闭工作簿
                    wb2.save(save_path)
                    wb.close()
                    wb2.close()
                    workbook = openpyxl.load_workbook('直梯.xlsx')
                    sheet = workbook[f'第{numtest1 - 1}页']
                    cell_g4 = sheet['BF4']
                    cell_g4.value = num
                    cell_g4 = sheet['L6']
                    cell_g4.value = dict["子分部工程名称"]
                    cell_g4 = sheet['AG6']
                    cell_g4.value = dict["子分部工程名称"]
                    cell_g4 = sheet['AZ6']
                    cell_g4.value = dict["分项工程名称"]
                    cell_g4 = sheet['L7']
                    cell_g4.value = dict["总包单位"]
                    cell_g4 = sheet['AG7']
                    cell_g4.value = dict["总包单位项目负责人"]
                    cell_g4 = sheet['AO23']
                    cell_g4.value = "符合要求" if "有" in dict["补偿链配备（有/无））"] else "/"
                    cell_g4 = sheet['BG23']
                    cell_g4.value = "合格" if "有" in dict["补偿链配备（有/无））"] else "/"
                    cell_g4 = sheet['L8']
                    cell_g4.value = dict["分包单位"]
                    cell_g4 = sheet['AG8']
                    cell_g4.value = dict["分包单位项目负责人"]
                    cell_g4 = sheet['AZ8']
                    cell_g4.value = dict["检验区域"]
                    cell_g4 = sheet['AZ32']
                    cell_g4.value = dict["建设单位项目负责人"]
                    workbook.save('直梯.xlsx')
                    workbook.close()

                # 运行异步主程序
                asyncio.run(main())

                async def main():
                    global numtest1
                    path = r"./直梯/59.3-鲁DT-_  整机安装批质量验收记录（每台一份）.xlsx"
                    save_path = "直梯.xlsx"
                    wb = openpyxl.load_workbook(path)
                    wb2 = openpyxl.Workbook()
                    sheetnames = wb.sheetnames

                    tasks = []
                    for numtest, sheetname in enumerate(sheetnames, start=1):
                        sheet = wb[sheetname]
                        sheet2 = wb2.create_sheet(f"第{numtest1}页")
                        print(numtest1)
                        numtest1 += 1
                        task = asyncio.create_task(copy_sheet(sheet, sheet2))
                        tasks.append(task)

                    # 等待所有任务完成
                    await asyncio.gather(*tasks)

                    # 保存并关闭工作簿
                    wb2.save(save_path)
                    wb.close()
                    wb2.close()
                    workbook = openpyxl.load_workbook('直梯.xlsx')
                    sheet = workbook[f'第{numtest1 - 1}页']
                    cell_g4 = sheet['BF4']
                    cell_g4.value = num
                    cell_g4 = sheet['L6']
                    cell_g4.value = dict["子分部工程名称"]
                    cell_g4 = sheet['AG6']
                    cell_g4.value = dict["子分部工程名称"]
                    cell_g4 = sheet['AZ6']
                    cell_g4.value = dict["分项工程名称"]
                    cell_g4 = sheet['L7']
                    cell_g4.value = dict["总包单位"]
                    cell_g4 = sheet['AG7']
                    cell_g4.value = dict["总包单位项目负责人"]
                    cell_g4 = sheet['AZ7']
                    cell_g4 = sheet['L8']
                    cell_g4.value = dict["分包单位"]
                    cell_g4 = sheet['AG8']
                    cell_g4.value = dict["分包单位项目负责人"]
                    cell_g4 = sheet['AZ8']
                    cell_g4.value = dict["检验区域"]
                    cell_g4 = sheet['AZ32']
                    cell_g4.value = dict["建设单位项目负责人"]
                    workbook.save('直梯.xlsx')
                    workbook.close()

                # 运行异步主程序
                asyncio.run(main())

                async def main():
                    global numtest1
                    path = r"./直梯/59.4-鲁DT-_ 安全部件检验批质量验收记录（每台一份）.xlsx"
                    save_path = "直梯.xlsx"
                    wb = openpyxl.load_workbook(path)
                    wb2 = openpyxl.Workbook()
                    sheetnames = wb.sheetnames

                    tasks = []
                    for numtest, sheetname in enumerate(sheetnames, start=1):
                        sheet = wb[sheetname]
                        sheet2 = wb2.create_sheet(f"第{numtest1}页")
                        print(numtest1)
                        numtest1 += 1
                        task = asyncio.create_task(copy_sheet(sheet, sheet2))
                        tasks.append(task)

                    # 等待所有任务完成
                    await asyncio.gather(*tasks)

                    # 保存并关闭工作簿
                    wb2.save(save_path)
                    wb.close()
                    wb2.close()
                    workbook = openpyxl.load_workbook('直梯.xlsx')
                    sheet = workbook[f'第{numtest1 - 1}页']
                    cell_g4 = sheet['BF4']
                    cell_g4.value = num
                    cell_g4 = sheet['L6']
                    cell_g4.value = dict["子分部工程名称"]
                    cell_g4 = sheet['AG6']
                    cell_g4.value = dict["子分部工程名称"]
                    cell_g4 = sheet['AZ6']
                    cell_g4.value = dict["分项工程名称"]
                    cell_g4 = sheet['L7']
                    cell_g4.value = dict["总包单位"]
                    cell_g4 = sheet['AG7']
                    cell_g4.value = dict["总包单位项目负责人"]
                    cell_g4 = sheet['L8']
                    cell_g4.value = dict["分包单位"]
                    cell_g4 = sheet['AG8']
                    cell_g4.value = dict["分包单位项目负责人"]
                    cell_g4 = sheet['AZ8']
                    cell_g4.value = dict["检验区域"]
                    cell_g4 = sheet['AZ32']
                    cell_g4.value = dict["建设单位项目负责人"]
                    workbook.save('直梯.xlsx')
                    workbook.close()

                # 运行异步主程序
                asyncio.run(main())

                async def main():
                    global numtest1
                    path = r"./直梯/59.5-鲁DT-_ 轿厢_检验批质量验收记录（每台一份）.xlsx"
                    save_path = "直梯.xlsx"
                    wb = openpyxl.load_workbook(path)
                    wb2 = openpyxl.Workbook()
                    sheetnames = wb.sheetnames

                    tasks = []
                    for numtest, sheetname in enumerate(sheetnames, start=1):
                        sheet = wb[sheetname]
                        sheet2 = wb2.create_sheet(f"第{numtest1}页")
                        print(numtest1)
                        numtest1 += 1
                        task = asyncio.create_task(copy_sheet(sheet, sheet2))
                        tasks.append(task)

                    # 等待所有任务完成
                    await asyncio.gather(*tasks)

                    # 保存并关闭工作簿
                    wb2.save(save_path)
                    wb.close()
                    wb2.close()
                    workbook = openpyxl.load_workbook('直梯.xlsx')
                    sheet = workbook[f'第{numtest1 - 1}页']
                    cell_g4 = sheet['BF4']
                    cell_g4.value = num
                    cell_g4 = sheet['L6']
                    cell_g4.value = dict["子分部工程名称"]
                    cell_g4 = sheet['AG6']
                    cell_g4.value = dict["子分部工程名称"]
                    cell_g4 = sheet['AZ6']
                    cell_g4.value = dict["分项工程名称"]
                    cell_g4 = sheet['L7']
                    cell_g4.value = dict["总包单位"]
                    cell_g4 = sheet['AG7']
                    cell_g4.value = dict["总包单位项目负责人"]
                    cell_g4 = sheet['L8']
                    cell_g4.value = dict["分包单位"]
                    cell_g4 = sheet['AG8']
                    cell_g4.value = dict["分包单位项目负责人"]
                    cell_g4 = sheet['AZ8']
                    cell_g4.value = dict["检验区域"]
                    cell_g4 = sheet['AZ32']
                    cell_g4.value = dict["建设单位项目负责人"]
                    workbook.save('直梯.xlsx')
                    workbook.close()

                # 运行异步主程序
                asyncio.run(main())

                async def main():
                    global numtest1
                    path = r"./直梯/59.6-鲁DT-_ 门系统_检验批质量验收记录（每台一份）.xlsx"
                    save_path = "直梯.xlsx"
                    wb = openpyxl.load_workbook(path)
                    wb2 = openpyxl.Workbook()
                    sheetnames = wb.sheetnames

                    tasks = []
                    for numtest, sheetname in enumerate(sheetnames, start=1):
                        sheet = wb[sheetname]
                        sheet2 = wb2.create_sheet(f"第{numtest1}页")
                        print(numtest1)
                        numtest1 += 1
                        task = asyncio.create_task(copy_sheet(sheet, sheet2))
                        tasks.append(task)

                    # 等待所有任务完成
                    await asyncio.gather(*tasks)

                    # 保存并关闭工作簿
                    wb2.save(save_path)
                    wb.close()
                    wb2.close()
                    workbook = openpyxl.load_workbook('直梯.xlsx')
                    sheet = workbook[f'第{numtest1 - 1}页']
                    cell_g4 = sheet['BF4']
                    cell_g4.value = num
                    cell_g4 = sheet['L6']
                    cell_g4.value = dict["子分部工程名称"]
                    cell_g4 = sheet['AG6']
                    cell_g4.value = dict["子分部工程名称"]
                    cell_g4 = sheet['AZ6']
                    cell_g4.value = dict["分项工程名称"]
                    cell_g4 = sheet['L7']
                    cell_g4.value = dict["总包单位"]
                    cell_g4 = sheet['AG7']
                    cell_g4.value = dict["总包单位项目负责人"]
                    cell_g4 = sheet['L8']
                    cell_g4.value = dict["分包单位"]
                    cell_g4 = sheet['AG8']
                    cell_g4.value = dict["分包单位项目负责人"]
                    cell_g4 = sheet['AZ8']
                    cell_g4.value = dict["检验区域"]
                    cell_g4 = sheet['AZ32']
                    cell_g4.value = dict["建设单位项目负责人"]
                    workbook.save('直梯.xlsx')
                    workbook.close()

                # 运行异步主程序
                asyncio.run(main())

                async def main():
                    global numtest1
                    path = r"./直梯/59.7-鲁DT-_导轨安装_检验批质量验收记录（每台一份）.xlsx"
                    save_path = "直梯.xlsx"
                    wb = openpyxl.load_workbook(path)
                    wb2 = openpyxl.Workbook()
                    sheetnames = wb.sheetnames

                    tasks = []
                    for numtest, sheetname in enumerate(sheetnames, start=1):
                        sheet = wb[sheetname]
                        sheet2 = wb2.create_sheet(f"第{numtest1}页")
                        print(numtest1)
                        numtest1 += 1
                        task = asyncio.create_task(copy_sheet(sheet, sheet2))
                        tasks.append(task)

                    # 等待所有任务完成
                    await asyncio.gather(*tasks)

                    # 保存并关闭工作簿
                    wb2.save(save_path)
                    wb.close()
                    wb2.close()
                    workbook = openpyxl.load_workbook('直梯.xlsx')
                    sheet = workbook[f'第{numtest1 - 1}页']
                    cell_g4 = sheet['BF4']
                    cell_g4.value = num
                    cell_g4 = sheet['L6']
                    cell_g4.value = dict["子分部工程名称"]
                    cell_g4 = sheet['AG6']
                    cell_g4.value = dict["子分部工程名称"]
                    cell_g4 = sheet['AZ6']
                    cell_g4.value = dict["分项工程名称"]
                    cell_g4 = sheet['L7']
                    cell_g4.value = dict["总包单位"]
                    cell_g4 = sheet['AG7']
                    cell_g4.value = dict["总包单位项目负责人"]
                    cell_g4 = sheet['L8']
                    cell_g4.value = dict["分包单位"]
                    cell_g4 = sheet['AG8']
                    cell_g4.value = dict["分包单位项目负责人"]
                    cell_g4 = sheet['AZ8']
                    cell_g4.value = dict["检验区域"]
                    cell_g4 = sheet['AZ32']
                    cell_g4.value = dict["建设单位项目负责人"]
                    workbook.save('直梯.xlsx')
                    workbook.close()

                # 运行异步主程序
                asyncio.run(main())

                async def main():
                    global numtest1
                    path = r"./直梯/59.8-鲁DT-_驱动主机_检验批质量验收记录（每台一份）.xlsx"
                    save_path = "直梯.xlsx"
                    wb = openpyxl.load_workbook(path)
                    wb2 = openpyxl.Workbook()
                    sheetnames = wb.sheetnames

                    tasks = []
                    for numtest, sheetname in enumerate(sheetnames, start=1):
                        sheet = wb[sheetname]
                        sheet2 = wb2.create_sheet(f"第{numtest1}页")
                        print(numtest1)
                        numtest1 += 1
                        task = asyncio.create_task(copy_sheet(sheet, sheet2))
                        tasks.append(task)

                    # 等待所有任务完成
                    await asyncio.gather(*tasks)

                    # 保存并关闭工作簿
                    wb2.save(save_path)
                    wb.close()
                    wb2.close()
                    workbook = openpyxl.load_workbook('直梯.xlsx')
                    sheet = workbook[f'第{numtest1 - 1}页']
                    cell_g4 = sheet['BF4']
                    cell_g4.value = num
                    cell_g4 = sheet['L6']
                    cell_g4.value = dict["子分部工程名称"]
                    cell_g4 = sheet['AG6']
                    cell_g4.value = dict["子分部工程名称"]
                    cell_g4 = sheet['AZ6']
                    cell_g4.value = dict["分项工程名称"]
                    cell_g4 = sheet['L7']
                    cell_g4.value = dict["总包单位"]
                    cell_g4 = sheet['AG7']
                    cell_g4.value = dict["总包单位项目负责人"]
                    cell_g4 = sheet['AZ7']
                    cell_g4 = sheet['L8']
                    cell_g4.value = dict["分包单位"]
                    cell_g4 = sheet['AG8']
                    cell_g4.value = dict["分包单位项目负责人"]
                    cell_g4 = sheet['AZ8']
                    cell_g4.value = dict["检验区域"]
                    cell_g4 = sheet['AO25']
                    cell_g4.value = "符合要求" if "有机房" in dict["有机房/无机房"] else "/"
                    cell_g4 = sheet['BG25']
                    cell_g4.value = "合格" if "有机房" in dict["有机房/无机房"] else "/"
                    cell_g4 = sheet['AZ32']
                    cell_g4.value = dict["建设单位项目负责人"]
                    workbook.save('直梯.xlsx')
                    workbook.close()

                # 运行异步主程序
                asyncio.run(main())

                async def main():
                    global numtest1
                    path = r"./直梯/59.9-鲁DT-_设备进场_检验批质量验收记录（每台一份）.xlsx"
                    save_path = "直梯.xlsx"
                    wb = openpyxl.load_workbook(path)
                    wb2 = openpyxl.Workbook()
                    sheetnames = wb.sheetnames

                    tasks = []
                    for numtest, sheetname in enumerate(sheetnames, start=1):
                        sheet = wb[sheetname]
                        sheet2 = wb2.create_sheet(f"第{numtest1}页")
                        print(numtest1)
                        numtest1 += 1
                        task = asyncio.create_task(copy_sheet(sheet, sheet2))
                        tasks.append(task)

                    # 等待所有任务完成
                    await asyncio.gather(*tasks)

                    # 保存并关闭工作簿
                    wb2.save(save_path)
                    wb.close()
                    wb2.close()
                    workbook = openpyxl.load_workbook('直梯.xlsx')
                    sheet = workbook[f'第{numtest1 - 1}页']
                    cell_g4 = sheet['BF4']
                    cell_g4.value = num
                    cell_g4 = sheet['L6']
                    cell_g4.value = dict["子分部工程名称"]
                    cell_g4 = sheet['AG6']
                    cell_g4.value = dict["子分部工程名称"]
                    cell_g4 = sheet['AZ6']
                    cell_g4.value = dict["分项工程名称"]
                    cell_g4 = sheet['L7']
                    cell_g4.value = dict["总包单位"]
                    cell_g4 = sheet['AG7']
                    cell_g4.value = dict["总包单位项目负责人"]
                    cell_g4 = sheet['L8']
                    cell_g4.value = dict["分包单位"]
                    cell_g4 = sheet['AG8']
                    cell_g4.value = dict["分包单位项目负责人"]
                    cell_g4 = sheet['AZ8']
                    cell_g4.value = dict["检验区域"]
                    cell_g4 = sheet['AZ32']
                    cell_g4.value = dict["建设单位项目负责人"]
                    workbook.save('直梯.xlsx')
                    workbook.close()

                # 运行异步主程序
                asyncio.run(main())

                async def main():
                    global numtest1
                    path = r"./直梯/59.10-鲁DT-_土建交接_检验批质量验收记录（每台一份）.xlsx"
                    save_path = "直梯.xlsx"
                    wb = openpyxl.load_workbook(path)
                    wb2 = openpyxl.Workbook()
                    sheetnames = wb.sheetnames

                    tasks = []
                    for numtest, sheetname in enumerate(sheetnames, start=1):
                        sheet = wb[sheetname]
                        sheet2 = wb2.create_sheet(f"第{numtest1}页")
                        print(numtest1)
                        numtest1 += 1
                        task = asyncio.create_task(copy_sheet(sheet, sheet2))
                        tasks.append(task)

                    # 等待所有任务完成
                    await asyncio.gather(*tasks)

                    # 保存并关闭工作簿
                    wb2.save(save_path)
                    wb.close()
                    wb2.close()
                    workbook = openpyxl.load_workbook('直梯.xlsx')
                    sheet = workbook[f'第{numtest1 - 1}页']
                    cell_g4 = sheet['BF4']
                    cell_g4.value = num
                    cell_g4 = sheet['L6']
                    cell_g4.value = dict["子分部工程名称"]
                    cell_g4 = sheet['AG6']
                    cell_g4.value = dict["子分部工程名称"]
                    cell_g4 = sheet['AZ6']
                    cell_g4.value = dict["分项工程名称"]
                    cell_g4 = sheet['L7']
                    cell_g4.value = dict["总包单位"]
                    cell_g4 = sheet['AG7']
                    cell_g4.value = dict["总包单位项目负责人"]
                    cell_g4 = sheet['L8']
                    cell_g4.value = dict["分包单位"]
                    cell_g4 = sheet['AG8']
                    cell_g4.value = dict["分包单位项目负责人"]
                    cell_g4 = sheet['AZ8']
                    cell_g4.value = dict["检验区域"]
                    cell_g4 = sheet['AO21']
                    cell_g4.value = "符合要求" if "有机房" in dict["有机房/无机房"] else "/"
                    cell_g4 = sheet['BG21']
                    cell_g4.value = "合格" if "有机房" in dict["有机房/无机房"] else "/"
                    cell_g4 = sheet['AZ32']
                    cell_g4.value = dict["建设单位项目负责人"]
                    workbook.save('直梯.xlsx')
                    workbook.close()

                # 运行异步主程序
                asyncio.run(main())
        except Exception as e:
            print(numtest1, e)

        try:
            async def main():
                global numtest1
                path = r"./直梯/60-鲁DT-_______检验批现场验收检查原始记录.xlsx"
                save_path = "直梯.xlsx"
                wb = openpyxl.load_workbook(path)
                wb2 = openpyxl.load_workbook(save_path)
                sheetnames = wb.sheetnames

                tasks = []
                for numtest, sheetname in enumerate(sheetnames, start=1):
                    sheet = wb[sheetname]
                    sheet2 = wb2.create_sheet(f"第{numtest1}页")
                    print(numtest1)
                    numtest1 += 1
                    task = asyncio.create_task(copy_sheet(sheet, sheet2))
                    tasks.append(task)

                # 等待所有任务完成
                await asyncio.gather(*tasks)

                # 保存并关闭工作簿
                wb2.save(save_path)
                wb.close()
                wb2.close()
                workbook = openpyxl.load_workbook('直梯.xlsx')
                sheet = workbook[f'第{numtest1 - 1}页']

                cell_g4 = sheet['L6']
                cell_g4.value = dict["子分部工程名称"]

                workbook.save('直梯.xlsx')
                workbook.close()

            # 运行异步主程序
            asyncio.run(main())
        except Exception as e:
            print(numtest1, e)

        try:
            for num in range(0, dict['电梯数量']):
                first_column = datas_exls.columns[num]
                date_value = datas_exls[first_column].iloc[45]
                if date_value is not None:
                    # 将日期格式化为字符串，例如"2023-01-13"
                    date_str = date_value.strftime("%Y-%m-%d")
                    # 将格式化后的日期字符串存储到字典中
                    dict["电梯噪声测试日期"] = date_str
                else:
                    dict["电梯噪声测试日期"] = ""
                dict["检验区域"] = datas_exls[first_column].iloc[60]
                dict["电梯型号规格"] = datas_exls[first_column].iloc[29]
                dict["设备梯号"] = datas_exls[first_column].iloc[0]
                dict["层站门"] = datas_exls[first_column].iloc[8]
                dict["起止层"] = datas_exls[first_column].iloc[9]
                dict["盲层"] = datas_exls[first_column].iloc[10]
                dict["初始层"] = dict["起止层"].rsplit('/')[0]
                dict["终止层"] = dict["起止层"].rsplit('/')[1]

                async def main():
                    global numtest1
                    path = r"./直梯/61设备进场验收（每台一份）.xlsx"
                    save_path = "直梯.xlsx"
                    wb = openpyxl.load_workbook(path)
                    wb2 = openpyxl.load_workbook(save_path)
                    sheetnames = wb.sheetnames

                    tasks = []
                    for numtest, sheetname in enumerate(sheetnames, start=1):
                        sheet = wb[sheetname]
                        sheet2 = wb2.create_sheet(f"第{numtest1}页")
                        print(numtest1)
                        numtest1 += 1
                        task = asyncio.create_task(copy_sheet(sheet, sheet2))
                        tasks.append(task)

                    # 等待所有任务完成
                    await asyncio.gather(*tasks)

                    # 保存并关闭工作簿
                    wb2.save(save_path)
                    wb.close()
                    wb2.close()
                    workbook = openpyxl.load_workbook('直梯.xlsx')
                    sheet = workbook[f'第{numtest1 - 1}页']
                    cell_g4 = sheet['BF4']
                    cell_g4.value = num
                    cell_g4 = sheet['L6']
                    cell_g4.value = dict["子分部工程名称"]
                    cell_g4 = sheet['AG6']
                    cell_g4.value = dict["子分部工程名称"]
                    cell_g4 = sheet['L7']
                    cell_g4.value = dict["总包单位"]
                    cell_g4 = sheet['AG7']
                    cell_g4.value = dict["总包单位项目负责人"]
                    cell_g4 = sheet['AZ7']

                    cell_g4 = sheet['L8']
                    cell_g4.value = dict["分包单位"]
                    cell_g4 = sheet['AG8']
                    cell_g4.value = dict["分包单位项目负责人"]

                    cell_g4 = sheet['AZ27']
                    cell_g4.value = dict["建设单位项目负责人"]
                    workbook.save('直梯.xlsx')
                    workbook.close()

                # 运行异步主程序
                asyncio.run(main())
        except Exception as e:
            print(numtest1, e)

        try:
            for num in range(0, dict['电梯数量']):
                first_column = datas_exls.columns[num]
                date_value = datas_exls[first_column].iloc[45]
                if date_value is not None:
                    # 将日期格式化为字符串，例如"2023-01-13"
                    date_str = date_value.strftime("%Y-%m-%d")
                    # 将格式化后的日期字符串存储到字典中
                    dict["电梯噪声测试日期"] = date_str
                else:
                    dict["电梯噪声测试日期"] = ""
                dict["检验区域"] = datas_exls[first_column].iloc[60]
                dict["电梯型号规格"] = datas_exls[first_column].iloc[29]
                dict["设备梯号"] = datas_exls[first_column].iloc[0]
                dict["有机房/无机房"] = datas_exls[first_column].iloc[5]
                dict["层站门"] = datas_exls[first_column].iloc[8]
                dict["起止层"] = datas_exls[first_column].iloc[9]
                dict["盲层"] = datas_exls[first_column].iloc[10]
                dict["初始层"] = dict["起止层"].rsplit('/')[0]
                dict["终止层"] = dict["起止层"].rsplit('/')[1]

                path = r"./直梯/62土建交接检验（每台一份）.xlsx"
                save_path = "直梯.xlsx"
                wb = openpyxl.load_workbook(path)
                wb2 = openpyxl.load_workbook(save_path)
                sheetnames = wb.sheetnames
                for sheetname in sheetnames:
                    print(f"第{numtest1}页")
                    sheet = wb[sheetname]
                    sheet2 = wb2.create_sheet(f"第{numtest1}页")
                    # Tab color
                    sheet2.sheet_properties.tabColor = sheet.sheet_properties.tabColor

                    # 复制列宽度
                    for column in sheet.column_dimensions:
                        sheet2.column_dimensions[column].width = sheet.column_dimensions[column].width

                    # 复制行高
                    for row in sheet.row_dimensions:
                        sheet2.row_dimensions[row].height = sheet.row_dimensions[row].height

                    # 处理合并单元格
                    wm = list(sheet.merged_cells)
                    if len(wm) > 0:
                        for i in range(0, len(wm)):
                            cell2 = str(wm[i]).replace('(<CellRange ', '').replace('>,)', '')
                            sheet2.merge_cells(cell2)

                    # 复制行、列和单元格值
                    for i, row in enumerate(sheet.iter_rows(max_row=60)):  # 复制前60行
                        if i >= 60:
                            break
                        for j, cell in enumerate(row):
                            if j >= 100:  # 仅复制前100列
                                break
                            # 获取目标文件中对应单元格的值
                            sheet2.cell(row=i + 1, column=j + 1, value=cell.value)

                            # 如果是合并单元格的一部分，则设置单元格格式和尺寸
                            if cell.coordinate in sheet.merged_cells:
                                for merged_cell in wm:
                                    if cell.coordinate in merged_cell:
                                        first_cell = sheet.cell(merged_cell.min_row, merged_cell.min_col)
                                        target_cell = sheet2.cell(i + 1, j + 1)
                                        sheet2.row_dimensions[i + 1].height = sheet.row_dimensions[
                                            first_cell.row].height
                                        sheet2.column_dimensions[get_column_letter(j + 1)].width = \
                                            sheet.column_dimensions[
                                                get_column_letter(first_cell.column)].width
                                        break

                            # 设置单元格格式
                            source_cell = sheet.cell(i + 1, j + 1)
                            target_cell = sheet2.cell(i + 1, j + 1)

                            # 创建新样式对象
                            font = Font(name=source_cell.font.name, size=source_cell.font.size,
                                        bold=source_cell.font.bold, italic=source_cell.font.italic,
                                        vertAlign=source_cell.font.vertAlign, underline=source_cell.font.underline,
                                        strike=source_cell.font.strike, color='FF000000')  # Set font color to black

                            black_side = Side(color='FF000000')
                            border = Border(
                                left=Side(color=black_side.color,
                                          border_style=source_cell.border.left.border_style),
                                right=Side(color=black_side.color,
                                           border_style=source_cell.border.right.border_style),
                                top=Side(color=black_side.color, border_style=source_cell.border.top.border_style),
                                bottom=Side(color=black_side.color,
                                            border_style=source_cell.border.bottom.border_style))

                            fill = PatternFill(fill_type=source_cell.fill.fill_type,
                                               fgColor=source_cell.fill.fgColor,
                                               bgColor=source_cell.fill.bgColor,
                                               patternType=source_cell.fill.patternType)

                            alignment = Alignment(horizontal=source_cell.alignment.horizontal,
                                                  vertical=source_cell.alignment.vertical,
                                                  text_rotation=source_cell.alignment.text_rotation,
                                                  wrap_text=source_cell.alignment.wrap_text,
                                                  shrink_to_fit=source_cell.alignment.shrink_to_fit,
                                                  indent=source_cell.alignment.indent)

                            protection = Protection(locked=source_cell.protection.locked,
                                                    hidden=source_cell.protection.hidden)

                            # 将新样式对象分配给目标单元格
                            target_cell.font = font
                            target_cell.border = border
                            target_cell.fill = fill
                            target_cell.alignment = alignment
                            target_cell.protection = protection
                            target_cell.number_format = source_cell.number_format
                    numtest1 = numtest1 + 1
                # 保存并关闭工作簿
                wb2.save(save_path)
                wb.close()
                wb2.close()
                workbook = openpyxl.load_workbook('直梯.xlsx')
                sheet = workbook[f'第{numtest1 - 1}页']
                cell_g4 = sheet['BF4']
                cell_g4.value = num
                cell_g4 = sheet['L6']
                cell_g4.value = dict["子分部工程名称"]
                cell_g4 = sheet['AG6']
                cell_g4.value = dict["子分部工程名称"]
                cell_g4 = sheet['L7']
                cell_g4.value = dict["总包单位"]
                cell_g4 = sheet['AG7']
                cell_g4.value = dict["总包单位项目负责人"]

                cell_g4 = sheet['L8']
                cell_g4.value = dict["分包单位"]
                cell_g4 = sheet['AG8']
                cell_g4.value = dict["分包单位项目负责人"]
                cell_g4 = sheet['L9']

                cell_g4 = sheet['BG16']
                cell_g4.value = "合格" if "有机房" in dict["有机房/无机房"] else "/"

                cell_g4 = sheet['AZ26']
                cell_g4.value = dict["建设单位项目负责人"]
                workbook.save('直梯.xlsx')
                workbook.close()
        except Exception as e:
            print(numtest1, e)

        try:
            for num in range(0, dict['电梯数量']):
                first_column = datas_exls.columns[num]
                date_value = datas_exls[first_column].iloc[45]
                if date_value is not None:
                    # 将日期格式化为字符串，例如"2023-01-13"
                    date_str = date_value.strftime("%Y-%m-%d")
                    # 将格式化后的日期字符串存储到字典中
                    dict["电梯噪声测试日期"] = date_str
                else:
                    dict["电梯噪声测试日期"] = ""
                dict["检验区域"] = datas_exls[first_column].iloc[60]
                dict["电梯型号规格"] = datas_exls[first_column].iloc[29]
                dict["设备梯号"] = datas_exls[first_column].iloc[0]
                dict["有机房/无机房"] = datas_exls[first_column].iloc[5]
                dict["层站门"] = datas_exls[first_column].iloc[8]
                dict["起止层"] = datas_exls[first_column].iloc[9]
                dict["盲层"] = datas_exls[first_column].iloc[10]
                dict["初始层"] = dict["起止层"].rsplit('/')[0]
                dict["终止层"] = dict["起止层"].rsplit('/')[1]

                async def main():
                    global numtest1
                    path = r"./直梯/63驱动主机（每台一份）.xlsx"
                    save_path = "直梯.xlsx"
                    wb = openpyxl.load_workbook(path)
                    wb2 = openpyxl.load_workbook(save_path)
                    sheetnames = wb.sheetnames

                    tasks = []
                    for numtest, sheetname in enumerate(sheetnames, start=1):
                        sheet = wb[sheetname]
                        sheet2 = wb2.create_sheet(f"第{numtest1}页")
                        print(numtest1)
                        numtest1 += 1
                        task = asyncio.create_task(copy_sheet(sheet, sheet2))
                        tasks.append(task)

                    # 等待所有任务完成
                    await asyncio.gather(*tasks)

                    # 保存并关闭工作簿
                    wb2.save(save_path)
                    wb.close()
                    wb2.close()
                    workbook = openpyxl.load_workbook('直梯.xlsx')
                    sheet = workbook[f'第{numtest1 - 1}页']
                    cell_g4 = sheet['BF4']
                    cell_g4.value = num
                    cell_g4 = sheet['L6']
                    cell_g4.value = dict["子分部工程名称"]
                    cell_g4 = sheet['AG6']
                    cell_g4.value = dict["子分部工程名称"]
                    cell_g4 = sheet['L7']
                    cell_g4.value = dict["总包单位"]
                    cell_g4 = sheet['AG7']
                    cell_g4.value = dict["总包单位项目负责人"]

                    cell_g4 = sheet['L8']
                    cell_g4.value = dict["分包单位"]
                    cell_g4 = sheet['AG8']
                    cell_g4.value = dict["分包单位项目负责人"]

                    cell_g4 = sheet['BG17']
                    cell_g4.value = "√" if "有机房" in dict["有机房/无机房"] else "/"

                    cell_g4 = sheet['AZ27']
                    cell_g4.value = dict["建设单位项目负责人"]
                    workbook.save('直梯.xlsx')
                    workbook.close()

                # 运行异步主程序
                asyncio.run(main())
        except Exception as e:
            print(numtest1, e)

        try:
            for num in range(0, dict['电梯数量']):
                first_column = datas_exls.columns[num]
                date_value = datas_exls[first_column].iloc[45]
                if date_value is not None:
                    # 将日期格式化为字符串，例如"2023-01-13"
                    date_str = date_value.strftime("%Y-%m-%d")
                    # 将格式化后的日期字符串存储到字典中
                    dict["电梯噪声测试日期"] = date_str
                else:
                    dict["电梯噪声测试日期"] = ""
                dict["检验区域"] = datas_exls[first_column].iloc[60]
                dict["电梯型号规格"] = datas_exls[first_column].iloc[29]
                dict["设备梯号"] = datas_exls[first_column].iloc[0]
                dict["层站门"] = datas_exls[first_column].iloc[8]
                dict["起止层"] = datas_exls[first_column].iloc[9]
                dict["盲层"] = datas_exls[first_column].iloc[10]
                dict["初始层"] = dict["起止层"].rsplit('/')[0]
                dict["终止层"] = dict["起止层"].rsplit('/')[1]

                async def main():
                    global numtest1
                    path = r"./直梯/64导轨（每台一份）.xlsx"
                    save_path = "直梯.xlsx"
                    wb = openpyxl.load_workbook(path)
                    wb2 = openpyxl.load_workbook(save_path)
                    sheetnames = wb.sheetnames

                    tasks = []
                    for numtest, sheetname in enumerate(sheetnames, start=1):
                        sheet = wb[sheetname]
                        sheet2 = wb2.create_sheet(f"第{numtest1}页")
                        print(numtest1)
                        numtest1 += 1
                        task = asyncio.create_task(copy_sheet(sheet, sheet2))
                        tasks.append(task)

                    # 等待所有任务完成
                    await asyncio.gather(*tasks)

                    # 保存并关闭工作簿
                    wb2.save(save_path)
                    wb.close()
                    wb2.close()
                    workbook = openpyxl.load_workbook('直梯.xlsx')
                    sheet = workbook[f'第{numtest1 - 1}页']
                    cell_g4 = sheet['BF4']
                    cell_g4.value = num
                    cell_g4 = sheet['L6']
                    cell_g4.value = dict["子分部工程名称"]
                    cell_g4 = sheet['AG6']
                    cell_g4.value = dict["子分部工程名称"]
                    cell_g4 = sheet['L7']
                    cell_g4.value = dict["总包单位"]
                    cell_g4 = sheet['AG7']
                    cell_g4.value = dict["总包单位项目负责人"]
                    cell_g4 = sheet['AZ7']

                    cell_g4 = sheet['L8']
                    cell_g4.value = dict["分包单位"]
                    cell_g4 = sheet['AG8']
                    cell_g4.value = dict["分包单位项目负责人"]
                    cell_g4 = sheet['L9']

                    cell_g4 = sheet['AR9']

                    cell_g4 = sheet['AZ26']
                    cell_g4.value = dict["建设单位项目负责人"]
                    workbook.save('直梯.xlsx')
                    workbook.close()

                # 运行异步主程序
                asyncio.run(main())
        except Exception as e:
            print(numtest1, e)

        try:
            for num in range(0, dict['电梯数量']):
                first_column = datas_exls.columns[num]
                date_value = datas_exls[first_column].iloc[45]
                if date_value is not None:
                    # 将日期格式化为字符串，例如"2023-01-13"
                    date_str = date_value.strftime("%Y-%m-%d")
                    # 将格式化后的日期字符串存储到字典中
                    dict["电梯噪声测试日期"] = date_str
                else:
                    dict["电梯噪声测试日期"] = ""
                dict["检验区域"] = datas_exls[first_column].iloc[60]
                dict["电梯型号规格"] = datas_exls[first_column].iloc[29]
                dict["设备梯号"] = datas_exls[first_column].iloc[0]
                dict["层站门"] = datas_exls[first_column].iloc[8]
                dict["起止层"] = datas_exls[first_column].iloc[9]
                dict["盲层"] = datas_exls[first_column].iloc[10]
                dict["初始层"] = dict["起止层"].rsplit('/')[0]
                dict["终止层"] = dict["起止层"].rsplit('/')[1]

                async def main():
                    global numtest1
                    path = r"./直梯/65门系统（每台一份）.xlsx"
                    save_path = "直梯.xlsx"
                    wb = openpyxl.load_workbook(path)
                    wb2 = openpyxl.load_workbook(save_path)
                    sheetnames = wb.sheetnames

                    tasks = []
                    for numtest, sheetname in enumerate(sheetnames, start=1):
                        sheet = wb[sheetname]
                        sheet2 = wb2.create_sheet(f"第{numtest1}页")
                        print(numtest1)
                        numtest1 += 1
                        task = asyncio.create_task(copy_sheet(sheet, sheet2))
                        tasks.append(task)

                    # 等待所有任务完成
                    await asyncio.gather(*tasks)

                    # 保存并关闭工作簿
                    wb2.save(save_path)
                    wb.close()
                    wb2.close()
                    workbook = openpyxl.load_workbook('直梯.xlsx')
                    sheet = workbook[f'第{numtest1 - 1}页']
                    cell_g4 = sheet['BF4']
                    cell_g4.value = num
                    cell_g4 = sheet['L6']
                    cell_g4.value = dict["子分部工程名称"]
                    cell_g4 = sheet['AG6']
                    cell_g4.value = dict["子分部工程名称"]
                    cell_g4 = sheet['L7']
                    cell_g4.value = dict["总包单位"]
                    cell_g4 = sheet['AG7']
                    cell_g4.value = dict["总包单位项目负责人"]
                    cell_g4 = sheet['AZ7']

                    cell_g4 = sheet['L8']
                    cell_g4.value = dict["分包单位"]
                    cell_g4 = sheet['AG8']
                    cell_g4.value = dict["分包单位项目负责人"]
                    cell_g4 = sheet['L9']

                    cell_g4 = sheet['AR9']

                    cell_g4 = sheet['AZ26']
                    cell_g4.value = dict["建设单位项目负责人"]
                    workbook.save('直梯.xlsx')
                    workbook.close()

                # 运行异步主程序
                asyncio.run(main())
        except Exception as e:
            print(numtest1, e)

        try:
            for num in range(0, dict['电梯数量']):
                first_column = datas_exls.columns[num]
                date_value = datas_exls[first_column].iloc[45]
                if date_value is not None:
                    # 将日期格式化为字符串，例如"2023-01-13"
                    date_str = date_value.strftime("%Y-%m-%d")
                    # 将格式化后的日期字符串存储到字典中
                    dict["电梯噪声测试日期"] = date_str
                else:
                    dict["电梯噪声测试日期"] = ""
                dict["检验区域"] = datas_exls[first_column].iloc[60]
                dict["电梯型号规格"] = datas_exls[first_column].iloc[29]
                dict["设备梯号"] = datas_exls[first_column].iloc[0]
                dict["层站门"] = datas_exls[first_column].iloc[8]
                dict["起止层"] = datas_exls[first_column].iloc[9]
                dict["盲层"] = datas_exls[first_column].iloc[10]
                dict["初始层"] = dict["起止层"].rsplit('/')[0]
                dict["终止层"] = dict["起止层"].rsplit('/')[1]

                async def main():
                    global numtest1
                    path = r"./直梯/66轿厢（每台一份）.xlsx"
                    save_path = "直梯.xlsx"
                    wb = openpyxl.load_workbook(path)
                    wb2 = openpyxl.load_workbook(save_path)
                    sheetnames = wb.sheetnames

                    tasks = []
                    for numtest, sheetname in enumerate(sheetnames, start=1):
                        sheet = wb[sheetname]
                        sheet2 = wb2.create_sheet(f"第{numtest1}页")
                        print(numtest1)
                        numtest1 += 1
                        task = asyncio.create_task(copy_sheet(sheet, sheet2))
                        tasks.append(task)

                    # 等待所有任务完成
                    await asyncio.gather(*tasks)

                    # 保存并关闭工作簿
                    wb2.save(save_path)
                    wb.close()
                    wb2.close()
                    workbook = openpyxl.load_workbook('直梯.xlsx')
                    sheet = workbook[f'第{numtest1 - 1}页']
                    cell_g4 = sheet['BF4']
                    cell_g4.value = num
                    cell_g4 = sheet['L6']
                    cell_g4.value = dict["子分部工程名称"]
                    cell_g4 = sheet['AG6']
                    cell_g4.value = dict["子分部工程名称"]
                    cell_g4 = sheet['L7']
                    cell_g4.value = dict["总包单位"]
                    cell_g4 = sheet['AG7']
                    cell_g4.value = dict["总包单位项目负责人"]
                    cell_g4 = sheet['AZ7']

                    cell_g4 = sheet['L8']
                    cell_g4.value = dict["分包单位"]
                    cell_g4 = sheet['AG8']
                    cell_g4.value = dict["分包单位项目负责人"]
                    cell_g4 = sheet['L9']

                    cell_g4 = sheet['AR9']

                    cell_g4 = sheet['AZ22']
                    cell_g4.value = dict["建设单位项目负责人"]
                    workbook.save('直梯.xlsx')
                    workbook.close()

                # 运行异步主程序
                asyncio.run(main())
        except Exception as e:
            print(numtest1, e)

        try:
            for num in range(0, dict['电梯数量']):
                first_column = datas_exls.columns[num]
                date_value = datas_exls[first_column].iloc[45]
                if date_value is not None:
                    # 将日期格式化为字符串，例如"2023-01-13"
                    date_str = date_value.strftime("%Y-%m-%d")
                    # 将格式化后的日期字符串存储到字典中
                    dict["电梯噪声测试日期"] = date_str
                else:
                    dict["电梯噪声测试日期"] = ""
                dict["检验区域"] = datas_exls[first_column].iloc[60]
                dict["电梯型号规格"] = datas_exls[first_column].iloc[29]
                dict["设备梯号"] = datas_exls[first_column].iloc[0]
                dict["缓冲器形式（耗能/蓄能）"] = datas_exls[first_column].iloc[18]
                dict["层站门"] = datas_exls[first_column].iloc[8]
                dict["起止层"] = datas_exls[first_column].iloc[9]
                dict["盲层"] = datas_exls[first_column].iloc[10]
                dict["初始层"] = dict["起止层"].rsplit('/')[0]
                dict["终止层"] = dict["起止层"].rsplit('/')[1]

                async def main():
                    global numtest1
                    path = r"./直梯/67安全部件（每台一份）.xlsx"
                    save_path = "直梯.xlsx"
                    wb = openpyxl.load_workbook(path)
                    wb2 = openpyxl.load_workbook(save_path)
                    sheetnames = wb.sheetnames

                    tasks = []
                    for numtest, sheetname in enumerate(sheetnames, start=1):
                        sheet = wb[sheetname]
                        sheet2 = wb2.create_sheet(f"第{numtest1}页")
                        print(numtest1)
                        numtest1 += 1
                        task = asyncio.create_task(copy_sheet(sheet, sheet2))
                        tasks.append(task)

                    # 等待所有任务完成
                    await asyncio.gather(*tasks)

                    # 保存并关闭工作簿
                    wb2.save(save_path)
                    wb.close()
                    wb2.close()
                    workbook = openpyxl.load_workbook('直梯.xlsx')
                    sheet = workbook[f'第{numtest1 - 1}页']
                    cell_g4 = sheet['BF4']
                    cell_g4.value = num
                    cell_g4 = sheet['L6']
                    cell_g4.value = dict["子分部工程名称"]
                    cell_g4 = sheet['AG6']
                    cell_g4.value = dict["子分部工程名称"]
                    cell_g4 = sheet['L7']
                    cell_g4.value = dict["总包单位"]
                    cell_g4 = sheet['AG7']
                    cell_g4.value = dict["总包单位项目负责人"]
                    cell_g4 = sheet['AZ7']

                    cell_g4 = sheet['L8']
                    cell_g4.value = dict["分包单位"]
                    cell_g4 = sheet['AG8']
                    cell_g4.value = dict["分包单位项目负责人"]
                    cell_g4 = sheet['L9']

                    cell_g4 = sheet['AR9']
                    cell_g4 = sheet['BB27']
                    cell_g4.value = "合格" if "液压" in dict["缓冲器形式（耗能/蓄能）"] else "/"

                    cell_g4 = sheet['AZ23']
                    cell_g4.value = dict["建设单位项目负责人"]
                    workbook.save('直梯.xlsx')
                    workbook.close()

                # 运行异步主程序
                asyncio.run(main())
        except Exception as e:
            print(numtest1, e)

        try:
            for num in range(0, dict['电梯数量']):
                first_column = datas_exls.columns[num]
                date_value = datas_exls[first_column].iloc[45]
                if date_value is not None:
                    # 将日期格式化为字符串，例如"2023-01-13"
                    date_str = date_value.strftime("%Y-%m-%d")
                    # 将格式化后的日期字符串存储到字典中
                    dict["电梯噪声测试日期"] = date_str
                else:
                    dict["电梯噪声测试日期"] = ""
                dict["检验区域"] = datas_exls[first_column].iloc[60]
                dict["电梯型号规格"] = datas_exls[first_column].iloc[29]
                dict["设备梯号"] = datas_exls[first_column].iloc[0]
                dict["补偿链配备（有/无））"] = datas_exls[first_column].iloc[19]
                dict["层站门"] = datas_exls[first_column].iloc[8]
                dict["起止层"] = datas_exls[first_column].iloc[9]
                dict["盲层"] = datas_exls[first_column].iloc[10]
                dict["初始层"] = dict["起止层"].rsplit('/')[0]
                dict["终止层"] = dict["起止层"].rsplit('/')[1]

                async def main():
                    global numtest1
                    path = r"./直梯/68悬挂装置、随行电缆、补偿装置（每台一份）.xlsx"
                    save_path = "直梯.xlsx"
                    wb = openpyxl.load_workbook(path)
                    wb2 = openpyxl.load_workbook(save_path)
                    sheetnames = wb.sheetnames

                    tasks = []
                    for numtest, sheetname in enumerate(sheetnames, start=1):
                        sheet = wb[sheetname]
                        sheet2 = wb2.create_sheet(f"第{numtest1}页")
                        print(numtest1)
                        numtest1 += 1
                        task = asyncio.create_task(copy_sheet(sheet, sheet2))
                        tasks.append(task)

                    # 等待所有任务完成
                    await asyncio.gather(*tasks)

                    # 保存并关闭工作簿
                    wb2.save(save_path)
                    wb.close()
                    wb2.close()
                    workbook = openpyxl.load_workbook('直梯.xlsx')
                    sheet = workbook[f'第{numtest1 - 1}页']
                    cell_g4 = sheet['BF4']
                    cell_g4.value = num
                    cell_g4 = sheet['L6']
                    cell_g4.value = dict["子分部工程名称"]
                    cell_g4 = sheet['AG6']
                    cell_g4.value = dict["子分部工程名称"]
                    cell_g4 = sheet['L7']
                    cell_g4.value = dict["总包单位"]
                    cell_g4 = sheet['AG7']
                    cell_g4.value = dict["总包单位项目负责人"]
                    cell_g4 = sheet['AZ7']

                    cell_g4 = sheet['L8']
                    cell_g4.value = dict["分包单位"]
                    cell_g4 = sheet['AG8']
                    cell_g4.value = dict["分包单位项目负责人"]
                    cell_g4 = sheet['L9']

                    cell_g4 = sheet['BB27']
                    cell_g4.value = "合格" if "合格" in dict["补偿链配备（有/无））"] else "/"

                    cell_g4 = sheet['AZ25']
                    cell_g4.value = dict["建设单位项目负责人"]
                    workbook.save('直梯.xlsx')
                    workbook.close()

                # 运行异步主程序
                asyncio.run(main())
        except Exception as e:
            print(numtest1, e)

        try:
            for num in range(0, dict['电梯数量']):
                first_column = datas_exls.columns[num]
                date_value = datas_exls[first_column].iloc[45]
                if date_value is not None:
                    # 将日期格式化为字符串，例如"2023-01-13"
                    date_str = date_value.strftime("%Y-%m-%d")
                    # 将格式化后的日期字符串存储到字典中
                    dict["电梯噪声测试日期"] = date_str
                else:
                    dict["电梯噪声测试日期"] = ""
                dict["检验区域"] = datas_exls[first_column].iloc[60]
                dict["电梯型号规格"] = datas_exls[first_column].iloc[29]
                dict["设备梯号"] = datas_exls[first_column].iloc[0]
                dict["层站门"] = datas_exls[first_column].iloc[8]
                dict["起止层"] = datas_exls[first_column].iloc[9]
                dict["盲层"] = datas_exls[first_column].iloc[10]
                dict["初始层"] = dict["起止层"].rsplit('/')[0]
                dict["终止层"] = dict["起止层"].rsplit('/')[1]

                async def main():
                    global numtest1
                    path = r"./直梯/69电气装置（每台一份）.xlsx"
                    save_path = "直梯.xlsx"
                    wb = openpyxl.load_workbook(path)
                    wb2 = openpyxl.load_workbook(save_path)
                    sheetnames = wb.sheetnames

                    tasks = []
                    for numtest, sheetname in enumerate(sheetnames, start=1):
                        sheet = wb[sheetname]
                        sheet2 = wb2.create_sheet(f"第{numtest1}页")
                        print(numtest1)
                        numtest1 += 1
                        task = asyncio.create_task(copy_sheet(sheet, sheet2))
                        tasks.append(task)

                    # 等待所有任务完成
                    await asyncio.gather(*tasks)

                    # 保存并关闭工作簿
                    wb2.save(save_path)
                    wb.close()
                    wb2.close()
                    workbook = openpyxl.load_workbook('直梯.xlsx')
                    sheet = workbook[f'第{numtest1 - 1}页']
                    cell_g4 = sheet['BF4']
                    cell_g4.value = num
                    cell_g4 = sheet['L6']
                    cell_g4.value = dict["子分部工程名称"]
                    cell_g4 = sheet['AG6']
                    cell_g4.value = dict["子分部工程名称"]
                    cell_g4 = sheet['L7']
                    cell_g4.value = dict["总包单位"]
                    cell_g4 = sheet['AG7']
                    cell_g4.value = dict["总包单位项目负责人"]
                    cell_g4 = sheet['AZ7']

                    cell_g4 = sheet['L8']
                    cell_g4.value = dict["分包单位"]
                    cell_g4 = sheet['AG8']
                    cell_g4.value = dict["分包单位项目负责人"]
                    cell_g4 = sheet['L9']

                    cell_g4 = sheet['AR9']

                    cell_g4 = sheet['AZ24']
                    cell_g4.value = dict["建设单位项目负责人"]
                    workbook.save('直梯.xlsx')
                    workbook.close()

                # 运行异步主程序
                asyncio.run(main())
        except Exception as e:
            print(numtest1, e)

        try:
            for num in range(0, dict['电梯数量']):
                first_column = datas_exls.columns[num]
                date_value = datas_exls[first_column].iloc[45]
                if date_value is not None:
                    # 将日期格式化为字符串，例如"2023-01-13"
                    date_str = date_value.strftime("%Y-%m-%d")
                    # 将格式化后的日期字符串存储到字典中
                    dict["电梯噪声测试日期"] = date_str
                else:
                    dict["电梯噪声测试日期"] = ""
                dict["检验区域"] = datas_exls[first_column].iloc[60]
                dict["电梯型号规格"] = datas_exls[first_column].iloc[29]
                dict["设备梯号"] = datas_exls[first_column].iloc[0]
                dict["层站门"] = datas_exls[first_column].iloc[8]
                dict["起止层"] = datas_exls[first_column].iloc[9]
                dict["盲层"] = datas_exls[first_column].iloc[10]
                dict["初始层"] = dict["起止层"].rsplit('/')[0]
                dict["终止层"] = dict["起止层"].rsplit('/')[1]

                async def main():
                    global numtest1
                    path = r"./直梯/70整机安装验收（每台一份）.xlsx"
                    save_path = "直梯.xlsx"
                    wb = openpyxl.load_workbook(path)
                    wb2 = openpyxl.load_workbook(save_path)
                    sheetnames = wb.sheetnames

                    tasks = []
                    for numtest, sheetname in enumerate(sheetnames, start=1):
                        sheet = wb[sheetname]
                        sheet2 = wb2.create_sheet(f"第{numtest1}页")
                        print(numtest1)
                        numtest1 += 1
                        task = asyncio.create_task(copy_sheet(sheet, sheet2))
                        tasks.append(task)

                    # 等待所有任务完成
                    await asyncio.gather(*tasks)

                    # 保存并关闭工作簿
                    wb2.save(save_path)
                    wb.close()
                    wb2.close()
                    workbook = openpyxl.load_workbook('直梯.xlsx')
                    sheet = workbook[f'第{numtest1 - 1}页']
                    cell_g4 = sheet['BF4']
                    cell_g4.value = num
                    cell_g4 = sheet['L6']
                    cell_g4.value = dict["子分部工程名称"]
                    cell_g4 = sheet['AG6']
                    cell_g4.value = dict["子分部工程名称"]
                    cell_g4 = sheet['L7']
                    cell_g4.value = dict["总包单位"]
                    cell_g4 = sheet['AG7']
                    cell_g4.value = dict["总包单位项目负责人"]
                    cell_g4 = sheet['AZ7']

                    cell_g4 = sheet['L8']
                    cell_g4.value = dict["分包单位"]
                    cell_g4 = sheet['AG8']
                    cell_g4.value = dict["分包单位项目负责人"]
                    cell_g4 = sheet['L9']

                    cell_g4 = sheet['AR9']

                    cell_g4 = sheet['AZ27']
                    cell_g4.value = dict["建设单位项目负责人"]
                    workbook.save('直梯.xlsx')
                    workbook.close()

                # 运行异步主程序
                asyncio.run(main())
        except Exception as e:
            print(numtest1, e)

        try:
            async def main():
                global numtest1
                path = r"./直梯/71-表B.0.10 单位工程竣工验收报审表（后附检验报告）.xlsx"
                save_path = "直梯.xlsx"
                wb = openpyxl.load_workbook(path)
                wb2 = openpyxl.load_workbook(save_path)
                sheetnames = wb.sheetnames

                tasks = []
                for numtest, sheetname in enumerate(sheetnames, start=1):
                    sheet = wb[sheetname]
                    sheet2 = wb2.create_sheet(f"第{numtest1}页")
                    print(numtest1)
                    numtest1 += 1
                    task = asyncio.create_task(copy_sheet(sheet, sheet2))
                    tasks.append(task)

                # 等待所有任务完成
                await asyncio.gather(*tasks)

                # 保存并关闭工作簿
                wb2.save(save_path)
                wb.close()
                wb2.close()
                workbook = openpyxl.load_workbook('直梯.xlsx')
                sheet = workbook[f'第{numtest1 - 1}页']

                cell_g4 = sheet['E3']
                cell_g4.value = dict["工程名称"]
                cell_g4 = sheet['D5']
                cell_g4.value = dict["监理单位"]
                cell_g4 = sheet['G6']
                cell_g4.value = dict["工程名称"]
                cell_g4 = sheet['K27']
                cell_g4.value = dict["监理单位"]
                workbook.save('直梯.xlsx')
                workbook.close()
                folder_path = r'.\检验报告'
                workbook = openpyxl.load_workbook('直梯.xlsx')
                # 遍历文件夹中的所有文件
                for filename in os.listdir(folder_path):
                    if filename.endswith('.jpg') or filename.endswith('.png'):
                        file_path = os.path.join(folder_path, filename)
                        # 创建一个新的工作表，并使用图片名称作为工作表名称
                        sheet = workbook.create_sheet(f"第{numtest1}页")
                        # 在工作表中插入图片
                        img = Image(file_path)
                        sheet.add_image(img, 'A1')
                        numtest1 = numtest1 + 1

                # 保存 Excel 文件
                workbook.save('直梯.xlsx')
                workbook.close()

            # 运行异步主程序
            asyncio.run(main())
        except Exception as e:
            print(numtest1, e)

        messagebox.showinfo("保存完成", "信息已成功保存。")


    def getdata(self, labels):
        for i in labels:
            if i in dict.keys():
                print()


if __name__ == '__main__':
    root = Tk()
    app = InformationCompletion(root)
    root.mainloop()
